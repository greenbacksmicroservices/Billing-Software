from django import forms
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import View, TemplateView, ListView, CreateView, UpdateView, DetailView, DeleteView
from django.contrib.auth.views import LoginView
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.forms import AuthenticationForm, PasswordChangeForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.urls import reverse_lazy
from django.http import JsonResponse, Http404, HttpResponse, FileResponse
from django.core.serializers.json import DjangoJSONEncoder
from django.db import transaction
from django.db.models import Sum, Count, Q, F
from django.utils.decorators import method_decorator
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import PermissionDenied

import json
import os
import mimetypes
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, date, timedelta

from .models import (
    Company, CustomUser, SubscriptionPlan, Customer, Supplier, Product, Category, Brand, Unit,
    HSNSACMaster, Warehouse, StockMovement, Invoice, InvoiceItem, Quotation, QuotationItem,
    SalesOrder, SalesOrderItem, PurchaseOrder, PurchaseOrderItem, PurchaseBill, PurchaseBillItem, PurchaseBillDocument, CreditNote, CreditNoteItem,
    DebitNote, DebitNoteItem, Payment, Expense, ExpenseCategory, AuditLog, Notification,
    SupportTicket, Announcement, GSTTransaction, CustomerLedger, SupplierLedger, PasswordResetOTP,
    GSTApplication
)
from .forms import CompanyForm, ProductForm, CustomerForm, SupplierForm, WarehouseForm, PlanForm, ExpenseForm
from .mixins import CompanyRequiredMixin, RoleRequiredMixin, CompanyQuerySetMixin, PaginationMixin, AjaxFormMixin
from .services import QuotationService, PurchaseOrderService
from .utils import (
    quantize_amount, calculate_item_gst, calculate_gst, recalculate_invoice_totals,
    recalculate_purchase_totals, recalculate_quotation_totals, recalculate_sales_order_totals,
    recalculate_purchase_order_totals, recalculate_credit_note_totals, recalculate_debit_note_totals,
    recalculate_proforma_totals, update_product_stock, generate_upi_qr_string, log_action,
    record_invoice_accounting, cancel_invoice_accounting, record_purchase_accounting,
    cancel_purchase_accounting, record_payment_accounting, record_credit_note_accounting,
    record_debit_note_accounting, parse_money, format_money, serialize_decimal, build_products_json
)

# --- AUTHENTICATION FLOWS ---

class UnifiedLoginView(View):
    def get(self, request):
        if request.user.is_authenticated:
            if request.user.role == 'SUPERADMIN' or request.user.is_superuser:
                return redirect('admin_dashboard')
            return redirect('company_dashboard')
        form = AuthenticationForm()
        return render(request, 'auth/login.html', {'form': form})

    def post(self, request):
        username_or_email = request.POST.get('username')
        password = request.POST.get('password')
        
        # Support email login by mapping to username
        username = username_or_email
        if '@' in username_or_email:
            user_by_email = CustomUser.objects.filter(email=username_or_email).first()
            if user_by_email:
                username = user_by_email.username
                
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.is_active:
                if user.role == 'SUPERADMIN' or user.is_superuser:
                    login(request, user)
                    log_action(user, 'LOGIN', 'AUTHENTICATION', request=request)
                    return redirect('admin_dashboard')
                else:
                    if user.company:
                        if user.company.is_active:
                            login(request, user)
                            log_action(user, 'LOGIN', 'AUTHENTICATION', request=request)
                            return redirect('company_dashboard')
                        else:
                            messages.error(request, "Your company account is inactive. Please contact support.")
                    else:
                        messages.error(request, "Your user profile does not belong to any business account.")
            else:
                messages.error(request, "This account has been disabled. Please contact support.")
        else:
            messages.error(request, "Invalid email/username or password.")
            
        form = AuthenticationForm(request, data=request.POST)
        return render(request, 'auth/login.html', {'form': form})


class AdminLoginView(UnifiedLoginView):
    pass


class CompanyLoginView(UnifiedLoginView):
    pass


def logout_view(request):
    user = request.user
    if user.is_authenticated:
        log_action(user, 'LOGOUT', 'AUTHENTICATION', request=request)
    logout(request)
    return redirect('login')


class UserProfileView(LoginRequiredMixin, View):
    def get(self, request):
        return render(request, 'auth/profile.html')

    def post(self, request):
        user = request.user
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        mobile = request.POST.get('mobile')
        
        if not first_name or not email:
            messages.error(request, "First Name and Email are required.")
            return render(request, 'auth/profile.html')
            
        user.first_name = first_name
        user.last_name = last_name or ''
        user.email = email
        user.mobile = mobile or ''
        user.save()
        
        messages.success(request, "Your profile has been updated successfully!")
        return redirect('profile')


class UserChangePasswordView(LoginRequiredMixin, View):
    def get(self, request):
        form = PasswordChangeForm(request.user)
        return render(request, 'auth/change_password.html', {'form': form})

    def post(self, request):
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, "Your password was successfully updated!")
            return redirect('profile')
        else:
            messages.error(request, "Please correct the error below.")
        return render(request, 'auth/change_password.html', {'form': form})


# --- PROFILE PHOTO & GENERIC DELETE API VIEWS ---
def api_profile_photo_upload(request):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)

    if request.method != 'POST' or 'photo' not in request.FILES:
        return JsonResponse({'status': 'error', 'message': 'Please upload an image file.'}, status=400)

    photo_file = request.FILES['photo']

    if photo_file.size > 5 * 1024 * 1024:
        return JsonResponse({'status': 'error', 'message': 'Profile image must be less than 5MB.'}, status=400)

    ext = os.path.splitext(photo_file.name)[1].lower()
    if ext not in ['.jpg', '.jpeg', '.png', '.webp']:
        return JsonResponse({'status': 'error', 'message': 'Please select a valid image file (JPG, JPEG, PNG, or WEBP).'}, status=400)

    try:
        user = request.user
        
        # 1. Update user profile photo
        if user.profile_photo:
            try:
                if user.profile_photo.path and os.path.isfile(user.profile_photo.path):
                    os.remove(user.profile_photo.path)
            except Exception:
                pass
            try:
                user.profile_photo.delete(save=False)
            except Exception:
                pass

        user.profile_photo = photo_file
        user.save()
        log_action(user, 'UPDATE_PROFILE_PHOTO', 'USER', user.id, request=request)

        # 2. Update company logo if user is linked to a company
        if user.company:
            if user.company.logo:
                try:
                    if user.company.logo.path and os.path.isfile(user.company.logo.path):
                        os.remove(user.company.logo.path)
                except Exception:
                    pass
                try:
                    user.company.logo.delete(save=False)
                except Exception:
                    pass
            user.company.logo = photo_file
            user.company.save()
            log_action(user, 'UPDATE_SETTINGS', 'COMPANY', user.company.id, request=request)

        return JsonResponse({
            'status': 'success',
            'message': 'Profile photo uploaded successfully.',
            'photo_url': user.get_profile_photo_url(),
            'profile_photo_name': user.get_profile_photo_filename()
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': 'Profile photo upload failed. Please try again.'}, status=400)


def api_profile_photo_remove(request):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST method required.'}, status=405)

    try:
        user = request.user
        
        # 1. Clear user profile photo
        if user.profile_photo:
            try:
                if user.profile_photo.path and os.path.isfile(user.profile_photo.path):
                    os.remove(user.profile_photo.path)
            except Exception:
                pass
            try:
                user.profile_photo.delete(save=False)
            except Exception:
                pass
            user.profile_photo = None
            user.save()
            log_action(user, 'REMOVE_PROFILE_PHOTO', 'USER', user.id, request=request)

        # 2. Clear company logo if user is linked to a company
        if user.company:
            if user.company.logo:
                try:
                    if user.company.logo.path and os.path.isfile(user.company.logo.path):
                        os.remove(user.company.logo.path)
                except Exception:
                    pass
                try:
                    user.company.logo.delete(save=False)
                except Exception:
                    pass
            user.company.logo = None
            user.company.save()
            log_action(user, 'UPDATE_SETTINGS', 'COMPANY', user.company.id, request=request)

        return JsonResponse({
            'status': 'success',
            'message': 'Profile photo removed successfully.',
            'photo_url': None,
            'profile_photo_name': 'No profile photo uploaded'
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': 'Failed to remove profile photo. Please try again.'}, status=400)


@transaction.atomic
def admin_user_change_password(request, pk):
    if not request.user.is_authenticated or request.user.role != 'SUPERADMIN':
        return JsonResponse({'status': 'error', 'message': 'Permission denied.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST method required.'}, status=405)

    try:
        data = json.loads(request.body) if request.body and request.content_type == 'application/json' else request.POST
        new_password = data.get('new_password', '').strip()
        confirm_password = data.get('confirm_password', '').strip()

        if not new_password or not confirm_password:
            return JsonResponse({'status': 'error', 'message': 'New password and Confirm password are required.'}, status=400)

        if new_password != confirm_password:
            return JsonResponse({'status': 'error', 'message': 'Passwords do not match.'}, status=400)

        if len(new_password) < 6:
            return JsonResponse({'status': 'error', 'message': 'Password must be at least 6 characters long.'}, status=400)

        target_user = get_object_or_404(CustomUser, pk=pk)
        target_user.set_password(new_password)
        target_user.save()

        log_action(request.user, 'CHANGE_USER_PASSWORD', 'USER', target_user.id, new_values={'user': target_user.username}, request=request)
        return JsonResponse({'status': 'success', 'message': f"Password for user '{target_user.username}' changed successfully."})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@transaction.atomic
def api_generic_delete(request, model_name, pk):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST method required.'}, status=405)

    user = request.user
    company = getattr(user, 'company', None)
    is_superadmin = (user.role == 'SUPERADMIN')

    ADMIN_MODELS = {
        'company': Company,
        'plan': SubscriptionPlan,
        'user': CustomUser,
        'ticket': SupportTicket,
        'audit_log': AuditLog,
        'hsn_sac': HSNSACMaster,
    }

    COMPANY_MODELS = {
        'product': Product,
        'customer': Customer,
        'supplier': Supplier,
        'warehouse': Warehouse,
        'quotation': Quotation,
        'sales_order': SalesOrder,
        'invoice': Invoice,
        'purchase_bill': PurchaseBill,
        'payment': Payment,
        'expense': Expense,
        'credit_note': CreditNote,
        'debit_note': DebitNote,
    }

    model_key = model_name.lower().replace('-', '_')

    try:
        if model_key in ADMIN_MODELS:
            if model_key == 'user' and user.role == 'ADMIN':
                model_cls = ADMIN_MODELS[model_key]
                obj = get_object_or_404(model_cls, pk=pk, company=company)
            else:
                if not is_superadmin:
                    return JsonResponse({'status': 'error', 'message': 'Only System Admins can perform this action.'}, status=403)
                model_cls = ADMIN_MODELS[model_key]
                obj = get_object_or_404(model_cls, pk=pk)
        elif model_key in COMPANY_MODELS:
            if not company and not is_superadmin:
                return JsonResponse({'status': 'error', 'message': 'Company context required.'}, status=403)
            model_cls = COMPANY_MODELS[model_key]
            if is_superadmin:
                obj = get_object_or_404(model_cls, pk=pk)
            else:
                obj = get_object_or_404(model_cls, pk=pk, company=company)
        else:
            return JsonResponse({'status': 'error', 'message': f"Unknown model type '{model_name}'."}, status=400)

        action_performed = 'deleted'
        msg = 'Record deleted successfully.'

        if model_key == 'company':
            if CustomUser.objects.filter(company=obj).exists() or Invoice.objects.filter(company=obj).exists():
                obj.is_active = False
                obj.save()
                action_performed = 'deactivated'
                msg = f"Company '{obj.name}' was deactivated because it is linked to existing users or financial records."
            else:
                obj.delete()
                msg = "Company deleted successfully."

        elif model_key == 'user':
            if obj.id == user.id:
                return JsonResponse({'success': False, 'status': 'error', 'message': 'You cannot delete your own active account.'}, status=400)
            
            # Guard: prevent deleting the last active Company Admin of the company
            if not is_superadmin and obj.role == 'ADMIN' and obj.company:
                admin_count = CustomUser.objects.filter(company=obj.company, role='ADMIN', is_active=True).count()
                if admin_count <= 1 and obj.is_active:
                    return JsonResponse({'success': False, 'status': 'error', 'message': 'You cannot delete or deactivate the last active Company Admin of this business.'}, status=400)
            
            # Soft-delete/deactivate accounts to protect audit logs
            obj.is_active = False
            obj.save()
            action_performed = 'deactivated'
            msg = f"User '{obj.username}' was deactivated because it is linked to existing users or financial records."

        elif model_key == 'product':
            is_ref = (
                StockMovement.objects.filter(product=obj).exists() or
                InvoiceItem.objects.filter(product=obj).exists() or
                QuotationItem.objects.filter(product=obj).exists() or
                SalesOrderItem.objects.filter(product=obj).exists() or
                PurchaseBillItem.objects.filter(product=obj).exists()
            )
            if is_ref:
                obj.is_active = False
                obj.save()
                action_performed = 'deactivated'
                msg = f"Product '{obj.name}' was deactivated because it is linked to existing users or financial records."
            else:
                obj.delete()
                msg = "Product deleted successfully."

        elif model_key == 'customer':
            is_ref = (
                Invoice.objects.filter(customer=obj).exists() or
                Payment.objects.filter(customer=obj).exists() or
                Quotation.objects.filter(customer=obj).exists() or
                CustomerLedger.objects.filter(customer=obj).exists()
            )
            if is_ref:
                obj.is_active = False
                obj.save()
                action_performed = 'deactivated'
                msg = f"Customer '{obj.name}' was deactivated because it is linked to existing users or financial records."
            else:
                obj.delete()
                msg = "Customer deleted successfully."

        elif model_key == 'supplier':
            is_ref = (
                PurchaseBill.objects.filter(supplier=obj).exists() or
                Payment.objects.filter(supplier=obj).exists() or
                SupplierLedger.objects.filter(supplier=obj).exists()
            )
            if is_ref:
                obj.is_active = False
                obj.save()
                action_performed = 'deactivated'
                msg = f"Supplier '{obj.name}' was deactivated because it is linked to existing users or financial records."
            else:
                obj.delete()
                msg = "Supplier deleted successfully."

        elif model_key == 'invoice':
            if obj.status in ['POSTED', 'PAID', 'PARTIALLY_PAID', 'APPROVED']:
                # Reverse stock adjustments
                movements = StockMovement.objects.filter(company=company, reference_id=obj.id, movement_type='SALE')
                for mv in movements:
                    # Create counter-balancing movement
                    StockMovement.objects.create(
                        company=company, product=mv.product, warehouse=mv.warehouse,
                        quantity=abs(mv.quantity), movement_type='SALES_RETURN', reference_id=obj.id,
                        reference_no=f"CNL: {obj.invoice_number}", created_by=user
                    )
                    update_product_stock(mv.product.id)
                # Rebalance customer receivable
                customer = obj.customer
                customer.outstanding_balance -= obj.grand_total
                customer.save()
                cancel_invoice_accounting(obj)
                obj.status = 'CANCELLED'
                obj.save()
                action_performed = 'deactivated'
                msg = f"Invoice {obj.invoice_number} cancelled and customer balance restored."
            else:
                obj.delete()
                msg = "Invoice deleted successfully."

        elif model_key == 'purchase_bill':
            if obj.status in ['POSTED', 'PAID', 'PARTIALLY_PAID']:
                # Reverse inventory
                movements = StockMovement.objects.filter(company=company, reference_id=obj.id, movement_type='PURCHASE')
                for mv in movements:
                    StockMovement.objects.create(
                        company=company, product=mv.product, warehouse=mv.warehouse,
                        quantity=-mv.quantity, movement_type='PURCHASE_RETURN', reference_id=obj.id,
                        reference_no=f"CNL: {obj.supplier_bill_no}", created_by=user
                    )
                    update_product_stock(mv.product.id)
                # Reverse outstanding
                supplier = obj.supplier
                supplier.outstanding_balance -= obj.grand_total
                supplier.save()
                cancel_purchase_accounting(obj)
                obj.status = 'CANCELLED'
                obj.save()
                action_performed = 'deactivated'
                msg = f"Purchase bill {obj.supplier_bill_no} cancelled."
            else:
                obj.delete()
                msg = "Purchase bill deleted successfully."

        elif model_key in ['quotation', 'sales_order']:
            if getattr(obj, 'status', '') in ['POSTED', 'PAID', 'PARTIALLY_PAID', 'APPROVED', 'SENT', 'CONVERTED', 'PENDING']:
                obj.status = 'CANCELLED'
                obj.save()
                action_performed = 'deactivated'
                msg = f"{model_key.replace('_', ' ').capitalize()} record has been CANCELLED."
            else:
                obj.delete()
                msg = f"{model_key.replace('_', ' ').capitalize()} deleted successfully."

        elif model_key == 'hsn_sac':
            is_ref = Product.objects.filter(hsn_sac=obj).exists() or InvoiceItem.objects.filter(hsn_sac_code=obj.code).exists()
            if is_ref:
                obj.is_active = False
                obj.save()
                action_performed = 'deactivated'
                msg = f"HSN/SAC Code '{obj.code}' was deactivated because it is linked to existing users or financial records."
            else:
                obj.delete()
                msg = f"HSN/SAC Code '{obj.code}' deleted successfully."

        elif model_key == 'payment':
            if obj.payment_type == 'RECEIPT':
                customer = obj.customer
                if customer:
                    customer.outstanding_balance += obj.amount
                    customer.save()
                invoice = obj.invoice
                if invoice:
                    invoice.paid_amount -= obj.amount
                    if invoice.paid_amount <= 0:
                        invoice.status = 'POSTED'
                    else:
                        invoice.status = 'PARTIALLY_PAID'
                    invoice.save()
                CustomerLedger.objects.filter(company=company, customer=customer, entry_type='PAYMENT', reference_id=obj.id).delete()
            else:
                supplier = obj.supplier
                if supplier:
                    supplier.outstanding_balance += obj.amount
                    supplier.save()
                bill = obj.purchase_bill
                if bill:
                    bill.paid_amount -= obj.amount
                    if bill.paid_amount <= 0:
                        bill.status = 'POSTED'
                    else:
                        bill.status = 'PARTIALLY_PAID'
                    bill.save()
                SupplierLedger.objects.filter(company=company, supplier=supplier, entry_type='PAYMENT', reference_id=obj.id).delete()
            obj.delete()
            msg = "Payment deleted successfully."

        elif model_key == 'credit_note':
            if obj.status == 'POSTED':
                customer = obj.invoice.customer
                customer.outstanding_balance += obj.grand_total
                customer.save()
                if obj.reason == 'SALES_RETURN':
                    StockMovement.objects.filter(company=company, reference_id=obj.id, movement_type='SALES_RETURN').delete()
                    for item in obj.invoice.items.all():
                        update_product_stock(item.product.id)
                GSTTransaction.objects.filter(company=company, transaction_type='CREDIT_NOTE', reference_id=obj.id).update(is_cancelled=True)
                CustomerLedger.objects.filter(company=company, customer=customer, entry_type='CREDIT_NOTE', reference_id=obj.id).delete()
                obj.status = 'CANCELLED'
                obj.save()
                action_performed = 'deactivated'
                msg = f"Credit note {obj.note_number} cancelled."
            else:
                obj.delete()
                msg = "Credit note deleted successfully."

        elif model_key == 'debit_note':
            if obj.status == 'POSTED':
                supplier = obj.purchase_bill.supplier
                supplier.outstanding_balance += obj.grand_total
                supplier.save()
                if obj.reason == 'PURCHASE_RETURN':
                    StockMovement.objects.filter(company=company, reference_id=obj.id, movement_type='PURCHASE_RETURN').delete()
                    for item in obj.purchase_bill.items.all():
                        update_product_stock(item.product.id)
                GSTTransaction.objects.filter(company=company, transaction_type='DEBIT_NOTE', reference_id=obj.id).update(is_cancelled=True)
                SupplierLedger.objects.filter(company=company, supplier=supplier, entry_type='DEBIT_NOTE', reference_id=obj.id).delete()
                obj.status = 'CANCELLED'
                obj.save()
                action_performed = 'deactivated'
                msg = f"Debit note {obj.note_number} cancelled."
            else:
                obj.delete()
                msg = "Debit note deleted successfully."

        else:
            obj.delete()
            msg = f"{model_key.replace('_', ' ').capitalize()} deleted successfully."

        log_action(user, f'DELETE_{model_key.upper()}', model_key.upper(), pk, request=request)
        return JsonResponse({
            'success': True,
            'status': 'success',
            'action': action_performed,
            'message': msg,
            'id': int(pk)
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'status': 'error',
            'message': str(e)
        }, status=400)


class CompanyDashboardChartDataView(CompanyRequiredMixin, View):
    def get(self, request):
        company = request.user.company
        period = request.GET.get('period', 'this_month')
        
        # Parse timezone-aware local date
        today_val = timezone.localtime(timezone.now()).date()
        if period == 'today':
            start_date = today_val
            end_date = today_val
        elif period == 'this_week':
            start_date = today_val - timedelta(days=today_val.weekday())
            end_date = today_val
        elif period == 'this_month':
            start_date = today_val.replace(day=1)
            end_date = today_val
        elif period == 'last_month':
            last_day_of_last_month = today_val.replace(day=1) - timedelta(days=1)
            start_date = last_day_of_last_month.replace(day=1)
            end_date = last_day_of_last_month
        elif period == 'this_year':
            start_date = today_val.replace(month=1, day=1)
            end_date = today_val
        else: # default to this month
            start_date = today_val.replace(day=1)
            end_date = today_val
            
        # Optional custom date range
        start_param = request.GET.get('start')
        end_param = request.GET.get('end')
        if start_param and end_param:
            try:
                start_date = datetime.strptime(start_param, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_param, '%Y-%m-%d').date()
            except ValueError:
                pass

        # Generate list of dates in the range
        delta = end_date - start_date
        dates_list = [start_date + timedelta(days=i) for i in range(delta.days + 1)]
        labels = [d.strftime('%d %b') for d in dates_list]
        
        # Sales Overview Trend (GRAPH 1) - Include all valid non-draft, non-cancelled invoices
        valid_sales_statuses = ['POSTED', 'PARTIALLY_PAID', 'PAID']
        sales_qs = Invoice.objects.filter(
            company=company,
            invoice_date__range=(start_date, end_date),
            status__in=valid_sales_statuses
        ).values('invoice_date').annotate(total=Sum('grand_total'))
        sales_map = {item['invoice_date']: float(item['total'] or 0.0) for item in sales_qs}
        sales_data = [round(sales_map.get(d, 0.0), 2) for d in dates_list]

        # Purchases Overview Trend (GRAPH 2) - Include all valid non-draft, non-cancelled purchase bills
        valid_purchase_statuses = ['POSTED', 'PARTIALLY_PAID', 'PAID']
        purchases_qs = PurchaseBill.objects.filter(
            company=company,
            bill_date__range=(start_date, end_date),
            status__in=valid_purchase_statuses
        ).values('bill_date').annotate(total=Sum('grand_total'))
        purchases_map = {item['bill_date']: float(item['total'] or 0.0) for item in purchases_qs}
        purchases_data = [round(purchases_map.get(d, 0.0), 2) for d in dates_list]

        # Expenses Category Donut (GRAPH 3)
        expenses_qs = Expense.objects.filter(
            company=company,
            created_at__date__range=(start_date, end_date)
        ).values('category__name').annotate(total=Sum('amount')).order_by('-total')
        
        expense_labels = []
        expense_values = []
        for item in expenses_qs:
            name = item['category__name'] or 'Uncategorized'
            expense_labels.append(name)
            expense_values.append(float(item['total'] or 0.0))
            
        if not expense_labels:
            expense_labels = ['No Expenses']
            expense_values = [0.0]

        # Revenue/Collection/Profit Trend (GRAPH 4)
        collections_qs = Payment.objects.filter(
            company=company,
            payment_type='RECEIPT',
            payment_date__range=(start_date, end_date)
        ).values('payment_date').annotate(total=Sum('amount'))
        collections_map = {item['payment_date']: float(item['total'] or 0.0) for item in collections_qs}
        collections_data = [round(collections_map.get(d, 0.0), 2) for d in dates_list]

        expenses_daily_qs = Expense.objects.filter(
            company=company,
            created_at__date__range=(start_date, end_date)
        ).values('created_at__date').annotate(total=Sum('amount'))
        expenses_map = {item['created_at__date']: float(item['total'] or 0.0) for item in expenses_daily_qs}
        expenses_data = [expenses_map.get(d, 0.0) for d in dates_list]

        profit_data = [round(sales_data[i] - purchases_data[i] - expenses_data[i], 2) for i in range(len(dates_list))]

        # Status-Wise Breakdown Graph (GRAPH 5)
        status_entity = request.GET.get('status_entity', 'invoice')
        status_labels = []
        status_values = []

        if status_entity == 'quotation':
            st_dict = dict(Quotation.STATUS_CHOICES)
            q_qs = Quotation.objects.filter(
                company=company,
                date__range=(start_date, end_date)
            ).values('status').annotate(count=Count('id'))
            counts = {item['status']: item['count'] for item in q_qs}
            for code, name in Quotation.STATUS_CHOICES:
                status_labels.append(name)
                status_values.append(counts.get(code, 0))
        elif status_entity == 'sales_order':
            st_dict = dict(SalesOrder.STATUS_CHOICES)
            so_qs = SalesOrder.objects.filter(
                company=company,
                order_date__range=(start_date, end_date)
            ).values('status').annotate(count=Count('id'))
            counts = {item['status']: item['count'] for item in so_qs}
            for code, name in SalesOrder.STATUS_CHOICES:
                status_labels.append(name)
                status_values.append(counts.get(code, 0))
        else: # default to invoice
            st_dict = dict(Invoice.STATUS_CHOICES)
            inv_qs = Invoice.objects.filter(
                company=company,
                invoice_date__range=(start_date, end_date)
            ).values('status').annotate(count=Count('id'))
            counts = {item['status']: item['count'] for item in inv_qs}
            for code, name in Invoice.STATUS_CHOICES:
                status_labels.append(name)
                status_values.append(counts.get(code, 0))

        return JsonResponse({
            'labels': labels,
            'sales_data': sales_data,
            'purchases_data': purchases_data,
            'expense_labels': expense_labels,
            'expense_values': expense_values,
            'collections_data': collections_data,
            'profit_data': profit_data,
            'status_labels': status_labels,
            'status_values': status_values,
            'status_entity': status_entity,
        })


class AdminDashboardChartDataView(View):
    def get(self, request):
        if request.user.role != 'SUPERADMIN' and not request.user.is_superuser:
            return JsonResponse({'error': 'Unauthorized'}, status=403)
            
        period = request.GET.get('period', 'this_month')
        
        # Parse timezone-aware local date
        today_val = timezone.localtime(timezone.now()).date()
        if period == 'today':
            start_date = today_val
            end_date = today_val
        elif period == 'this_week':
            start_date = today_val - timedelta(days=today_val.weekday())
            end_date = today_val
        elif period == 'this_month':
            start_date = today_val.replace(day=1)
            end_date = today_val
        elif period == 'last_month':
            last_day_of_last_month = today_val.replace(day=1) - timedelta(days=1)
            start_date = last_day_of_last_month.replace(day=1)
            end_date = last_day_of_last_month
        elif period == 'this_year':
            start_date = today_val.replace(month=1, day=1)
            end_date = today_val
        else: # default to this month
            start_date = today_val.replace(day=1)
            end_date = today_val

        # Generate list of dates in the range
        delta = end_date - start_date
        dates_list = [start_date + timedelta(days=i) for i in range(delta.days + 1)]
        labels = [d.strftime('%d %b') for d in dates_list]

        # Registrations Trend (GRAPH 1)
        companies_qs = Company.objects.filter(
            created_at__date__range=(start_date, end_date)
        ).values('created_at__date').annotate(count=Count('id'))
        companies_map = {item['created_at__date']: item['count'] for item in companies_qs}
        registrations_data = [companies_map.get(d, 0) for d in dates_list]

        # SaaS Estimated Subscription Revenue (GRAPH 2)
        rev_qs = Company.objects.filter(
            subscription_status='ACTIVE',
            created_at__date__range=(start_date, end_date)
        ).values('created_at__date').annotate(total=Sum('plan__monthly_price'))
        rev_map = {item['created_at__date']: float(item['total'] or 0.0) for item in rev_qs}
        revenue_data = [round(rev_map.get(d, 0.0), 2) for d in dates_list]

        # Plan Distribution Donut (GRAPH 3)
        plan_qs = Company.objects.values('plan__name').annotate(count=Count('id')).order_by('-count')
        plan_labels = [item['plan__name'] or 'No Plan' for item in plan_qs]
        plan_values = [item['count'] for item in plan_qs]
        if not plan_labels:
            plan_labels = ['No Plans']
            plan_values = [0]

        # Support Ticket Status Distribution (GRAPH 4)
        ticket_map = dict(SupportTicket.STATUS_CHOICES)
        ticket_qs = SupportTicket.objects.filter(
            created_at__date__range=(start_date, end_date)
        ).values('status').annotate(count=Count('id')).order_by('-count')
        
        ticket_labels = []
        ticket_values = []
        for item in ticket_qs:
            status_code = item['status']
            ticket_labels.append(ticket_map.get(status_code, status_code))
            ticket_values.append(item['count'])
            
        if not ticket_labels:
            ticket_labels = ['No Tickets']
            ticket_values = [0]

        # Company Subscription Status Distribution (GRAPH 5)
        company_status_map = dict(Company.SUBSCRIPTION_STATUS_CHOICES)
        comp_status_qs = Company.objects.values('subscription_status').annotate(count=Count('id'))
        comp_counts = {item['subscription_status']: item['count'] for item in comp_status_qs}
        
        company_status_labels = []
        company_status_values = []
        for code, name in Company.SUBSCRIPTION_STATUS_CHOICES:
            company_status_labels.append(name)
            company_status_values.append(comp_counts.get(code, 0))

        # Realtime Recent companies
        recent_comps = Company.objects.all().order_by('-created_at')[:5]
        companies_data = []
        for c in recent_comps:
            companies_data.append({
                'id': c.id,
                'name': c.name,
                'gstin': c.gstin or '-',
                'mobile': c.mobile,
                'status': c.subscription_status,
                'status_display': c.get_subscription_status_display(),
                'created_at': c.created_at.strftime('%d %b %Y')
            })

        # Realtime Recent logs
        recent_logs_qs = AuditLog.objects.select_related('user').order_by('-timestamp')[:10]
        logs_data = []
        for l in recent_logs_qs:
            local_time = timezone.localtime(l.timestamp)
            logs_data.append({
                'username': l.user.username if l.user else 'System',
                'time': local_time.strftime('%H:%M'),
                'action': l.action,
                'module': l.module
            })

        return JsonResponse({
            'labels': labels,
            'registrations_data': registrations_data,
            'revenue_data': revenue_data,
            'plan_labels': plan_labels,
            'plan_values': plan_values,
            'ticket_labels': ticket_labels,
            'ticket_values': ticket_values,
            'company_status_labels': company_status_labels,
            'company_status_values': company_status_values,
            'recent_companies': companies_data,
            'recent_logs': logs_data,
        })


# --- ADMIN PANEL VIEWS ---

class AdminDashboardView(TemplateView):
    template_name = 'admin/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.role != 'SUPERADMIN' and not self.request.user.is_superuser:
            raise PermissionDenied
        
        # Realtime metrics
        context['total_companies'] = Company.objects.count()
        context['active_companies'] = Company.objects.filter(subscription_status='ACTIVE').count()
        context['trial_companies'] = Company.objects.filter(subscription_status='TRIAL').count()
        context['expired_companies'] = Company.objects.filter(subscription_status='EXPIRED').count()
        context['suspended_companies'] = Company.objects.filter(subscription_status='SUSPENDED').count()
        context['total_users'] = CustomUser.objects.exclude(role='SUPERADMIN').count()
        
        # Invoices stat
        context['total_invoices'] = Invoice.objects.filter(status='PAID').count()
        context['today_invoices'] = Invoice.objects.filter(invoice_date=date.today()).count()
        
        # Audit logs & Companies
        context['recent_logs'] = AuditLog.objects.order_by('-timestamp')[:10]
        context['companies'] = Company.objects.order_by('-created_at')[:5]
        
        return context


class AdminDeliveryView(TemplateView):
    template_name = 'admin/delivery_coming_soon.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if not self.request.user.is_authenticated:
            raise PermissionDenied
        if self.request.user.role != 'SUPERADMIN' and not self.request.user.is_superuser:
            raise PermissionDenied
        return context


class AdminCompaniesListView(PaginationMixin, ListView):
    model = Company
    template_name = 'admin/company_list.html'
    context_object_name = 'companies'

    def get_queryset(self):
        if self.request.user.role != 'SUPERADMIN':
            raise PermissionDenied
        qs = Company.objects.filter(is_active=True).order_by('-created_at')
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(subscription_status=status)
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(name__icontains=search) |
                Q(trade_name__icontains=search) |
                Q(gstin__icontains=search) |
                Q(email__icontains=search) |
                Q(mobile__icontains=search) |
                Q(city__icontains=search) |
                Q(state__icontains=search) |
                Q(subscription_status__icontains=search)
            )
        return qs


class AdminCompanyCreateView(View):
    def get(self, request):
        if request.user.role != 'SUPERADMIN':
            raise PermissionDenied
        from .utils import INDIAN_STATES_AND_UTS
        plans = SubscriptionPlan.objects.filter(is_active=True)
        return render(request, 'admin/company_add.html', {'plans': plans, 'indian_states': INDIAN_STATES_AND_UTS})

    @transaction.atomic
    def post(self, request):
        if request.user.role != 'SUPERADMIN':
            raise PermissionDenied
        
        # Extract business fields
        name = request.POST.get('company_name')
        trade_name = request.POST.get('trade_name')
        business_type = request.POST.get('business_type') or 'PROPRIETORSHIP'
        gst_status = request.POST.get('gst_status') or 'UNREGISTERED'
        gstin = request.POST.get('gstin')
        pan = request.POST.get('pan')
        email = request.POST.get('email')
        mobile = request.POST.get('mobile')
        address = request.POST.get('address')
        city = request.POST.get('city')
        state = request.POST.get('state')
        state_code = request.POST.get('state_code')
        pincode = request.POST.get('pincode')
        
        from .utils import validate_state_and_code, INDIAN_STATES_AND_UTS
        is_valid, msg_or_code = validate_state_and_code(state, state_code)
        if not is_valid:
            messages.error(request, msg_or_code)
            plans = SubscriptionPlan.objects.filter(is_active=True)
            return render(request, 'admin/company_add.html', {'plans': plans, 'indian_states': INDIAN_STATES_AND_UTS, 'form_data': request.POST})
        state_code = msg_or_code
        
        # Owner fields
        owner_name = request.POST.get('owner_name')
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # Plan fields
        plan_id = request.POST.get('plan')
        sub_status = request.POST.get('sub_status', 'TRIAL')
        
        plan = SubscriptionPlan.objects.get(id=plan_id)
        
        # Create company
        company = Company.objects.create(
            name=name, trade_name=trade_name, business_type=business_type,
            gst_status=gst_status, gstin=gstin, pan=pan, email=email, mobile=mobile,
            address=address, city=city, state=state, state_code=state_code, pincode=pincode,
            plan=plan, subscription_status=sub_status,
            subscription_start_date=date.today(),
            subscription_end_date=date.today() + timedelta(days=plan.trial_days)
        )
        
        # Create Company Admin user
        user = CustomUser.objects.create_user(
            username=username, password=password, email=email,
            first_name=owner_name, role='ADMIN', company=company
        )
        
        # Add basic default records (default warehouses, units)
        Warehouse.objects.create(company=company, name="Main Warehouse", code="MAIN", is_active=True)
        Unit.objects.create(company=company, name="Pieces", code="PCS")
        Unit.objects.create(company=company, name="Kilograms", code="KGS")
        
        # Create HSN Default records
        HSNSACMaster.objects.create(company=company, code="9983", description="IT Services", type="SAC", gst_rate=Decimal("18.00"))
        
        log_action(request.user, 'CREATE_COMPANY', 'COMPANY', company.id, new_values={'name': name}, request=request)
        
        messages.success(request, f"Company '{name}' and Admin User '{username}' created successfully!")
        return redirect('admin_companies_list')


class AdminCompanyDetailView(DetailView):
    model = Company
    template_name = 'admin/company_detail.html'
    context_object_name = 'company'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.role != 'SUPERADMIN':
            raise PermissionDenied
        
        comp = self.get_object()
        context['users'] = comp.users.all()
        context['invoices'] = Invoice.objects.filter(company=comp).order_by('-invoice_date')[:10]
        context['purchases'] = PurchaseBill.objects.filter(company=comp).order_by('-bill_date')[:10]
        context['audit_logs'] = AuditLog.objects.filter(company=comp).order_by('-timestamp')[:10]
        
        # Stats summary
        context['invoice_count'] = Invoice.objects.filter(company=comp).count()
        context['sales_total'] = Invoice.objects.filter(company=comp, status='PAID').aggregate(sum=Sum('grand_total'))['sum'] or Decimal('0.00')
        context['product_count'] = Product.objects.filter(company=comp).count()
        context['customer_count'] = Customer.objects.filter(company=comp).count()
        return context


class AdminCompanyUpdateView(UpdateView):
    model = Company
    form_class = CompanyForm
    template_name = 'admin/company_edit.html'
    success_url = reverse_lazy('admin_companies_list')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'SUPERADMIN':
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .utils import INDIAN_STATES_AND_UTS
        context['indian_states'] = INDIAN_STATES_AND_UTS
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        log_action(self.request.user, 'UPDATE_COMPANY', 'COMPANY', self.object.id, new_values=form.cleaned_data, request=self.request)
        messages.success(self.request, "Company details updated successfully!")
        return response



def admin_company_status_change(request, pk, status):
    if request.user.role != 'SUPERADMIN':
        raise PermissionDenied
    company = get_object_or_404(Company, id=pk)
    if status in ['ACTIVE', 'SUSPENDED', 'EXPIRED', 'TRIAL']:
        company.subscription_status = status
        company.save()
        log_action(request.user, 'CHANGE_COMPANY_STATUS', 'COMPANY', company.id, new_values={'status': status}, request=request)
        messages.success(request, f"Company subscription status changed to {status}")
    return redirect('admin_company_view', pk=pk)


def admin_login_as_company(request, pk):
    if request.user.role != 'SUPERADMIN':
        raise PermissionDenied
    company = get_object_or_404(Company, id=pk)
    company_admin = company.users.filter(role='ADMIN').first()
    if company_admin:
        logout(request)
        login(request, company_admin)
        messages.success(request, f"Logged in impersonating company user: {company_admin.username}")
        return redirect('company_dashboard')
    else:
        messages.error(request, "No admin user found for this company.")
        return redirect('admin_company_view', pk=pk)


# --- PLANS ---

class AdminPlansView(PaginationMixin, ListView):
    model = SubscriptionPlan
    template_name = 'admin/plans.html'
    context_object_name = 'plans'

    def get_queryset(self):
        if self.request.user.role != 'SUPERADMIN':
            raise PermissionDenied
        qs = SubscriptionPlan.objects.filter(is_active=True).order_by('-id')
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
        return qs


class AdminPlanCreateView(AjaxFormMixin, CreateView):
    model = SubscriptionPlan
    form_class = PlanForm
    template_name = 'admin/plan_add.html'
    success_url = reverse_lazy('admin_plans')


class AdminPlanUpdateView(AjaxFormMixin, UpdateView):
    model = SubscriptionPlan
    form_class = PlanForm
    template_name = 'admin/plan_edit.html'
    success_url = reverse_lazy('admin_plans')


class AdminPlanDeleteView(DeleteView):
    model = SubscriptionPlan
    template_name = 'admin/plan_delete.html'
    success_url = reverse_lazy('admin_plans')


class AdminUsersView(PaginationMixin, ListView):
    model = CustomUser
    template_name = 'admin/users.html'
    context_object_name = 'users'

    def get_queryset(self):
        if self.request.user.role != 'SUPERADMIN':
            raise PermissionDenied
        qs = CustomUser.objects.filter(is_active=True).exclude(role='SUPERADMIN').select_related('company').order_by('-id')
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(username__icontains=search) |
                Q(email__icontains=search) |
                Q(mobile__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(company__name__icontains=search) |
                Q(role__icontains=search)
            )
        return qs


class AdminTicketsView(PaginationMixin, ListView):
    model = SupportTicket
    template_name = 'admin/tickets.html'
    context_object_name = 'tickets'

    def get_queryset(self):
        if self.request.user.role != 'SUPERADMIN':
            raise PermissionDenied
        qs = SupportTicket.objects.select_related('company', 'user').order_by('-created_at')
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(id__icontains=search) |
                Q(subject__icontains=search) |
                Q(description__icontains=search) |
                Q(company__name__icontains=search) |
                Q(user__username__icontains=search) |
                Q(priority__icontains=search) |
                Q(status__icontains=search)
            )
        return qs


def admin_ticket_reply(request, pk):
    if request.user.role != 'SUPERADMIN':
        raise PermissionDenied
    ticket = get_object_or_404(SupportTicket, id=pk)
    if request.method == 'POST':
        reply = request.POST.get('reply')
        ticket.status = 'RESOLVED'
        ticket.save()
        messages.success(request, "Ticket updated successfully!")
    return redirect('admin_tickets')


class AdminReportsView(TemplateView):
    template_name = 'admin/reports.html'


class AdminAuditLogsView(PaginationMixin, ListView):
    model = AuditLog
    template_name = 'admin/audit_logs.html'
    context_object_name = 'logs'

    def get_queryset(self):
        if self.request.user.role != 'SUPERADMIN':
            raise PermissionDenied
        qs = AuditLog.objects.select_related('user', 'company').order_by('-timestamp')
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(action__icontains=search) |
                Q(model_name__icontains=search) |
                Q(user__username__icontains=search) |
                Q(company__name__icontains=search) |
                Q(ip_address__icontains=search)
            )
        return qs


class AdminSettingsView(TemplateView):
    template_name = 'admin/settings.html'


# --- ADMIN HSN/SAC CODE ASSIGNMENT MASTER ---

class AdminHSNSACListView(PaginationMixin, ListView):
    model = HSNSACMaster
    template_name = 'admin/hsn_sac_list.html'
    context_object_name = 'hsn_codes'

    def get_paginate_by(self, queryset):
        try:
            limit = int(self.request.GET.get('entries') or self.request.GET.get('per_page') or self.request.GET.get('page_size') or 10)
            if limit in [10, 25, 50, 100]:
                return limit
        except (ValueError, TypeError):
            pass
        return 10

    def get_queryset(self):
        if not self.request.user.is_authenticated or self.request.user.role != 'SUPERADMIN':
            raise PermissionDenied("Only System Admins can manage HSN/SAC Master.")

        qs = HSNSACMaster.objects.filter(company__isnull=True)

        code_type = self.request.GET.get('type')
        if code_type in ['HSN', 'SAC']:
            qs = qs.filter(type=code_type)

        gst_rate = self.request.GET.get('gst_rate')
        if gst_rate:
            if gst_rate == 'other':
                qs = qs.exclude(gst_rate__in=[Decimal('0.00'), Decimal('5.00'), Decimal('12.00'), Decimal('18.00'), Decimal('28.00')])
            else:
                try:
                    qs = qs.filter(gst_rate=Decimal(gst_rate))
                except Exception:
                    pass

        status = self.request.GET.get('status')
        if status == 'active':
            qs = qs.filter(is_active=True)
        elif status == 'inactive':
            qs = qs.filter(is_active=False)

        category = self.request.GET.get('category')
        if category:
            qs = qs.filter(category__iexact=category)

        sub_category = self.request.GET.get('sub_category')
        if sub_category:
            qs = qs.filter(sub_category__iexact=sub_category)

        uqc = self.request.GET.get('uqc')
        if uqc:
            qs = qs.filter(uqc__iexact=uqc)

        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')
        if from_date:
            try:
                qs = qs.filter(effective_from__gte=datetime.strptime(from_date, '%Y-%m-%d').date())
            except Exception:
                pass
        if to_date:
            try:
                qs = qs.filter(effective_from__lte=datetime.strptime(to_date, '%Y-%m-%d').date())
            except Exception:
                pass

        search = self.request.GET.get('search')
        if search:
            qs = qs.filter(
                Q(code__icontains=search) |
                Q(description__icontains=search) |
                Q(category__icontains=search) |
                Q(sub_category__icontains=search) |
                Q(uqc__icontains=search) |
                Q(type__icontains=search)
            )

        sort_by = self.request.GET.get('sort_by', 'code_asc')
        if sort_by == 'code_asc':
            qs = qs.order_by('code', 'id')
        elif sort_by == 'code_desc':
            qs = qs.order_by('-code', '-id')
        elif sort_by == 'desc_asc':
            qs = qs.order_by('description', 'id')
        elif sort_by == 'desc_desc':
            qs = qs.order_by('-description', '-id')
        elif sort_by == 'rate_asc':
            qs = qs.order_by('gst_rate', 'code')
        elif sort_by == 'rate_desc':
            qs = qs.order_by('-gst_rate', 'code')
        elif sort_by == 'date_asc':
            qs = qs.order_by('created_at', 'id')
        elif sort_by == 'date_desc':
            qs = qs.order_by('-created_at', '-id')
        else:
            qs = qs.order_by('code', 'id')

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        base_qs = HSNSACMaster.objects.filter(company__isnull=True)

        summary_stats = base_qs.aggregate(
            total_codes=Count('id'),
            active_codes=Count('id', filter=Q(is_active=True)),
            inactive_codes=Count('id', filter=Q(is_active=False)),
            hsn_codes=Count('id', filter=Q(type='HSN')),
            sac_codes=Count('id', filter=Q(type='SAC'))
        )

        context['summary'] = summary_stats
        context['categories'] = base_qs.exclude(category__isnull=True).exclude(category='').values_list('category', flat=True).distinct().order_by('category')
        context['sub_categories'] = base_qs.exclude(sub_category__isnull=True).exclude(sub_category='').values_list('sub_category', flat=True).distinct().order_by('sub_category')
        context['uqcs'] = base_qs.exclude(uqc__isnull=True).exclude(uqc='').values_list('uqc', flat=True).distinct().order_by('uqc')
        context['current_time'] = timezone.localtime(timezone.now())
        context['entries'] = str(self.get_paginate_by(self.get_queryset()))
        return context

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format in ['csv', 'excel']:
            return self.export_data(export_format)

        print_format = request.GET.get('print')
        if print_format == 'true':
            self.object_list = self.get_queryset()
            context = self.get_context_data()
            return render(request, 'admin/hsn_sac_print.html', context)

        return super().get(request, *args, **kwargs)

    def export_data(self, export_format):
        qs = self.get_queryset()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        headers = ['Code', 'Type', 'Description', 'Category', 'Sub Category', 'GST Rate', 'CGST Rate', 'SGST Rate', 'IGST Rate', 'Cess Rate', 'UQC', 'Status']

        if export_format == 'excel':
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            import io

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "HSN SAC Master"

            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin', color='D1D5DB'),
                right=Side(style='thin', color='D1D5DB'),
                top=Side(style='thin', color='D1D5DB'),
                bottom=Side(style='thin', color='D1D5DB')
            )

            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for item in qs:
                ws.append([
                    item.code,
                    item.type,
                    item.description,
                    item.category or '',
                    item.sub_category or '',
                    float(item.gst_rate),
                    float(item.get_cgst_rate()),
                    float(item.get_sgst_rate()),
                    float(item.get_igst_rate()),
                    float(item.cess_rate),
                    item.uqc or '',
                    'Active' if item.is_active else 'Inactive'
                ])

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border
                    cell.font = Font(name="Arial", size=9)

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="HSN_SAC_Master_{timestamp}.xlsx"'
            return response
        else:
            import csv
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="HSN_SAC_Master_{timestamp}.csv"'
            writer = csv.writer(response)
            writer.writerow(headers)
            for item in qs:
                writer.writerow([
                    item.code,
                    item.type,
                    item.description,
                    item.category or '',
                    item.sub_category or '',
                    item.gst_rate,
                    item.get_cgst_rate(),
                    item.get_sgst_rate(),
                    item.get_igst_rate(),
                    item.cess_rate,
                    item.uqc or '',
                    'Active' if item.is_active else 'Inactive'
                ])
            return response


def admin_hsn_sac_detail(request, pk):
    if not request.user.is_authenticated or request.user.role != 'SUPERADMIN':
        return JsonResponse({'status': 'error', 'message': 'Permission denied.'}, status=403)

    obj = get_object_or_404(HSNSACMaster, id=pk, company__isnull=True)
    return JsonResponse({
        'status': 'success',
        'data': {
            'id': obj.id,
            'code': obj.code,
            'type': obj.type,
            'description': obj.description,
            'category': obj.category or '',
            'sub_category': obj.sub_category or '',
            'gst_rate': str(obj.gst_rate),
            'cgst_rate': str(obj.get_cgst_rate()),
            'sgst_rate': str(obj.get_sgst_rate()),
            'igst_rate': str(obj.get_igst_rate()),
            'cess_rate': str(obj.cess_rate),
            'uqc': obj.uqc or 'PCS',
            'effective_from': obj.effective_from.strftime('%Y-%m-%d') if obj.effective_from else '',
            'effective_to': obj.effective_to.strftime('%Y-%m-%d') if obj.effective_to else '',
            'is_active': obj.is_active,
            'created_at': obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else '',
            'updated_at': obj.updated_at.strftime('%Y-%m-%d %H:%M:%S') if obj.updated_at else '',
        }
    })


@transaction.atomic
def admin_hsn_sac_add(request):
    if not request.user.is_authenticated or request.user.role != 'SUPERADMIN':
        return JsonResponse({'status': 'error', 'message': 'Permission denied.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)

    try:
        data = {}
        if request.body:
            try:
                data = json.loads(request.body)
            except Exception:
                data = request.POST
        else:
            data = request.POST

        code_type = data.get('type', 'HSN').upper()
        code = data.get('code', '').strip()
        description = data.get('description', '').strip()
        category = data.get('category', '').strip() or None
        sub_category = data.get('sub_category', '').strip() or None
        gst_rate = Decimal(str(data.get('gst_rate', '18.00')))
        cess_rate = Decimal(str(data.get('cess_rate', '0.00'))) if data.get('cess_rate') not in [None, ''] else Decimal('0.00')
        uqc = data.get('uqc', 'PCS').strip().upper() or 'PCS'
        is_active = str(data.get('is_active', 'true')).lower() in ['true', '1', 'active']

        if not code or not description:
            return JsonResponse({'status': 'error', 'message': 'Invalid HSN/SAC data. Code and Description are required.'}, status=400)

        if code_type not in ['HSN', 'SAC']:
            return JsonResponse({'status': 'error', 'message': 'Invalid HSN/SAC data.'}, status=400)

        if code_type == 'HSN':
            if not code.isdigit() or len(code) < 2 or len(code) > 8:
                return JsonResponse({'status': 'error', 'message': 'Invalid HSN code. HSN code must be a numeric value between 2 and 8 digits.'}, status=400)
        elif code_type == 'SAC':
            if not code.isdigit() or not code.startswith('99') or len(code) != 6:
                return JsonResponse({'status': 'error', 'message': 'Invalid SAC code. SAC code must be a 6-digit number starting with 99 (e.g., 998311).'}, status=400)

        if HSNSACMaster.objects.filter(company__isnull=True, type=code_type, code=code).exists():
            return JsonResponse({'status': 'error', 'message': 'HSN/SAC code already exists.'}, status=400)

        cgst = Decimal(str(data.get('cgst_rate'))) if data.get('cgst_rate') not in [None, ''] else (gst_rate / Decimal('2.00')).quantize(Decimal('0.01'))
        sgst = Decimal(str(data.get('sgst_rate'))) if data.get('sgst_rate') not in [None, ''] else (gst_rate / Decimal('2.00')).quantize(Decimal('0.01'))
        igst = Decimal(str(data.get('igst_rate'))) if data.get('igst_rate') not in [None, ''] else gst_rate

        eff_from = datetime.strptime(data.get('effective_from'), '%Y-%m-%d').date() if data.get('effective_from') else None
        eff_to = datetime.strptime(data.get('effective_to'), '%Y-%m-%d').date() if data.get('effective_to') else None

        obj = HSNSACMaster.objects.create(
            company=None,
            type=code_type,
            code=code,
            description=description,
            category=category,
            sub_category=sub_category,
            gst_rate=gst_rate,
            cgst_rate=cgst,
            sgst_rate=sgst,
            igst_rate=igst,
            cess_rate=cess_rate,
            uqc=uqc,
            effective_from=eff_from,
            effective_to=eff_to,
            is_active=is_active,
            created_by=request.user
        )

        log_action(request.user, 'CREATE_HSN_SAC', 'HSN_SAC', obj.id, new_values={'code': code, 'type': code_type, 'gst_rate': str(gst_rate)}, request=request)
        if hasattr(request, '_messages'):
            messages.success(request, f"HSN/SAC code added successfully.")
        return JsonResponse({'status': 'success', 'message': f"HSN/SAC code added successfully."})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Invalid HSN/SAC data: {str(e)}'}, status=400)


@transaction.atomic
def admin_hsn_sac_edit(request, pk):
    if not request.user.is_authenticated or request.user.role != 'SUPERADMIN':
        return JsonResponse({'status': 'error', 'message': 'Permission denied.'}, status=403)

    obj = get_object_or_404(HSNSACMaster, id=pk, company__isnull=True)

    if request.method == 'GET':
        return JsonResponse({
            'status': 'success',
            'data': {
                'id': obj.id,
                'type': obj.type,
                'code': obj.code,
                'description': obj.description,
                'category': obj.category or '',
                'sub_category': obj.sub_category or '',
                'gst_rate': str(obj.gst_rate),
                'cgst_rate': str(obj.get_cgst_rate()),
                'sgst_rate': str(obj.get_sgst_rate()),
                'igst_rate': str(obj.get_igst_rate()),
                'cess_rate': str(obj.cess_rate),
                'uqc': obj.uqc or '',
                'effective_from': obj.effective_from.strftime('%Y-%m-%d') if obj.effective_from else '',
                'effective_to': obj.effective_to.strftime('%Y-%m-%d') if obj.effective_to else '',
                'is_active': obj.is_active,
                'created_at': obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else '',
                'updated_at': obj.updated_at.strftime('%Y-%m-%d %H:%M:%S') if obj.updated_at else ''
            }
        })

    try:
        data = json.loads(request.body) if request.body and request.content_type == 'application/json' else request.POST
        code_type = data.get('type', obj.type).upper()
        code = data.get('code', obj.code).strip()
        description = data.get('description', obj.description).strip()
        category = data.get('category', '').strip() or None
        sub_category = data.get('sub_category', '').strip() or None
        gst_rate = Decimal(str(data.get('gst_rate', obj.gst_rate)))
        cess_rate = Decimal(str(data.get('cess_rate', obj.cess_rate))) if data.get('cess_rate') not in [None, ''] else Decimal('0.00')
        uqc = data.get('uqc', 'PCS').strip().upper() or 'PCS'
        is_active = str(data.get('is_active', obj.is_active)).lower() in ['true', '1', 'active']

        if not code or not description:
            return JsonResponse({'status': 'error', 'message': 'Invalid HSN/SAC data. Code and Description are required.'}, status=400)

        if code_type not in ['HSN', 'SAC']:
            return JsonResponse({'status': 'error', 'message': 'Invalid HSN/SAC data.'}, status=400)

        if code_type == 'HSN':
            if not code.isdigit() or len(code) < 2 or len(code) > 8:
                return JsonResponse({'status': 'error', 'message': 'Invalid HSN code. HSN code must be a numeric value between 2 and 8 digits.'}, status=400)
        elif code_type == 'SAC':
            if not code.isdigit() or not code.startswith('99') or len(code) != 6:
                return JsonResponse({'status': 'error', 'message': 'Invalid SAC code. SAC code must be a 6-digit number starting with 99 (e.g., 998311).'}, status=400)

        if HSNSACMaster.objects.filter(company__isnull=True, type=code_type, code=code).exclude(id=obj.id).exists():
            return JsonResponse({'status': 'error', 'message': 'HSN/SAC code already exists.'}, status=400)

        cgst = Decimal(str(data.get('cgst_rate'))) if data.get('cgst_rate') not in [None, ''] else (gst_rate / Decimal('2.00')).quantize(Decimal('0.01'))
        sgst = Decimal(str(data.get('sgst_rate'))) if data.get('sgst_rate') not in [None, ''] else (gst_rate / Decimal('2.00')).quantize(Decimal('0.01'))
        igst = Decimal(str(data.get('igst_rate'))) if data.get('igst_rate') not in [None, ''] else gst_rate

        eff_from = datetime.strptime(data.get('effective_from'), '%Y-%m-%d').date() if data.get('effective_from') else None
        eff_to = datetime.strptime(data.get('effective_to'), '%Y-%m-%d').date() if data.get('effective_to') else None

        obj.type = code_type
        obj.code = code
        obj.description = description
        obj.category = category
        obj.sub_category = sub_category
        obj.gst_rate = gst_rate
        obj.cgst_rate = cgst
        obj.sgst_rate = sgst
        obj.igst_rate = igst
        obj.cess_rate = cess_rate
        obj.uqc = uqc
        obj.effective_from = eff_from
        obj.effective_to = eff_to
        obj.is_active = is_active
        obj.save()

        log_action(request.user, 'UPDATE_HSN_SAC', 'HSN_SAC', obj.id, new_values={'code': code, 'type': code_type, 'gst_rate': str(gst_rate)}, request=request)
        return JsonResponse({'status': 'success', 'message': f"HSN/SAC code updated successfully."})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': 'Invalid HSN/SAC data.'}, status=400)


@transaction.atomic
def admin_hsn_sac_status_change(request, pk, status):
    if not request.user.is_authenticated or request.user.role != 'SUPERADMIN':
        return JsonResponse({'status': 'error', 'message': 'Permission denied.'}, status=403)

    obj = get_object_or_404(HSNSACMaster, id=pk, company__isnull=True)
    is_active = (status.lower() == 'active')
    obj.is_active = is_active
    obj.save()

    log_action(request.user, 'STATUS_HSN_SAC', 'HSN_SAC', obj.id, new_values={'is_active': is_active}, request=request)
    if hasattr(request, '_messages'):
        messages.success(request, f"HSN/SAC Code '{obj.code}' status changed to {'Active' if is_active else 'Inactive'}.")
    return redirect('admin_hsn_sac_list')


@transaction.atomic
def admin_hsn_sac_delete(request, pk):
    if not request.user.is_authenticated or request.user.role != 'SUPERADMIN':
        return JsonResponse({'success': False, 'status': 'error', 'message': 'Permission denied.'}, status=403)

    obj = get_object_or_404(HSNSACMaster, id=pk, company__isnull=True)

    is_referenced = (
        Product.objects.filter(hsn_sac=obj).exists() or
        InvoiceItem.objects.filter(hsn_sac_code=obj.code).exists() or
        QuotationItem.objects.filter(product__hsn_sac=obj).exists() or
        SalesOrderItem.objects.filter(product__hsn_sac=obj).exists() or
        PurchaseBillItem.objects.filter(hsn_sac_code=obj.code).exists() or
        CreditNoteItem.objects.filter(hsn_sac_code=obj.code).exists() or
        DebitNoteItem.objects.filter(hsn_sac_code=obj.code).exists() or
        GSTTransaction.objects.filter(hsn_sac_code=obj.code).exists()
    )

    if is_referenced:
        obj.is_active = False
        obj.save()
        log_action(request.user, 'DEACTIVATE_HSN_SAC', 'HSN_SAC', obj.id, request=request)
        msg = f"HSN/SAC code deactivated because it is referenced in billing records."
        if hasattr(request, '_messages'):
            messages.warning(request, msg)
        return JsonResponse({
            'success': True,
            'status': 'warning',
            'action': 'deactivated',
            'message': msg,
            'id': int(pk)
        })

    code = obj.code
    obj.delete()
    log_action(request.user, 'DELETE_HSN_SAC', 'HSN_SAC', pk, request=request)
    msg = f"HSN/SAC code deleted successfully."
    if hasattr(request, '_messages'):
        messages.success(request, msg)
    return JsonResponse({
        'success': True,
        'status': 'success',
        'action': 'deleted',
        'message': msg,
        'id': int(pk)
    })


def admin_hsn_sac_sample_template(request):
    import csv
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="HSN_SAC_Bulk_Template.csv"'

    writer = csv.writer(response)
    writer.writerow(['Code', 'Type', 'Description', 'GST Rate', 'CGST Rate', 'SGST Rate', 'IGST Rate', 'Cess Rate', 'Status'])
    writer.writerow(['1001', 'HSN', 'Wheat and meslin', '5.00', '2.50', '2.50', '5.00', '0.00', 'Active'])
    writer.writerow(['998311', 'SAC', 'Management consulting services', '18.00', '9.00', '9.00', '18.00', '0.00', 'Active'])
    return response


def admin_hsn_sac_bulk_preview(request):
    if not request.user.is_authenticated or request.user.role != 'SUPERADMIN':
        return JsonResponse({'status': 'error', 'message': 'Permission denied.'}, status=403)

    if request.method != 'POST' or 'file' not in request.FILES:
        return JsonResponse({'status': 'error', 'message': 'Please upload a CSV or Excel (.xlsx) file.'}, status=400)

    uploaded_file = request.FILES['file']
    filename = uploaded_file.name.lower()

    rows_data = []

    try:
        if filename.endswith('.csv'):
            import csv
            file_data = uploaded_file.read().decode('utf-8-sig', errors='ignore').splitlines()
            reader = csv.reader(file_data)
            header = None
            for row in reader:
                if not row or not any(row):
                    continue
                if header is None:
                    header = [c.strip().lower() for c in row]
                    continue
                rows_data.append(dict(zip(header, [c.strip() for c in row])))
        elif filename.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            sheet = wb.active
            header = None
            for row in sheet.iter_rows(values_only=True):
                if not row or not any(row):
                    continue
                row_vals = [str(c).strip() if c is not None else '' for c in row]
                if header is None:
                    header = [c.lower() for c in row_vals]
                    continue
                rows_data.append(dict(zip(header, row_vals)))
        else:
            return JsonResponse({'status': 'error', 'message': 'Unsupported file format. Please upload a .csv or .xlsx file.'}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f"Error parsing file: {str(e)}"}, status=400)

    existing_db_codes = set(HSNSACMaster.objects.filter(company__isnull=True).values_list('type', 'code'))
    seen_file_codes = set()

    parsed_rows = []
    valid_count = 0
    invalid_count = 0
    new_count = 0
    update_count = 0

    for i, raw in enumerate(rows_data, 1):
        def extract_hsn_row_data(raw_dict):
            norm = {}
            for k, v in raw_dict.items():
                clean_k = str(k or '').lower().strip().replace('_', ' ').replace('-', ' ')
                clean_k = ' '.join(clean_k.split())
                val = str(v).strip() if v is not None else ''
                if clean_k in ['code type', 'type', 'item type', 'hsn/sac type', 'hsn type', 'sac type']:
                    norm['type'] = val
                elif clean_k in ['hsn/sac code', 'hsn code', 'sac code', 'hsn / sac code', 'hsn_sac_code', 'code', 'hsn', 'sac']:
                    norm['code'] = val
                elif clean_k in ['description', 'desc', 'hsn description', 'sac description', 'details', 'item description']:
                    norm['description'] = val
                elif clean_k in ['category', 'cat']:
                    norm['category'] = val
                elif clean_k in ['sub category', 'sub_category', 'subcategory', 'sub cat', 'sub-category']:
                    norm['sub_category'] = val
                elif clean_k in ['gst rate', 'gst %', 'gst rate (%)', 'gst', 'tax rate', 'tax %', 'gst_rate']:
                    norm['gst_rate'] = val
                elif clean_k in ['cgst rate', 'cgst %', 'cgst rate (%)', 'cgst', 'cgst_rate']:
                    norm['cgst_rate'] = val
                elif clean_k in ['sgst rate', 'sgst %', 'sgst rate (%)', 'sgst', 'sgst_rate']:
                    norm['sgst_rate'] = val
                elif clean_k in ['igst rate', 'igst %', 'igst rate (%)', 'igst', 'igst_rate']:
                    norm['igst_rate'] = val
                elif clean_k in ['uqc', 'unit', 'uom', 'uqc (unit)', 'uqc unit']:
                    norm['uqc'] = val
                elif clean_k in ['effective from', 'from date', 'from', 'effective_from']:
                    norm['effective_from'] = val
                elif clean_k in ['effective to', 'to date', 'to', 'effective_to']:
                    norm['effective_to'] = val
                elif clean_k in ['status', 'is active', 'active', 'is_active']:
                    norm['status'] = val
            return norm

        row_map = extract_hsn_row_data(raw)
        code_type = (row_map.get('type') or '').upper()
        if 'GOOD' in code_type: code_type = 'HSN'
        if 'SERV' in code_type: code_type = 'SAC'
        if not code_type: code_type = 'HSN'

        code = row_map.get('code', '').strip()
        description = row_map.get('description', '').strip()
        category = row_map.get('category', '').strip()
        sub_category = row_map.get('sub_category', '').strip()
        gst_str = row_map.get('gst_rate', '').replace('%', '').strip()
        cgst_str = row_map.get('cgst_rate', '').replace('%', '').strip()
        sgst_str = row_map.get('sgst_rate', '').replace('%', '').strip()
        igst_str = row_map.get('igst_rate', '').replace('%', '').strip()
        uqc = row_map.get('uqc', '').upper().strip() or 'PCS'
        eff_from = row_map.get('effective_from', '').strip()
        eff_to = row_map.get('effective_to', '').strip()
        status_str = row_map.get('status', '').strip().lower()
        is_active = status_str not in ['inactive', '0', 'false', 'disabled']

        errors = []

        if code_type not in ['HSN', 'SAC']:
            errors.append("Invalid Code Type (Must be HSN or SAC)")

        if not code:
            errors.append("Missing HSN/SAC Code")

        if not description:
            errors.append("Missing Description")

        gst_rate = Decimal('18.00')
        if not gst_str:
            errors.append("GST rate is not configured")
        else:
            try:
                gst_rate = Decimal(gst_str)
                if gst_rate < 0 or gst_rate > 100:
                    errors.append("GST Rate must be between 0 and 100")
            except Exception:
                errors.append("Invalid GST Rate format")

        cgst_rate = None
        if cgst_str:
            try: cgst_rate = str(Decimal(cgst_str))
            except Exception: pass

        sgst_rate = None
        if sgst_str:
            try: sgst_rate = str(Decimal(sgst_str))
            except Exception: pass

        igst_rate = None
        if igst_str:
            try: igst_rate = str(Decimal(igst_str))
            except Exception: pass

        file_key = (code_type, code)
        if file_key in seen_file_codes:
            errors.append(f"Duplicate code '{code}' in uploaded file")
        elif code:
            seen_file_codes.add(file_key)

        is_update = file_key in existing_db_codes
        action = 'UPDATE' if is_update else 'NEW'

        is_valid = len(errors) == 0
        if is_valid:
            valid_count += 1
            if is_update:
                update_count += 1
            else:
                new_count += 1
        else:
            invalid_count += 1

        parsed_rows.append({
            'row_num': i,
            'type': code_type,
            'code': code,
            'description': description,
            'category': category,
            'sub_category': sub_category,
            'gst_rate': str(gst_rate),
            'cgst_rate': cgst_rate or str((gst_rate / Decimal('2.00')).quantize(Decimal('0.01'))),
            'sgst_rate': sgst_rate or str((gst_rate / Decimal('2.00')).quantize(Decimal('0.01'))),
            'igst_rate': igst_rate or str(gst_rate),
            'uqc': uqc,
            'effective_from': eff_from,
            'effective_to': eff_to,
            'is_active': is_active,
            'is_valid': is_valid,
            'errors': '; '.join(errors),
            'action': action
        })

    return JsonResponse({
        'status': 'success',
        'summary': {
            'total_rows': len(rows_data),
            'valid_rows': valid_count,
            'invalid_rows': invalid_count,
            'new_rows': new_count,
            'update_rows': update_count
        },
        'rows': parsed_rows
    })


@transaction.atomic
def admin_hsn_sac_bulk_import(request):
    if not request.user.is_authenticated or request.user.role != 'SUPERADMIN':
        return JsonResponse({'status': 'error', 'message': 'Permission denied.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)

    try:
        data = json.loads(request.body)
        valid_rows = data.get('valid_rows', [])
        import_mode = data.get('import_mode', 'add_update')

        imported_new = 0
        updated_existing = 0

        for row in valid_rows:
            code_type = row.get('type')
            code = row.get('code')
            if not code_type or not code:
                continue

            obj = HSNSACMaster.objects.filter(company__isnull=True, type=code_type, code=code).first()

            if obj and import_mode in ['add_update', 'update_only']:
                obj.description = row.get('description', obj.description)
                obj.category = row.get('category') or None
                obj.sub_category = row.get('sub_category') or None
                obj.gst_rate = Decimal(str(row.get('gst_rate', '18.00')))
                obj.cgst_rate = Decimal(str(row.get('cgst_rate'))) if row.get('cgst_rate') else None
                obj.sgst_rate = Decimal(str(row.get('sgst_rate'))) if row.get('sgst_rate') else None
                obj.igst_rate = Decimal(str(row.get('igst_rate'))) if row.get('igst_rate') else None
                obj.uqc = row.get('uqc', 'PCS')
                obj.is_active = row.get('is_active', True)
                if row.get('effective_from'):
                    try: obj.effective_from = datetime.strptime(row.get('effective_from'), '%Y-%m-%d').date()
                    except Exception: pass
                if row.get('effective_to'):
                    try: obj.effective_to = datetime.strptime(row.get('effective_to'), '%Y-%m-%d').date()
                    except Exception: pass
                obj.save()
                updated_existing += 1
            elif not obj and import_mode in ['add_update', 'add_only']:
                eff_from = None
                eff_to = None
                if row.get('effective_from'):
                    try: eff_from = datetime.strptime(row.get('effective_from'), '%Y-%m-%d').date()
                    except Exception: pass
                if row.get('effective_to'):
                    try: eff_to = datetime.strptime(row.get('effective_to'), '%Y-%m-%d').date()
                    except Exception: pass

                HSNSACMaster.objects.create(
                    company=None,
                    type=code_type,
                    code=code,
                    description=row.get('description', ''),
                    category=row.get('category') or None,
                    sub_category=row.get('sub_category') or None,
                    gst_rate=Decimal(str(row.get('gst_rate', '18.00'))),
                    cgst_rate=Decimal(str(row.get('cgst_rate'))) if row.get('cgst_rate') else None,
                    sgst_rate=Decimal(str(row.get('sgst_rate'))) if row.get('sgst_rate') else None,
                    igst_rate=Decimal(str(row.get('igst_rate'))) if row.get('igst_rate') else None,
                    uqc=row.get('uqc', 'PCS'),
                    effective_from=eff_from,
                    effective_to=eff_to,
                    is_active=row.get('is_active', True),
                    created_by=request.user
                )
                imported_new += 1

        log_action(request.user, 'BULK_IMPORT_HSN_SAC', 'HSN_SAC', 0, new_values={'new': imported_new, 'updated': updated_existing}, request=request)
        messages.success(request, f"Bulk Import Completed: {imported_new} new codes added, {updated_existing} existing codes updated.")
        return JsonResponse({'status': 'success', 'message': f"Import Completed: {imported_new} new codes added, {updated_existing} codes updated."})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def admin_hsn_sac_export_error_report(request):
    import csv
    if request.method != 'POST':
        return HttpResponse("Method not allowed", status=405)

    try:
        data = json.loads(request.body)
        invalid_rows = data.get('invalid_rows', [])

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="HSN_SAC_Import_Error_Report.csv"'

        writer = csv.writer(response)
        writer.writerow(['Row', 'Code Type', 'HSN/SAC Code', 'Description', 'Category', 'GST Rate', 'UQC', 'Error Message'])

        for row in invalid_rows:
            writer.writerow([
                row.get('row_num', ''),
                row.get('type', ''),
                row.get('code', ''),
                row.get('description', ''),
                row.get('category', ''),
                row.get('gst_rate', ''),
                row.get('uqc', ''),
                row.get('errors', 'Validation failed')
            ])

        return response
    except Exception as e:
        return HttpResponse(f"Error generating report: {str(e)}", status=400)


# --- COMPANY PANEL VIEWS ---

def get_company_dashboard_metrics(company):
    today = timezone.localtime(timezone.now()).date()
    yesterday = today - timedelta(days=1)
    
    valid_sales_statuses = ['POSTED', 'PARTIALLY_PAID', 'PAID']
    valid_purchase_statuses = ['POSTED', 'PARTIALLY_PAID', 'PAID']

    today_sales = Invoice.objects.filter(company=company, invoice_date=today, status__in=valid_sales_statuses).aggregate(sum=Sum('grand_total'))['sum'] or Decimal('0.00')
    today_purchase = PurchaseBill.objects.filter(company=company, bill_date=today, status__in=valid_purchase_statuses).aggregate(sum=Sum('grand_total'))['sum'] or Decimal('0.00')
    today_collection = Payment.objects.filter(company=company, payment_date=today, payment_type='RECEIPT').aggregate(sum=Sum('amount'))['sum'] or Decimal('0.00')
    
    yesterday_sales = Invoice.objects.filter(company=company, invoice_date=yesterday, status__in=valid_sales_statuses).aggregate(sum=Sum('grand_total'))['sum'] or Decimal('0.00')
    yesterday_purchase = PurchaseBill.objects.filter(company=company, bill_date=yesterday, status__in=valid_purchase_statuses).aggregate(sum=Sum('grand_total'))['sum'] or Decimal('0.00')
    yesterday_collection = Payment.objects.filter(company=company, payment_date=yesterday, payment_type='RECEIPT').aggregate(sum=Sum('amount'))['sum'] or Decimal('0.00')
    
    if yesterday_sales > 0:
        sales_trend = round(float((today_sales - yesterday_sales) / yesterday_sales * 100), 1)
    else:
        sales_trend = 0.0 if today_sales == 0 else 100.0
        
    if yesterday_purchase > 0:
        purchase_trend = round(float((today_purchase - yesterday_purchase) / yesterday_purchase * 100), 1)
    else:
        purchase_trend = 0.0 if today_purchase == 0 else 100.0
        
    if yesterday_collection > 0:
        collection_trend = round(float((today_collection - yesterday_collection) / yesterday_collection * 100), 1)
    else:
        collection_trend = 0.0 if today_collection == 0 else 100.0
    
    receivable = Customer.objects.filter(company=company).aggregate(sum=Sum('outstanding_balance'))['sum'] or Decimal('0.00')
    payable = Supplier.objects.filter(company=company).aggregate(sum=Sum('outstanding_balance'))['sum'] or Decimal('0.00')
    
    total_stock_value = Product.objects.filter(company=company).aggregate(val=Sum(F('current_stock') * F('selling_price')))['val'] or Decimal('0.00')
    low_stock_items = Product.objects.filter(company=company, current_stock__lte=F('min_stock'), track_inventory=True).count()
    
    first_day_of_month = today.replace(day=1)
    mtd_sales = Invoice.objects.filter(company=company, invoice_date__range=(first_day_of_month, today), status__in=valid_sales_statuses).aggregate(sum=Sum('grand_total'))['sum'] or Decimal('0.00')
    mtd_purchases = PurchaseBill.objects.filter(company=company, bill_date__range=(first_day_of_month, today), status__in=valid_purchase_statuses).aggregate(sum=Sum('grand_total'))['sum'] or Decimal('0.00')
    mtd_expenses = Expense.objects.filter(company=company, created_at__date__range=(first_day_of_month, today)).aggregate(sum=Sum('amount'))['sum'] or Decimal('0.00')
    profit_mtd = mtd_sales - mtd_purchases - mtd_expenses
    
    last_day_of_last_month = first_day_of_month - timedelta(days=1)
    first_day_of_last_month = last_day_of_last_month.replace(day=1)
    lm_sales = Invoice.objects.filter(company=company, invoice_date__range=(first_day_of_last_month, last_day_of_last_month), status__in=valid_sales_statuses).aggregate(sum=Sum('grand_total'))['sum'] or Decimal('0.00')
    lm_purchases = PurchaseBill.objects.filter(company=company, bill_date__range=(first_day_of_last_month, last_day_of_last_month), status__in=valid_purchase_statuses).aggregate(sum=Sum('grand_total'))['sum'] or Decimal('0.00')
    lm_expenses = Expense.objects.filter(company=company, created_at__date__range=(first_day_of_last_month, last_day_of_last_month)).aggregate(sum=Sum('amount'))['sum'] or Decimal('0.00')
    lm_profit = lm_sales - lm_purchases - lm_expenses
    
    if lm_profit > 0:
        profit_trend = round(float((profit_mtd - lm_profit) / lm_profit * 100), 1)
    else:
        profit_trend = 0.0 if profit_mtd == 0 else 100.0

    input_gst = PurchaseBill.objects.filter(company=company, status='POSTED').aggregate(cgst=Sum('cgst_total'), sgst=Sum('sgst_total'), igst=Sum('igst_total'))
    output_gst = Invoice.objects.filter(company=company, status='POSTED').aggregate(cgst=Sum('cgst_total'), sgst=Sum('sgst_total'), igst=Sum('igst_total'))
    
    recent_invoices = Invoice.objects.filter(company=company).order_by('-created_at')[:5]
    recent_payments = Payment.objects.filter(company=company).order_by('-created_at')[:5]
    low_stock_alerts = Product.objects.filter(company=company, current_stock__lte=F('min_stock'), track_inventory=True)[:5]
    
    return {
        'today_sales': today_sales,
        'today_purchase': today_purchase,
        'today_collection': today_collection,
        'sales_trend': sales_trend,
        'purchase_trend': purchase_trend,
        'collection_trend': collection_trend,
        'receivable': receivable,
        'payable': payable,
        'total_stock_value': total_stock_value,
        'low_stock_items': low_stock_items,
        'profit_mtd': profit_mtd,
        'profit_trend': profit_trend,
        'input_gst': input_gst,
        'output_gst': output_gst,
        'recent_invoices': recent_invoices,
        'recent_payments': recent_payments,
        'low_stock_alerts': low_stock_alerts,
    }


class HowToUseView(LoginRequiredMixin, TemplateView):
    template_name = 'company/how_to_use.html'


class CompanyDashboardView(CompanyRequiredMixin, TemplateView):
    template_name = 'company/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.user.company
        metrics = get_company_dashboard_metrics(company)
        context.update(metrics)
        return context


class CompanyDashboardDataApiView(CompanyRequiredMixin, View):
    def get(self, request):
        company = request.user.company
        metrics = get_company_dashboard_metrics(company)
        return JsonResponse({
            'today_sales': float(metrics['today_sales']),
            'today_purchase': float(metrics['today_purchase']),
            'today_collection': float(metrics['today_collection']),
            'sales_trend': metrics['sales_trend'],
            'purchase_trend': metrics['purchase_trend'],
            'collection_trend': metrics['collection_trend'],
            'receivable': float(metrics['receivable']),
            'payable': float(metrics['payable']),
            'total_stock_value': float(metrics['total_stock_value']),
            'low_stock_items': metrics['low_stock_items'],
            'profit_mtd': float(metrics['profit_mtd']),
            'profit_trend': metrics['profit_trend'],
        })


class CompanyDeliveryView(CompanyRequiredMixin, TemplateView):
    template_name = 'company/delivery_coming_soon.html'


class CompanySettingsView(CompanyRequiredMixin, View):
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if request.user.role != 'ADMIN':
            raise PermissionDenied("Only Company Admins are allowed to access and manage business settings.")
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        company = request.user.company
        form = CompanyForm(instance=company)
        users = company.users.all()
        logs = AuditLog.objects.filter(company=company).order_by('-timestamp')[:30]
        return render(request, 'company/settings.html', {'form': form, 'users': users, 'logs': logs})

    def post(self, request):
        company = request.user.company
        form = CompanyForm(request.POST, request.FILES, instance=company)
        if form.is_valid():
            form.save()
            log_action(request.user, 'UPDATE_SETTINGS', 'COMPANY', company.id, request=request)
            messages.success(request, "Business Settings updated successfully!")
            return redirect('company_settings')
        users = company.users.all()
        logs = AuditLog.objects.filter(company=company).order_by('-timestamp')[:30]
        return render(request, 'company/settings.html', {'form': form, 'users': users, 'logs': logs})


# --- MASTERS: CUSTOMERS ---

class CustomerListView(CompanyRequiredMixin, CompanyQuerySetMixin, PaginationMixin, ListView):
    model = Customer
    template_name = 'company/customer_list.html'
    context_object_name = 'customers'

    def get_queryset(self):
        qs = super().get_queryset().filter(is_active=True)
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(name__icontains=search) |
                Q(business_name__icontains=search) |
                Q(gstin__icontains=search) |
                Q(mobile__icontains=search) |
                Q(email__icontains=search) |
                Q(billing_city__icontains=search) |
                Q(billing_state__icontains=search)
            )
        return qs


class CustomerCreateView(CompanyRequiredMixin, CreateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'company/customer_add.html'
    success_url = reverse_lazy('customer_list')

    def is_ajax(self):
        return (
            self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in self.request.META.get('HTTP_ACCEPT', '') or
            self.request.content_type == 'application/json'
        )

    def form_valid(self, form):
        form.instance.company = self.request.user.company
        form.instance.outstanding_balance = form.instance.opening_balance if form.instance.opening_balance_type == 'DR' else -form.instance.opening_balance
        response = super().form_valid(form)
        log_action(self.request.user, 'CREATE_CUSTOMER', 'CUSTOMER', self.object.id, new_values=form.cleaned_data, request=self.request)
        messages.success(self.request, "Customer created successfully!")
        if self.is_ajax():
            return JsonResponse({
                'success': True,
                'status': 'success',
                'message': "Customer created successfully!",
                'customer_id': self.object.id,
                'redirect_url': str(self.success_url)
            }, encoder=DjangoJSONEncoder)
        return response

    def form_invalid(self, form):
        if self.is_ajax():
            errors = [f"{field.capitalize()}: {errs[0]}" for field, errs in form.errors.items()]
            return JsonResponse({'success': False, 'status': 'error', 'message': ', '.join(errors)}, status=400, encoder=DjangoJSONEncoder)
        return super().form_invalid(form)


class CustomerUpdateView(CompanyRequiredMixin, CompanyQuerySetMixin, UpdateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'company/customer_add.html'
    success_url = reverse_lazy('customer_list')

    def is_ajax(self):
        return (
            self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in self.request.META.get('HTTP_ACCEPT', '') or
            self.request.content_type == 'application/json'
        )

    def form_valid(self, form):
        response = super().form_valid(form)
        log_action(self.request.user, 'UPDATE_CUSTOMER', 'CUSTOMER', self.object.id, new_values=form.cleaned_data, request=self.request)
        messages.success(self.request, "Customer details updated successfully!")
        if self.is_ajax():
            return JsonResponse({
                'success': True,
                'status': 'success',
                'message': "Customer details updated successfully!",
                'customer_id': self.object.id,
                'redirect_url': str(self.success_url)
            }, encoder=DjangoJSONEncoder)
        return response

    def form_invalid(self, form):
        if self.is_ajax():
            errors = [f"{field.capitalize()}: {errs[0]}" for field, errs in form.errors.items()]
            return JsonResponse({'success': False, 'status': 'error', 'message': ', '.join(errors)}, status=400, encoder=DjangoJSONEncoder)
        return super().form_invalid(form)


class CustomerDetailView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = Customer
    template_name = 'company/customer_detail.html'
    context_object_name = 'customer'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        customer = self.get_object()
        context['invoices'] = Invoice.objects.filter(customer=customer).order_by('-invoice_date')
        context['payments'] = Payment.objects.filter(customer=customer).order_by('-payment_date')
        context['ledger_entries'] = customer.ledger_entries.all().order_by('-date', '-id')
        return context


def customer_search_api(request):
    if not request.user.is_authenticated or not request.user.company:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
        
    company = request.user.company
    q = request.GET.get('q', '').strip()
    
    customers = Customer.objects.filter(company=company, is_active=True)
    if q:
        customers = customers.filter(
            Q(name__icontains=q) | 
            Q(business_name__icontains=q) | 
            Q(gstin__icontains=q) |
            Q(mobile__icontains=q)
        )
    
    customers = customers.order_by('name')[:20]
    
    data = []
    for c in customers:
        data.append({
            'id': c.id,
            'name': c.name,
            'balance': str(c.outstanding_balance),
            'business_name': c.business_name or '',
            'customer_type': c.customer_type,
            'gstin': c.gstin or '',
            'mobile': c.mobile or '',
            'email': c.email or '',
            'billing_address': c.billing_address or '',
            'billing_city': c.billing_city or '',
            'billing_state': c.billing_state or '',
            'billing_state_code': c.billing_state_code or '',
            'billing_pincode': c.billing_pincode or '',
        })
        
    return JsonResponse({'status': 'success', 'customers': data, 'results': data})


def supplier_search_api(request):
    if not request.user.is_authenticated or not request.user.company:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
        
    company = request.user.company
    q = request.GET.get('q', '').strip()
    
    suppliers = Supplier.objects.filter(company=company, is_active=True)
    if q:
        suppliers = suppliers.filter(
            Q(name__icontains=q) | 
            Q(business_name__icontains=q) | 
            Q(gstin__icontains=q) |
            Q(mobile__icontains=q)
        )
    
    suppliers = suppliers.order_by('name')[:20]
    
    data = []
    for s in suppliers:
        data.append({
            'id': s.id,
            'name': s.name,
            'balance': str(s.outstanding_balance),
            'business_name': s.business_name or '',
            'supplier_type': s.supplier_type,
            'gstin': s.gstin or '',
            'mobile': s.mobile or '',
            'email': s.email or '',
            'address': s.address or '',
            'city': s.city or '',
            'state': s.state or '',
            'state_code': s.state_code or '',
            'pincode': s.pincode or '',
        })
        
    return JsonResponse({'status': 'success', 'suppliers': data, 'results': data})


def hsn_sac_search_api(request):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
        
    company = getattr(request.user, 'company', None)
    q = request.GET.get('q', '').strip()
    
    qs = HSNSACMaster.objects.filter(is_active=True)
    if company:
        qs = qs.filter(Q(company=company) | Q(company__isnull=True))
    else:
        qs = qs.filter(company__isnull=True)
        
    if q:
        qs = qs.filter(Q(code__icontains=q) | Q(description__icontains=q))
        
    records = qs.order_by('code')[:25]
    
    data = []
    for h in records:
        data.append({
            'id': h.id,
            'code': h.code,
            'description': h.description,
            'type': h.type,
            'gst_rate': float(h.gst_rate),
            'cgst_rate': float(h.get_cgst_rate()),
            'sgst_rate': float(h.get_sgst_rate()),
            'igst_rate': float(h.get_igst_rate()),
            'cess_rate': float(h.cess_rate),
            'uqc': h.uqc or 'PCS',
            'category': h.category or '',
            'sub_category': h.sub_category or ''
        })
        
    return JsonResponse({'status': 'success', 'hsn_sac_list': data})


def category_search_api(request):
    if not request.user.is_authenticated or not request.user.company:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
        
    company = request.user.company
    q = request.GET.get('q', '').strip()
    
    categories = Category.objects.filter(company=company)
    if q:
        categories = categories.filter(name__icontains=q)
    
    categories = categories.order_by('name')[:15]
    
    data = [{'id': c.id, 'name': c.name} for c in categories]
    return JsonResponse({'status': 'success', 'results': data, 'categories': data})


def category_quick_add(request):
    if not request.user.is_authenticated or not request.user.company:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
        
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST request required.'}, status=405)
        
    company = request.user.company
    name = request.POST.get('name', '').strip()
    if not name:
        return JsonResponse({'status': 'error', 'message': 'Category name is required.'}, status=400)
        
    existing = Category.objects.filter(company=company, name__iexact=name).first()
    if existing:
        return JsonResponse({
            'status': 'success',
            'success': True,
            'id': existing.id,
            'name': existing.name,
            'created': False,
            'message': 'Category already exists.'
        })
        
    try:
        category = Category.objects.create(company=company, name=name)
        log_action(request.user, 'CREATE_CATEGORY', 'CATEGORY', category.id, new_values={'name': name}, request=request)
        return JsonResponse({
            'status': 'success',
            'success': True,
            'id': category.id,
            'name': category.name,
            'created': True,
            'message': f"Category '{name}' created successfully."
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'success': False, 'message': f"Unable to create category: {str(e)}"}, status=400)


def brand_search_api(request):
    if not request.user.is_authenticated or not request.user.company:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
        
    company = request.user.company
    q = request.GET.get('q', '').strip()
    
    brands = Brand.objects.filter(company=company)
    if q:
        brands = brands.filter(name__icontains=q)
    
    brands = brands.order_by('name')[:15]
    
    data = [{'id': b.id, 'name': b.name} for b in brands]
    return JsonResponse({'status': 'success', 'results': data, 'brands': data})


def brand_quick_add(request):
    if not request.user.is_authenticated or not request.user.company:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
        
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST request required.'}, status=405)
        
    company = request.user.company
    name = request.POST.get('name', '').strip()
    if not name:
        return JsonResponse({'status': 'error', 'message': 'Brand name is required.'}, status=400)
        
    existing = Brand.objects.filter(company=company, name__iexact=name).first()
    if existing:
        return JsonResponse({
            'status': 'success',
            'success': True,
            'id': existing.id,
            'name': existing.name,
            'created': False,
            'message': 'Brand already exists.'
        })
        
    try:
        brand = Brand.objects.create(company=company, name=name)
        log_action(request.user, 'CREATE_BRAND', 'BRAND', brand.id, new_values={'name': name}, request=request)
        return JsonResponse({
            'status': 'success',
            'success': True,
            'id': brand.id,
            'name': brand.name,
            'created': True,
            'message': f"Brand '{name}' created successfully."
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'success': False, 'message': f"Unable to create brand: {str(e)}"}, status=400)


def indian_states_search_api(request):
    from .utils import INDIAN_STATES_AND_UTS
    q = request.GET.get('q', '').strip().lower()
    results = []
    for item in INDIAN_STATES_AND_UTS:
        display = f"{item['name']} — {item['code']}"
        if not q or q in item['name'].lower() or q in item['code'] or q in display.lower():
            results.append({
                'code': item['code'],
                'name': item['name'],
                'display': display
            })
    return JsonResponse({'status': 'success', 'results': results})



class GSTCalculateView(CompanyRequiredMixin, View):
    def post(self, request):
        company = request.user.company
        try:
            data = json.loads(request.body)
            hsn_sac = data.get('hsn_sac')
            taxable_value = data.get('taxable_value', 0)
            place_of_supply_code = data.get('place_of_supply_code')
            
            if not hsn_sac:
                return JsonResponse({'status': 'error', 'message': 'Please select HSN/SAC Code.'}, status=400)
            if not place_of_supply_code:
                return JsonResponse({'status': 'error', 'message': 'Please select a valid Place of Supply.'}, status=400)
                
            res = calculate_gst(
                hsn_sac_code=hsn_sac,
                taxable_value=taxable_value,
                supplier_state_code=company.state_code,
                place_of_supply_code=place_of_supply_code,
                company=company
            )
            return JsonResponse({'status': 'success', **res})
        except ValueError as ve:
            return JsonResponse({'status': 'error', 'message': str(ve)}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@transaction.atomic
def customer_quick_add(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST method required.'}, status=405)
        
    if not request.user.is_authenticated or not request.user.company:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
        
    company = request.user.company
    
    try:
        data = json.loads(request.body)
    except Exception:
        data = request.POST
        
    name = (data.get('name') or data.get('customer_name') or '').strip()
    if not name:
        return JsonResponse({'status': 'error', 'message': 'Customer name is required.'}, status=400)
        
    force = data.get('force', False)
    
    existing = Customer.objects.filter(company=company, name__iexact=name).first()
    if existing and not force:
        return JsonResponse({
            'success': False,
            'status': 'duplicate',
            'message': f"A customer with the name '{existing.name}' already exists.",
            'customer': {
                'id': existing.id,
                'name': existing.name,
                'business_name': existing.business_name or '',
                'customer_type': existing.customer_type,
                'gstin': existing.gstin or '',
                'mobile': existing.mobile or '',
                'email': existing.email or '',
                'billing_address': existing.billing_address or '',
                'billing_city': existing.billing_city or '',
                'billing_state': existing.billing_state or '',
                'billing_state_code': existing.billing_state_code or '',
                'billing_pincode': existing.billing_pincode or '',
            }
        })
        
    state_map = {
        '01': 'Jammu & Kashmir', '02': 'Himachal Pradesh', '03': 'Punjab', '04': 'Chandigarh',
        '05': 'Uttarakhand', '06': 'Haryana', '07': 'Delhi', '08': 'Rajasthan',
        '09': 'Uttar Pradesh', '10': 'Bihar', '11': 'Sikkim', '12': 'Arunachal Pradesh',
        '13': 'Nagaland', '14': 'Manipur', '15': 'Mizoram', '16': 'Tripura',
        '17': 'Meghalaya', '18': 'Assam', '19': 'West Bengal', '20': 'Jharkhand',
        '21': 'Odisha', '22': 'Chhattisgarh', '23': 'Madhya Pradesh', '24': 'Gujarat',
        '26': 'Dadra & Nagar Haveli and Daman & Diu', '27': 'Maharashtra', '28': 'Andhra Pradesh (Old)',
        '29': 'Karnataka', '30': 'Goa', '31': 'Lakshadweep', '32': 'Kerala',
        '33': 'Tamil Nadu', '34': 'Puducherry', '35': 'Andaman & Nicobar Islands',
        '36': 'Telangana', '37': 'Andhra Pradesh', '38': 'Ladakh', '97': 'Other Territory',
        '99': 'Centre Jurisdiction'
    }

    gstin = (data.get('gstin') or '').strip().upper()
    state = (data.get('billing_state') or company.state or 'Maharashtra').strip()
    state_code = (data.get('billing_state_code') or company.state_code or '27').strip().zfill(2)

    # In India, GSTIN prefix defines the state code automatically
    if gstin and len(gstin) >= 2 and gstin[:2].isdigit():
        derived_code = gstin[:2]
        if derived_code in state_map:
            state_code = derived_code
            state = state_map[derived_code]
        else:
            state_code = derived_code
            
    customer_type = data.get('customer_type', 'REGISTERED' if gstin else 'UNREGISTERED')
    if customer_type not in ['REGISTERED', 'COMPOSITION', 'UNREGISTERED', 'CONSUMER', 'SEZ', 'EXPORT']:
        customer_type = 'REGISTERED' if gstin else 'UNREGISTERED'

    billing_addr = (data.get('billing_address') or state).strip()
    billing_city = (data.get('billing_city') or company.city or state).strip()
    billing_pincode = (data.get('billing_pincode') or company.pincode or '400001').strip()
    mobile = (data.get('mobile') or '').strip()

    customer = Customer.objects.create(
        company=company,
        name=name,
        business_name=(data.get('business_name') or '').strip(),
        customer_type=customer_type,
        gstin=gstin,
        mobile=mobile,
        email=(data.get('email') or '').strip(),
        billing_address=billing_addr,
        billing_city=billing_city,
        billing_state=state,
        billing_state_code=state_code,
        billing_pincode=billing_pincode,
    )
    
    log_action(request.user, 'CREATE_CUSTOMER', 'CUSTOMER', customer.id, new_values={'name': name}, request=request)
    
    return JsonResponse({
        'success': True,
        'status': 'success',
        'message': f"Customer '{customer.name}' created successfully.",
        'customer': {
            'id': customer.id,
            'name': customer.name,
            'business_name': customer.business_name or '',
            'customer_type': customer.customer_type,
            'gstin': customer.gstin or '',
            'mobile': customer.mobile or '',
            'email': customer.email or '',
            'billing_address': customer.billing_address or '',
            'billing_city': customer.billing_city or '',
            'billing_state': customer.billing_state or '',
            'billing_state_code': customer.billing_state_code or '',
            'billing_pincode': customer.billing_pincode or '',
        },
        'data': {
            'id': customer.id,
            'name': customer.name,
            'billing_state_code': customer.billing_state_code or '27'
        }
    })


@transaction.atomic
def supplier_quick_add(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'status': 'error', 'message': 'POST method required.'}, status=405)
        
    if not request.user.is_authenticated or not request.user.company:
        return JsonResponse({'success': False, 'status': 'error', 'message': 'Authentication required.'}, status=403)
        
    company = request.user.company
    
    try:
        data = json.loads(request.body)
    except Exception:
        data = request.POST
        
    name = (data.get('name') or '').strip()
    if not name:
        return JsonResponse({'success': False, 'status': 'error', 'message': 'Supplier name is required.'}, status=400)
        
    force = data.get('force', False)
    
    existing = Supplier.objects.filter(company=company, name__iexact=name).first()
    if existing and not force:
        return JsonResponse({
            'success': False,
            'status': 'duplicate',
            'message': 'Supplier already exists.',
            'supplier': {
                'id': existing.id,
                'name': existing.name,
                'business_name': existing.business_name or '',
                'supplier_type': existing.supplier_type,
                'gstin': existing.gstin or '',
                'mobile': existing.mobile or '',
                'email': existing.email or '',
                'address': existing.address or '',
                'city': existing.city or '',
                'state': existing.state or '',
                'state_code': existing.state_code or '',
                'pincode': existing.pincode or '',
            }
        })
        
    state_map = {
        '01': 'Jammu & Kashmir', '02': 'Himachal Pradesh', '03': 'Punjab', '04': 'Chandigarh',
        '05': 'Uttarakhand', '06': 'Haryana', '07': 'Delhi', '08': 'Rajasthan',
        '09': 'Uttar Pradesh', '10': 'Bihar', '11': 'Sikkim', '12': 'Arunachal Pradesh',
        '13': 'Nagaland', '14': 'Manipur', '15': 'Mizoram', '16': 'Tripura',
        '17': 'Meghalaya', '18': 'Assam', '19': 'West Bengal', '20': 'Jharkhand',
        '21': 'Odisha', '22': 'Chhattisgarh', '23': 'Madhya Pradesh', '24': 'Gujarat',
        '26': 'Dadra & Nagar Haveli and Daman & Diu', '27': 'Maharashtra', '28': 'Andhra Pradesh (Old)',
        '29': 'Karnataka', '30': 'Goa', '31': 'Lakshadweep', '32': 'Kerala',
        '33': 'Tamil Nadu', '34': 'Puducherry', '35': 'Andaman & Nicobar Islands',
        '36': 'Telangana', '37': 'Andhra Pradesh', '38': 'Ladakh', '97': 'Other Territory',
        '99': 'Centre Jurisdiction'
    }

    gstin = (data.get('gstin') or '').strip().upper()
    state = (data.get('state') or company.state or 'Maharashtra').strip()
    state_code = (data.get('state_code') or company.state_code or '27').strip().zfill(2)

    if gstin and len(gstin) >= 2 and gstin[:2].isdigit():
        derived_code = gstin[:2]
        if derived_code in state_map:
            state_code = derived_code
            state = state_map[derived_code]
        else:
            state_code = derived_code

    supplier_type = data.get('supplier_type', 'REGISTERED' if gstin else 'UNREGISTERED')
    if supplier_type not in ['REGISTERED', 'COMPOSITION', 'UNREGISTERED', 'SEZ', 'EXPORT']:
        supplier_type = 'REGISTERED' if gstin else 'UNREGISTERED'

    address = (data.get('address') or state).strip()
    city = (data.get('city') or company.city or state).strip()
    pincode = (data.get('pincode') or company.pincode or '400001').strip()
    mobile = (data.get('mobile') or '').strip()

    supplier = Supplier.objects.create(
        company=company,
        name=name,
        business_name=(data.get('business_name') or '').strip(),
        supplier_type=supplier_type,
        gstin=gstin,
        mobile=mobile,
        email=(data.get('email') or '').strip(),
        address=address,
        city=city,
        state=state,
        state_code=state_code,
        pincode=pincode,
    )
    
    log_action(request.user, 'CREATE_SUPPLIER', 'SUPPLIER', supplier.id, new_values={'name': name}, request=request)
    
    return JsonResponse({
        'success': True,
        'status': 'success',
        'message': f"Supplier '{supplier.name}' created successfully.",
        'supplier': {
            'id': supplier.id,
            'name': supplier.name,
            'business_name': supplier.business_name or '',
            'supplier_type': supplier.supplier_type,
            'gstin': supplier.gstin or '',
            'mobile': supplier.mobile or '',
            'email': supplier.email or '',
            'address': supplier.address or '',
            'city': supplier.city or '',
            'state': supplier.state or '',
            'state_code': supplier.state_code or '',
            'pincode': supplier.pincode or '',
        },
        'data': {
            'id': supplier.id,
            'name': supplier.name,
            'state_code': supplier.state_code or '27'
        }
    })



# --- MASTERS: SUPPLIERS ---

class SupplierListView(CompanyRequiredMixin, CompanyQuerySetMixin, PaginationMixin, ListView):
    model = Supplier
    template_name = 'company/supplier_list.html'
    context_object_name = 'suppliers'

    def get_queryset(self):
        qs = super().get_queryset().filter(is_active=True)
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(name__icontains=search) |
                Q(business_name__icontains=search) |
                Q(gstin__icontains=search) |
                Q(mobile__icontains=search) |
                Q(email__icontains=search) |
                Q(city__icontains=search) |
                Q(state__icontains=search)
            )
        return qs


class SupplierCreateView(CompanyRequiredMixin, CreateView):
    model = Supplier
    form_class = SupplierForm
    template_name = 'company/supplier_add.html'
    success_url = reverse_lazy('supplier_list')

    def is_ajax(self):
        return (
            self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in self.request.META.get('HTTP_ACCEPT', '') or
            self.request.content_type == 'application/json'
        )

    def form_valid(self, form):
        form.instance.company = self.request.user.company
        form.instance.outstanding_balance = form.instance.opening_balance if form.instance.opening_balance_type == 'CR' else -form.instance.opening_balance
        response = super().form_valid(form)
        log_action(self.request.user, 'CREATE_SUPPLIER', 'SUPPLIER', self.object.id, new_values=form.cleaned_data, request=self.request)
        messages.success(self.request, "Supplier created successfully!")
        if self.is_ajax():
            return JsonResponse({
                'success': True,
                'status': 'success',
                'message': "Supplier created successfully!",
                'supplier_id': self.object.id,
                'redirect_url': str(self.success_url)
            }, encoder=DjangoJSONEncoder)
        return response

    def form_invalid(self, form):
        if self.is_ajax():
            errors = [f"{field.capitalize()}: {errs[0]}" for field, errs in form.errors.items()]
            return JsonResponse({'success': False, 'status': 'error', 'message': ', '.join(errors)}, status=400, encoder=DjangoJSONEncoder)
        return super().form_invalid(form)


class SupplierUpdateView(CompanyRequiredMixin, CompanyQuerySetMixin, UpdateView):
    model = Supplier
    form_class = SupplierForm
    template_name = 'company/supplier_add.html'
    success_url = reverse_lazy('supplier_list')

    def is_ajax(self):
        return (
            self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in self.request.META.get('HTTP_ACCEPT', '') or
            self.request.content_type == 'application/json'
        )

    def form_valid(self, form):
        response = super().form_valid(form)
        log_action(self.request.user, 'UPDATE_SUPPLIER', 'SUPPLIER', self.object.id, new_values=form.cleaned_data, request=self.request)
        messages.success(self.request, "Supplier details updated successfully!")
        if self.is_ajax():
            return JsonResponse({
                'success': True,
                'status': 'success',
                'message': "Supplier details updated successfully!",
                'supplier_id': self.object.id,
                'redirect_url': str(self.success_url)
            }, encoder=DjangoJSONEncoder)
        return response

    def form_invalid(self, form):
        if self.is_ajax():
            errors = [f"{field.capitalize()}: {errs[0]}" for field, errs in form.errors.items()]
            return JsonResponse({'success': False, 'status': 'error', 'message': ', '.join(errors)}, status=400, encoder=DjangoJSONEncoder)
        return super().form_invalid(form)


class SupplierDetailView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = Supplier
    template_name = 'company/supplier_detail.html'
    context_object_name = 'supplier'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        supplier = self.get_object()
        context['bills'] = PurchaseBill.objects.filter(supplier=supplier).order_by('-bill_date')
        context['payments'] = Payment.objects.filter(supplier=supplier).order_by('-payment_date')
        context['ledger_entries'] = supplier.ledger_entries.all().order_by('-date', '-id')
        return context


# --- MASTERS: PRODUCTS ---

class ProductListView(CompanyRequiredMixin, CompanyQuerySetMixin, PaginationMixin, ListView):
    model = Product
    template_name = 'company/product_list.html'
    context_object_name = 'products'

    def get_queryset(self):
        qs = super().get_queryset().filter(is_active=True)
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(name__icontains=search) |
                Q(sku__icontains=search) |
                Q(barcode__icontains=search) |
                Q(hsn_sac__code__icontains=search) |
                Q(category__name__icontains=search) |
                Q(brand__name__icontains=search)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import Warehouse
        context['warehouses'] = Warehouse.objects.filter(company=self.request.user.company, is_active=True)
        return context


class ProductCreateView(CompanyRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'company/product_add.html'
    success_url = reverse_lazy('product_list')

    def is_ajax(self):
        return (
            self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in self.request.META.get('HTTP_ACCEPT', '') or
            self.request.content_type == 'application/json'
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.user.company
        return kwargs

    @transaction.atomic
    def form_valid(self, form):
        form.instance.company = self.request.user.company
        response = super().form_valid(form)
        
        # Log opening stock movement if quantity > 0
        qty = self.request.POST.get('opening_stock', 0)
        if Decimal(qty) > 0:
            wh = Warehouse.objects.filter(company=self.request.user.company, is_active=True).first()
            if wh:
                StockMovement.objects.create(
                    company=self.request.user.company,
                    product=self.object,
                    warehouse=wh,
                    quantity=Decimal(qty),
                    movement_type='OPENING',
                    reference_no="OPENING STOCK",
                    created_by=self.request.user
                )
                update_product_stock(self.object.id)
                
        log_action(self.request.user, 'CREATE_PRODUCT', 'PRODUCT', self.object.id, new_values=form.cleaned_data, request=self.request)
        messages.success(self.request, "Product created successfully!")
        if self.is_ajax():
            return JsonResponse({
                'success': True,
                'status': 'success',
                'message': "Product created successfully!",
                'product_id': self.object.id,
                'redirect_url': str(self.success_url)
            }, encoder=DjangoJSONEncoder)
        return response

    def form_invalid(self, form):
        if self.is_ajax():
            errors = [f"{field.capitalize()}: {errs[0]}" for field, errs in form.errors.items()]
            return JsonResponse({'success': False, 'status': 'error', 'message': ', '.join(errors)}, status=400, encoder=DjangoJSONEncoder)
        return super().form_invalid(form)


class ProductUpdateView(CompanyRequiredMixin, CompanyQuerySetMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'company/product_add.html'
    success_url = reverse_lazy('product_list')

    def is_ajax(self):
        return (
            self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in self.request.META.get('HTTP_ACCEPT', '') or
            self.request.content_type == 'application/json'
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.user.company
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        log_action(self.request.user, 'UPDATE_PRODUCT', 'PRODUCT', self.object.id, new_values=form.cleaned_data, request=self.request)
        messages.success(self.request, "Product details updated successfully!")
        if self.is_ajax():
            return JsonResponse({
                'success': True,
                'status': 'success',
                'message': "Product details updated successfully!",
                'product_id': self.object.id,
                'redirect_url': str(self.success_url)
            }, encoder=DjangoJSONEncoder)
        return response

    def form_invalid(self, form):
        if self.is_ajax():
            errors = [f"{field.capitalize()}: {errs[0]}" for field, errs in form.errors.items()]
            return JsonResponse({'success': False, 'status': 'error', 'message': ', '.join(errors)}, status=400, encoder=DjangoJSONEncoder)
        return super().form_invalid(form)


class ProductDetailView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = Product
    template_name = 'company/product_detail.html'
    context_object_name = 'product'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        product = self.get_object()

        # Stock Movements history
        context['stock_movements'] = StockMovement.objects.filter(
            product=product, company=self.request.user.company
        ).select_related('warehouse', 'created_by').order_by('-created_at')[:50]

        # Recent Sales history (Invoice items)
        context['recent_sales'] = InvoiceItem.objects.filter(
            product=product, invoice__company=self.request.user.company
        ).select_related('invoice', 'invoice__customer').order_by('-invoice__invoice_date', '-id')[:30]

        # Recent Purchase history (Purchase bill items)
        context['recent_purchases'] = PurchaseBillItem.objects.filter(
            product=product, purchase_bill__company=self.request.user.company
        ).select_related('purchase_bill', 'purchase_bill__supplier').order_by('-purchase_bill__bill_date', '-id')[:30]

        # Warehouses for adjustment modal
        from .models import Warehouse
        context['warehouses'] = Warehouse.objects.filter(company=self.request.user.company, is_active=True)

        # Calculated metrics (Profit margin per unit)
        if product.selling_price and product.purchase_price and product.purchase_price > Decimal('0.00'):
            profit = product.selling_price - product.purchase_price
            margin_pct = (profit / product.purchase_price) * Decimal('100.00')
        else:
            profit = Decimal('0.00')
            margin_pct = Decimal('0.00')

        context['profit_per_unit'] = profit
        context['profit_margin_pct'] = round(margin_pct, 2)

        return context


def product_duplicate(request, pk):
    company = request.user.company
    prod = get_object_or_404(Product, id=pk, company=company)
    new_prod = Product.objects.create(
        company=company, name=f"{prod.name} (Copy)", product_type=prod.product_type,
        sku=f"{prod.sku}-copy" if prod.sku else None, barcode=None,
        category=prod.category, brand=prod.brand, hsn_sac=prod.hsn_sac, unit=prod.unit,
        purchase_price=prod.purchase_price, selling_price=prod.selling_price, mrp=prod.mrp,
        wholesale_price=prod.wholesale_price, retail_price=prod.retail_price,
        min_selling_price=prod.min_selling_price, tax_inclusive=prod.tax_inclusive,
        track_inventory=prod.track_inventory, allow_negative_stock=prod.allow_negative_stock,
        min_stock=prod.min_stock, max_stock=prod.max_stock
    )
    log_action(request.user, 'DUPLICATE_PRODUCT', 'PRODUCT', new_prod.id, request=request)
    messages.success(request, f"Product duplicated successfully into '{new_prod.name}'!")
    return redirect('product_list')


@transaction.atomic
def product_adjust_stock(request, pk):
    if not request.user.is_authenticated or not request.user.company:
        return JsonResponse({'success': False, 'status': 'error', 'message': 'Authentication required.'}, status=403)

    company = request.user.company
    prod = get_object_or_404(Product, id=pk, company=company)
    if request.method == 'POST':
        try:
            data = json.loads(request.body) if request.body and request.content_type == 'application/json' else request.POST
            qty = parse_money(data.get('quantity', 0))
            reason = data.get('reason', 'ADJUSTMENT')
            wh_id = data.get('warehouse_id')
            wh = get_object_or_404(Warehouse, id=wh_id, company=company)
            
            StockMovement.objects.create(
                company=company, product=prod, warehouse=wh,
                quantity=qty, movement_type='ADJUSTMENT',
                reference_no=f"Adj: {reason}", created_by=request.user
            )
            update_product_stock(prod.id)
            prod.refresh_from_db()
                
            log_action(request.user, 'STOCK_ADJUST', 'PRODUCT', prod.id, new_values={'qty': str(qty)}, request=request)
            return JsonResponse({
                'success': True,
                'status': 'success',
                'message': f"Stock adjusted successfully! Current stock is {prod.current_stock}.",
                'current_stock': str(prod.current_stock)
            })
        except Exception as e:
            return JsonResponse({'success': False, 'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'status': 'error', 'message': 'Invalid request'}, status=400)


def product_scan_barcode(request):
    barcode = request.GET.get('barcode')
    company = request.user.company
    prod = Product.objects.filter(company=company, barcode=barcode, is_active=True).first()
    if prod:
        return JsonResponse({
            'status': 'success',
            'product': {
                'id': prod.id,
                'name': prod.name,
                'selling_price': str(prod.selling_price),
                'gst_rate': str(prod.hsn_sac.gst_rate) if prod.hsn_sac else '18.00',
                'unit': prod.unit.name if prod.unit else 'PCS'
            }
        })
    return JsonResponse({'status': 'error', 'message': 'Product not found'}, status=404)


def hsn_lookup(request):
    code = request.GET.get('code')
    company = request.user.company
    hsn = HSNSACMaster.objects.filter(Q(company=company) | Q(company__isnull=True), code=code, is_active=True).first()
    if hsn:
        return JsonResponse({
            'status': 'success',
            'hsn': {
                'id': hsn.id,
                'code': hsn.code,
                'gst_rate': str(hsn.gst_rate),
                'cess_rate': str(hsn.cess_rate)
            }
        })
    return JsonResponse({'status': 'error', 'message': 'HSN/SAC not found'}, status=404)


def check_company_permission(user):
    if not user.is_authenticated:
        return False, "Not authenticated"
    if user.role == 'SUPERADMIN':
        return False, "Superadmins cannot access Company Panel"
    if not user.company or not user.company.is_active:
        return False, "Inactive company account"
    if user.company.subscription_status == 'SUSPENDED':
        return False, "Subscription is suspended"
    return True, ""


def normalize_hsn_sac(val):
    if val is None:
        return ""
    val_str = str(val).strip()
    if not val_str:
        return ""
    if 'e' in val_str.lower():
        try:
            val_float = float(val_str)
            if val_float.is_integer():
                val_str = str(int(val_float))
        except Exception:
            pass
    elif val_str.endswith('.0') and len(val_str) > 2:
        val_str = val_str[:-2]
    return val_str


def normalize_barcode(val):
    if val is None:
        return ""
    val_str = str(val).strip()
    if not val_str:
        return ""
    if 'e' in val_str.lower():
        try:
            val_float = float(val_str)
            if val_float.is_integer():
                val_str = str(int(val_float))
        except Exception:
            pass
    elif val_str.endswith('.0') and len(val_str) > 2:
        val_str = val_str[:-2]
    return val_str


def company_product_sample_template(request):
    allowed, err_msg = check_company_permission(request.user)
    if not allowed:
        return HttpResponse(err_msg, status=403)

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Template"

    headers = [
        "Product Name", "Product Type", "SKU", "Barcode", "Category", "Brand", "Unit", "Description",
        "HSN/SAC Code", "Purchase Price", "Selling Price", "MRP", "Wholesale Price", "Retail Price",
        "Min Selling Price", "Tax Inclusive", "Track Inventory", "Allow Negative Stock", "Minimum Stock",
        "Maximum Stock", "Opening Stock", "Warehouse", "Status", "GST Rate"
    ]
    ws.append(headers)

    first_wh = Warehouse.objects.filter(company=request.user.company, is_active=True).first()
    wh_name = first_wh.name if first_wh else "Main Warehouse"

    ws.append([
        "Sample Product A", "GOODS", "SKU-001", "123456789012", "Electronics", "Sony", "PCS", "A great electronic device",
        "8517", "1000.00", "1200.00", "1500.00", "1100.00", "1200.00", "1150.00", "No", "Yes", "No", "10.00",
        "100.00", "50.00", wh_name, "Active", "18.00"
    ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Inventory_Bulk_Template.xlsx"'
    wb.save(response)
    return response


def company_product_bulk_preview(request):
    allowed, err_msg = check_company_permission(request.user)
    if not allowed:
        return JsonResponse({'status': 'error', 'message': err_msg}, status=403)

    if request.method != 'POST' or 'file' not in request.FILES:
        return JsonResponse({'status': 'error', 'message': 'Please upload a valid CSV or Excel (.xlsx) file.'}, status=400)

    uploaded_file = request.FILES['file']
    filename = uploaded_file.name.lower()
    company = request.user.company

    rows_data = []

    try:
        if filename.endswith('.csv'):
            import csv
            file_data = uploaded_file.read().decode('utf-8-sig', errors='ignore').splitlines()
            reader = csv.reader(file_data)
            header = None
            for row in reader:
                if not row or not any(row):
                    continue
                if header is None:
                    header = [c.strip().lower() for c in row]
                    continue
                rows_data.append(dict(zip(header, [c.strip() for c in row])))
        elif filename.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            sheet = wb.active
            header = None
            for row in sheet.iter_rows(values_only=True):
                if not row or not any(row):
                    continue
                row_vals = [str(c).strip() if c is not None else '' for c in row]
                if header is None:
                    header = [c.lower() for c in row_vals]
                    continue
                rows_data.append(dict(zip(header, row_vals)))
        else:
            return JsonResponse({'status': 'error', 'message': 'Unsupported file format. Please upload a .csv or .xlsx file.'}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f"Error parsing file: {str(e)}"}, status=400)

    existing_products = Product.objects.filter(company=company)
    sku_to_prod = {p.sku.strip().lower(): p for p in existing_products if p.sku}
    barcode_to_prod = {p.barcode.strip().lower(): p for p in existing_products if p.barcode}
    name_to_prod = {p.name.strip().lower(): p for p in existing_products if p.name}

    seen_file_skus = {}
    seen_file_barcodes = {}
    seen_file_names = {}

    parsed_rows = []
    valid_count = 0
    invalid_count = 0
    new_count = 0
    update_count = 0
    duplicate_count = 0

    for i, raw in enumerate(rows_data, 1):
        def get_val(keys, default=''):
            for k in keys:
                for header_key in raw:
                    if k == header_key.lower() or k in header_key.lower():
                        return raw[header_key]
            return default

        name = get_val(['product name', 'product_name', 'name', 'title']).strip()
        product_type = get_val(['product type', 'product_type', 'type']).strip().upper()
        if not product_type: product_type = 'GOODS'
        if 'SERV' in product_type: product_type = 'SERVICES'
        else: product_type = 'GOODS'

        sku = get_val(['sku', 'sku code']).strip()
        barcode = get_val(['barcode', 'barcode code', 'product code', 'product_code']).strip()
        category = get_val(['category', 'category name']).strip()
        brand = get_val(['brand', 'brand name']).strip()
        unit = get_val(['unit', 'unit name']).strip() or 'PCS'
        description = get_val(['description', 'desc', 'details']).strip()
        hsn_sac = get_val(['hsn/sac code', 'hsn/sac', 'hsn_sac code', 'hsn/sac code', 'hsn code', 'sac code', 'hsn_sac']).strip()
        
        purchase_price_str = get_val(['purchase price', 'purchase_price']).strip()
        selling_price_str = get_val(['selling price', 'selling_price']).strip()
        mrp_str = get_val(['mrp']).strip()
        wholesale_price_str = get_val(['wholesale price', 'wholesale_price']).strip()
        retail_price_str = get_val(['retail price', 'retail_price']).strip()
        min_selling_price_str = get_val(['min selling price', 'min_selling_price', 'minimum selling price']).strip()
        
        tax_inclusive_str = get_val(['tax inclusive', 'tax_inclusive']).strip().lower()
        track_inventory_str = get_val(['track inventory', 'track_inventory']).strip().lower()
        allow_negative_stock_str = get_val(['allow negative stock', 'allow_negative_stock']).strip().lower()
        
        min_stock_str = get_val(['minimum stock', 'min_stock', 'min stock']).strip()
        max_stock_str = get_val(['maximum stock', 'max_stock', 'max stock']).strip()
        
        opening_stock_str = get_val(['opening stock', 'opening_stock', 'quantity', 'qty']).strip()
        warehouse_str = get_val(['warehouse', 'warehouse name', 'warehouse code']).strip()
        
        status_str = get_val(['status', 'active', 'is_active']).strip().lower()
        gst_rate_str = get_val(['gst rate', 'gst %', 'gst']).replace('%', '').strip()

        errors = []

        if not name:
            errors.append("Product Name is required")
        elif len(name) > 200:
            errors.append("Product Name cannot exceed 200 characters")

        def parse_decimal(val_str, field_name, default_val=Decimal('0.00')):
            if not val_str:
                return default_val, None
            try:
                val = Decimal(val_str)
                if val < 0:
                    return default_val, f"{field_name} cannot be negative"
                return val, None
            except Exception:
                return default_val, f"Invalid {field_name} format"

        purchase_price, err = parse_decimal(purchase_price_str, "Purchase Price")
        if err: errors.append(err)

        selling_price, err = parse_decimal(selling_price_str, "Selling Price")
        if err: errors.append(err)

        mrp, err = parse_decimal(mrp_str, "MRP")
        if err: errors.append(err)

        wholesale_price, err = parse_decimal(wholesale_price_str, "Wholesale Price")
        if err: errors.append(err)

        retail_price, err = parse_decimal(retail_price_str, "Retail Price")
        if err: errors.append(err)

        min_selling_price, err = parse_decimal(min_selling_price_str, "Min Selling Price")
        if err: errors.append(err)

        min_stock, err = parse_decimal(min_stock_str, "Minimum Stock")
        if err: errors.append(err)

        max_stock, err = parse_decimal(max_stock_str, "Maximum Stock")
        if err: errors.append(err)

        opening_stock, err = parse_decimal(opening_stock_str, "Opening Stock")
        if err: errors.append(err)

        def parse_bool(val_str, default_val=False):
            if not val_str:
                return default_val
            return val_str in ['yes', 'true', '1', 'active', 'y']

        tax_inclusive = parse_bool(tax_inclusive_str, False)
        track_inventory = parse_bool(track_inventory_str, True)
        allow_negative_stock = parse_bool(allow_negative_stock_str, False)
        is_active = status_str not in ['inactive', '0', 'false', 'disabled']

        hsn_sac = normalize_hsn_sac(hsn_sac)
        hsn_sac_id = None
        gst_rate = None
        if hsn_sac:
            hsn_obj = HSNSACMaster.objects.filter(
                Q(company=company) | Q(company__isnull=True),
                code=hsn_sac,
                is_active=True
            ).first()
            if not hsn_obj:
                errors.append(f"HSN/SAC Code '{hsn_sac}' does not exist")
            else:
                hsn_sac_id = hsn_obj.id
                gst_rate = hsn_obj.gst_rate
                if gst_rate_str:
                    try:
                        excel_gst = Decimal(gst_rate_str)
                        if abs(excel_gst - gst_rate) > Decimal('0.01'):
                            errors.append(f"GST Rate conflict: Excel says {excel_gst}%, but Master HSN/SAC has {gst_rate}%")
                    except Exception:
                        errors.append("Invalid GST Rate format in Excel")

        warehouse_id = None
        if opening_stock > 0 and track_inventory:
            if warehouse_str:
                wh_obj = Warehouse.objects.filter(
                    Q(name__iexact=warehouse_str) | Q(code__iexact=warehouse_str),
                    company=company,
                    is_active=True
                ).first()
                if not wh_obj:
                    errors.append(f"Warehouse '{warehouse_str}' not found or inactive")
                else:
                    warehouse_id = wh_obj.id
            else:
                wh_obj = Warehouse.objects.filter(company=company, is_active=True).first()
                if not wh_obj:
                    errors.append("No active warehouse exists to receive opening stock")
                else:
                    warehouse_id = wh_obj.id

        barcode = normalize_barcode(barcode)
        
        is_barcode_duplicate = False
        is_barcode_conflict = False
        barcode_error_msg = ""
        
        if barcode:
            # File-level barcode checks
            if barcode.lower() in seen_file_barcodes:
                prev = seen_file_barcodes[barcode.lower()]
                is_same_product = False
                if sku and prev['sku']:
                    is_same_product = (sku.lower() == prev['sku'].lower())
                elif not sku and not prev['sku']:
                    is_same_product = (name.lower() == prev['name'].lower())
                
                if is_same_product:
                    is_barcode_duplicate = True
                    barcode_error_msg = "Same barcode in uploaded file"
                else:
                    is_barcode_conflict = True
                    barcode_error_msg = f"Barcode belongs to another product in row {prev['row_num']}"
            else:
                seen_file_barcodes[barcode.lower()] = {'sku': sku, 'name': name, 'row_num': i}

            # DB-level barcode checks (only if not already a conflict)
            if not is_barcode_conflict:
                if barcode.lower() in barcode_to_prod:
                    db_prod = barcode_to_prod[barcode.lower()]
                    is_same_product = False
                    if sku and db_prod.sku:
                        is_same_product = (sku.lower() == db_prod.sku.lower())
                    elif not sku and not db_prod.sku:
                        is_same_product = (name.lower() == db_prod.name.lower())
                    
                    if not is_same_product:
                        is_barcode_conflict = True
                        barcode_error_msg = f"Barcode conflict: '{barcode}' is already assigned to another product '{db_prod.name}'"

        is_sku_duplicate = False
        if sku:
            if sku.lower() in seen_file_skus:
                is_sku_duplicate = True
                prev = seen_file_skus[sku.lower()]
                barcode_error_msg = "Same SKU in uploaded file"
            else:
                seen_file_skus[sku.lower()] = {'name': name, 'row_num': i}

        is_name_duplicate = False
        if not sku and not barcode and name:
            if name.lower() in seen_file_names:
                is_name_duplicate = True
                prev = seen_file_names[name.lower()]
                barcode_error_msg = "Same product name in uploaded file"
            else:
                seen_file_names[name.lower()] = {'row_num': i}

        # Check DB duplicates for NEW vs UPDATE mapping
        is_db_duplicate = False
        if len(errors) == 0 and not is_barcode_conflict:
            if sku and sku.lower() in sku_to_prod:
                is_db_duplicate = True
            elif barcode and barcode.lower() in barcode_to_prod:
                is_db_duplicate = True
            elif not sku and not barcode and name and name.lower() in name_to_prod:
                is_db_duplicate = True

        # Determine action and validity
        if is_barcode_conflict:
            errors.append(barcode_error_msg)
            is_valid = False
            action = 'CONFLICT'
        else:
            is_valid = len(errors) == 0
            if is_valid:
                if is_barcode_duplicate or is_sku_duplicate or is_name_duplicate:
                    action = 'DUPLICATE'
                    errors.append(barcode_error_msg or "Duplicate row in uploaded file")
                else:
                    action = 'UPDATE' if is_db_duplicate else 'NEW'
            else:
                action = 'ERROR'

        if is_valid:
            valid_count += 1
            if action == 'DUPLICATE':
                duplicate_count += 1
            elif action == 'UPDATE':
                update_count += 1
            elif action == 'NEW':
                new_count += 1
        else:
            invalid_count += 1

        parsed_rows.append({
            'row_num': i,
            'name': name,
            'product_type': product_type,
            'sku': sku,
            'barcode': barcode,
            'category': category,
            'brand': brand,
            'unit': unit,
            'description': description,
            'hsn_sac': hsn_sac,
            'hsn_sac_id': hsn_sac_id,
            'purchase_price': str(purchase_price),
            'selling_price': str(selling_price),
            'mrp': str(mrp),
            'wholesale_price': str(wholesale_price),
            'retail_price': str(retail_price),
            'min_selling_price': str(min_selling_price),
            'tax_inclusive': tax_inclusive,
            'track_inventory': track_inventory,
            'allow_negative_stock': allow_negative_stock,
            'min_stock': str(min_stock),
            'max_stock': str(max_stock),
            'opening_stock': str(opening_stock),
            'warehouse': warehouse_str,
            'warehouse_id': warehouse_id,
            'status': 'Active' if is_active else 'Inactive',
            'is_active': is_active,
            'gst_rate': str(gst_rate) if gst_rate is not None else (gst_rate_str or ''),
            'is_valid': is_valid,
            'errors': '; '.join(errors),
            'action': action
        })

    return JsonResponse({
        'status': 'success',
        'summary': {
            'total_rows': len(rows_data),
            'valid_rows': valid_count,
            'invalid_rows': invalid_count,
            'new_rows': new_count,
            'update_rows': update_count,
            'duplicate_rows': duplicate_count
        },
        'rows': parsed_rows
    })


@transaction.atomic
def company_product_bulk_import(request):
    allowed, err_msg = check_company_permission(request.user)
    if not allowed:
        return JsonResponse({'status': 'error', 'message': err_msg}, status=403)

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)

    try:
        data = json.loads(request.body)
        valid_rows = data.get('valid_rows', [])
        import_mode = data.get('import_mode', 'add_update')
        barcode_mode = data.get('barcode_mode', 'update_existing')
        company = request.user.company

        imported_new = 0
        updated_existing = 0
        skipped_count = 0

        categories_cache = {c.name.strip().lower(): c for c in Category.objects.filter(company=company)}
        brands_cache = {b.name.strip().lower(): b for b in Brand.objects.filter(company=company)}
        units_cache = {u.name.strip().lower(): u for u in Unit.objects.filter(company=company)}
        warehouses_cache = {str(w.id): w for w in Warehouse.objects.filter(company=company)}

        for row in valid_rows:
            action = row.get('action')
            if barcode_mode == 'skip_duplicates' and action == 'DUPLICATE':
                skipped_count += 1
                continue

            sku = row.get('sku', '').strip()
            barcode = row.get('barcode', '').strip()
            name = row.get('name', '').strip()

            product_obj = None
            if sku:
                product_obj = Product.objects.filter(company=company, sku=sku).first()
            if not product_obj and barcode:
                product_obj = Product.objects.filter(company=company, barcode=barcode).first()
            if not product_obj and name:
                product_obj = Product.objects.filter(company=company, name=name).first()

            is_duplicate = product_obj is not None

            if is_duplicate and import_mode == 'add_only':
                skipped_count += 1
                continue
            if not is_duplicate and import_mode == 'update_only':
                skipped_count += 1
                continue

            category_name = row.get('category', '').strip()
            category_obj = None
            if category_name:
                cat_lower = category_name.lower()
                if cat_lower in categories_cache:
                    category_obj = categories_cache[cat_lower]
                else:
                    category_obj = Category.objects.create(company=company, name=category_name)
                    categories_cache[cat_lower] = category_obj

            brand_name = row.get('brand', '').strip()
            brand_obj = None
            if brand_name:
                brand_lower = brand_name.lower()
                if brand_lower in brands_cache:
                    brand_obj = brands_cache[brand_lower]
                else:
                    brand_obj = Brand.objects.create(company=company, name=brand_name)
                    brands_cache[brand_lower] = brand_obj

            unit_name = row.get('unit', '').strip()
            unit_obj = None
            if unit_name:
                unit_lower = unit_name.lower()
                if unit_lower in units_cache:
                    unit_obj = units_cache[unit_lower]
                else:
                    unit_obj = Unit.objects.create(company=company, name=unit_name, code=unit_name.upper()[:10])
                    units_cache[unit_lower] = unit_obj

            hsn_sac_id = row.get('hsn_sac_id')
            hsn_sac_obj = None
            if hsn_sac_id:
                hsn_sac_obj = HSNSACMaster.objects.filter(id=hsn_sac_id).first()

            product_type = row.get('product_type', 'GOODS')
            description = row.get('description', '')
            purchase_price = Decimal(row.get('purchase_price', '0.00'))
            selling_price = Decimal(row.get('selling_price', '0.00'))
            mrp = Decimal(row.get('mrp', '0.00'))
            wholesale_price = Decimal(row.get('wholesale_price', '0.00'))
            retail_price = Decimal(row.get('retail_price', '0.00'))
            min_selling_price = Decimal(row.get('min_selling_price', '0.00'))
            tax_inclusive = bool(row.get('tax_inclusive', False))
            track_inventory = bool(row.get('track_inventory', True))
            allow_negative_stock = bool(row.get('allow_negative_stock', False))
            min_stock = Decimal(row.get('min_stock', '0.00'))
            max_stock = Decimal(row.get('max_stock', '0.00'))
            is_active = bool(row.get('is_active', True))

            if product_obj:
                product_obj.name = name
                product_obj.product_type = product_type
                product_obj.sku = sku or product_obj.sku
                product_obj.barcode = barcode or product_obj.barcode
                product_obj.category = category_obj
                product_obj.brand = brand_obj
                product_obj.unit = unit_obj
                product_obj.description = description
                product_obj.hsn_sac = hsn_sac_obj
                product_obj.purchase_price = purchase_price
                product_obj.selling_price = selling_price
                product_obj.mrp = mrp
                product_obj.wholesale_price = wholesale_price
                product_obj.retail_price = retail_price
                product_obj.min_selling_price = min_selling_price
                product_obj.tax_inclusive = tax_inclusive
                product_obj.track_inventory = track_inventory
                product_obj.allow_negative_stock = allow_negative_stock
                product_obj.min_stock = min_stock
                product_obj.max_stock = max_stock
                product_obj.is_active = is_active
                product_obj.save()
                updated_existing += 1
            else:
                new_prod = Product.objects.create(
                    company=company,
                    name=name,
                    product_type=product_type,
                    sku=sku or None,
                    barcode=barcode or None,
                    category=category_obj,
                    brand=brand_obj,
                    unit=unit_obj,
                    description=description,
                    hsn_sac=hsn_sac_obj,
                    purchase_price=purchase_price,
                    selling_price=selling_price,
                    mrp=mrp,
                    wholesale_price=wholesale_price,
                    retail_price=retail_price,
                    min_selling_price=min_selling_price,
                    tax_inclusive=tax_inclusive,
                    track_inventory=track_inventory,
                    allow_negative_stock=allow_negative_stock,
                    min_stock=min_stock,
                    max_stock=max_stock,
                    is_active=is_active
                )
                imported_new += 1

                opening_stock = Decimal(row.get('opening_stock', '0.00'))
                warehouse_id = row.get('warehouse_id')
                if opening_stock > 0 and track_inventory and warehouse_id:
                    wh_obj = warehouses_cache.get(str(warehouse_id))
                    if wh_obj:
                        StockMovement.objects.create(
                            company=company,
                            product=new_prod,
                            warehouse=wh_obj,
                            quantity=opening_stock,
                            movement_type='OPENING',
                            reference_no="OPENING STOCK",
                            created_by=request.user
                        )
                        update_product_stock(new_prod.id)

        new_values = {
            'import_mode': import_mode,
            'barcode_mode': barcode_mode,
            'total_rows': len(valid_rows),
            'inserted': imported_new,
            'updated': updated_existing,
            'skipped': skipped_count
        }
        log_action(
            user=request.user,
            action='BULK_IMPORT_PRODUCT',
            module='PRODUCT',
            record_id=0,
            new_values=new_values,
            company=company,
            request=request
        )

        messages.success(request, f"Bulk Import Completed: {imported_new} new products added, {updated_existing} existing products updated, {skipped_count} products skipped.")
        return JsonResponse({
            'status': 'success',
            'message': f"Import Completed: {imported_new} products added, {updated_existing} updated, {skipped_count} skipped."
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def company_product_download_error_report(request):
    allowed, err_msg = check_company_permission(request.user)
    if not allowed:
        return HttpResponse(err_msg, status=403)

    import csv
    if request.method != 'POST':
        return HttpResponse("Method not allowed", status=405)

    try:
        data = json.loads(request.body)
        invalid_rows = data.get('invalid_rows', [])

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="Inventory_Upload_Error_Report.csv"'

        writer = csv.writer(response)
        writer.writerow([
            "Row", "Product Name", "Product Type", "SKU", "Barcode", "Category", "Brand", "Unit", "Description",
            "HSN/SAC Code", "Purchase Price", "Selling Price", "MRP", "Wholesale Price", "Retail Price",
            "Min Selling Price", "Tax Inclusive", "Track Inventory", "Allow Negative Stock", "Minimum Stock",
            "Maximum Stock", "Opening Stock", "Warehouse", "Status", "GST Rate", "Error Message"
        ])

        for row in invalid_rows:
            writer.writerow([
                row.get('row_num', ''),
                row.get('name', ''),
                row.get('product_type', ''),
                row.get('sku', ''),
                row.get('barcode', ''),
                row.get('category', ''),
                row.get('brand', ''),
                row.get('unit', ''),
                row.get('description', ''),
                row.get('hsn_sac', ''),
                row.get('purchase_price', ''),
                row.get('selling_price', ''),
                row.get('mrp', ''),
                row.get('wholesale_price', ''),
                row.get('retail_price', ''),
                row.get('min_selling_price', ''),
                row.get('tax_inclusive', ''),
                row.get('track_inventory', ''),
                row.get('allow_negative_stock', ''),
                row.get('min_stock', ''),
                row.get('max_stock', ''),
                row.get('opening_stock', ''),
                row.get('warehouse', ''),
                row.get('status', ''),
                row.get('gst_rate', ''),
                row.get('errors', '')
            ])

        return response
    except Exception as e:
        return HttpResponse(f"Error generating report: {str(e)}", status=400)



# --- WAREHOUSES & STOCK LEDGER ---

class WarehouseListView(CompanyRequiredMixin, CompanyQuerySetMixin, PaginationMixin, ListView):
    model = Warehouse
    template_name = 'company/warehouse_list.html'
    context_object_name = 'warehouses'

    def get_queryset(self):
        qs = super().get_queryset().filter(is_active=True)
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(name__icontains=search) |
                Q(code__icontains=search) |
                Q(address__icontains=search) |
                Q(manager__icontains=search) |
                Q(contact__icontains=search)
            )
        return qs


class WarehouseCreateView(CompanyRequiredMixin, AjaxFormMixin, CreateView):
    model = Warehouse
    form_class = WarehouseForm
    template_name = 'company/warehouse_add.html'
    success_url = reverse_lazy('warehouse_list')

    def form_valid(self, form):
        form.instance.company = self.request.user.company
        response = super().form_valid(form)
        if hasattr(self, 'object') and self.object:
            log_action(self.request.user, 'CREATE_WAREHOUSE', 'WAREHOUSE', self.object.id, new_values=form.cleaned_data, request=self.request)
        return response


class WarehouseUpdateView(CompanyRequiredMixin, CompanyQuerySetMixin, AjaxFormMixin, UpdateView):
    model = Warehouse
    form_class = WarehouseForm
    template_name = 'company/warehouse_edit.html'
    success_url = reverse_lazy('warehouse_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        if hasattr(self, 'object') and self.object:
            log_action(self.request.user, 'UPDATE_WAREHOUSE', 'WAREHOUSE', self.object.id, new_values=form.cleaned_data, request=self.request)
        return response


class StockLedgerView(CompanyRequiredMixin, CompanyQuerySetMixin, PaginationMixin, ListView):
    model = StockMovement
    template_name = 'company/stock_ledger.html'
    context_object_name = 'movements'

    def get_queryset(self):
        qs = super().get_queryset().order_by('-created_at')
        product_id = self.request.GET.get('product')
        if product_id:
            qs = qs.filter(product_id=product_id)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['products'] = Product.objects.filter(company=self.request.user.company)
        return context


# --- SALES MODULE: QUOTATIONS ---

class QuotationListView(CompanyRequiredMixin, CompanyQuerySetMixin, PaginationMixin, ListView):
    model = Quotation
    template_name = 'company/quotation_list.html'
    context_object_name = 'quotations'

    def get_queryset(self):
        qs = super().get_queryset().order_by('-date', '-id')
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(quotation_number__icontains=search) |
                Q(customer__name__icontains=search) |
                Q(customer__gstin__icontains=search) |
                Q(status__icontains=search)
            )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        else:
            qs = qs.exclude(status='CANCELLED')
        return qs


class QuotationCreateView(CompanyRequiredMixin, View):
    def get(self, request):
        company = request.user.company
        customers = Customer.objects.filter(company=company, is_active=True)
        products = Product.objects.filter(company=company, is_active=True)
        
        # Generate next unique quotation number
        count = Quotation.objects.filter(company=company).count() + 1
        q_no = f"QTN-{company.financial_year}-{str(count).zfill(5)}"
        while Quotation.objects.filter(company=company, quotation_number=q_no).exists():
            count += 1
            q_no = f"QTN-{company.financial_year}-{str(count).zfill(5)}"
        
        from .utils import get_or_create_predefined_quotation_terms
        predefined_terms = get_or_create_predefined_quotation_terms(company)

        return render(request, 'company/quotation_add.html', {
            'customers': customers,
            'products': products,
            'products_json': build_products_json(products),
            'quotation_number': q_no,
            'predefined_terms': predefined_terms
        })

    def post(self, request):
        company = request.user.company
        is_ajax = (
            request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in request.META.get('HTTP_ACCEPT', '') or
            request.content_type == 'application/json'
        )

        try:
            if request.content_type == 'application/json' and request.body:
                data = json.loads(request.body)
            else:
                data = request.POST.dict()
                if 'items_json' in request.POST:
                    data['items'] = request.POST['items_json']
        except Exception:
            data = request.POST.dict()

        result = QuotationService.create_quotation(company, request.user, data)
        if result.get('success'):
            messages.success(request, result.get('message', 'Sales quotation saved successfully!'))
            if is_ajax:
                return JsonResponse(result, status=200, encoder=DjangoJSONEncoder)
            return redirect('quotation_list')
        else:
            err_msg = result.get('message', 'Unable to save sales quotation. Please check the entered details.')
            messages.error(request, err_msg)
            if is_ajax:
                return JsonResponse(result, status=400, encoder=DjangoJSONEncoder)
            customers = Customer.objects.filter(company=company, is_active=True)
            products = Product.objects.filter(company=company, is_active=True)
            from .utils import get_or_create_predefined_quotation_terms
            predefined_terms = get_or_create_predefined_quotation_terms(company)
            return render(request, 'company/quotation_add.html', {
                'customers': customers,
                'products': products,
                'products_json': build_products_json(products),
                'quotation_number': data.get('quotation_number', ''),
                'predefined_terms': predefined_terms,
                'form_data': data,
                'errors': result.get('errors', {})
            })


class QuotationUpdateView(CompanyRequiredMixin, View):
    def get(self, request, pk):
        company = request.user.company
        quotation = get_object_or_404(Quotation, id=pk, company=company)
        
        # Guard: converted quotations cannot be edited
        if quotation.status == 'CONVERTED':
            messages.error(request, "Converted quotations cannot be modified.")
            return redirect('quotation_view', pk=quotation.id)
            
        customers = Customer.objects.filter(company=company, is_active=True)
        products = Product.objects.filter(company=company, is_active=True)
        
        # Build items JSON for editing grid
        items_data = []
        for item in quotation.items.all():
            items_data.append({
                'product_id': item.product.id,
                'product_name': item.product.name,
                'quantity': float(item.quantity),
                'rate': float(item.rate),
                'discount': float(item.discount),
                'gst_rate': float(item.gst_rate),
                'taxable_value': float(item.taxable_value),
                'total_amount': float(item.total_amount)
            })

        from .utils import get_or_create_predefined_quotation_terms
        predefined_terms = get_or_create_predefined_quotation_terms(company)
        selected_terms = list(quotation.selected_terms.all().order_by('display_order', 'id'))
        selected_texts = set(st.term_text for st in selected_terms)
        has_selected = len(selected_terms) > 0

        predefined_terms_list = []
        for pt in predefined_terms:
            pt.is_selected_in_quotation = (pt.term_text in selected_texts) if has_selected else True
            predefined_terms_list.append(pt)
            
        return render(request, 'company/quotation_edit.html', {
            'customers': customers,
            'products': products,
            'products_json': build_products_json(products),
            'quotation': quotation,
            'items_json': json.dumps(items_data, cls=DjangoJSONEncoder),
            'predefined_terms': predefined_terms_list,
            'selected_terms': selected_terms
        })

    def post(self, request, pk):
        company = request.user.company
        quotation = get_object_or_404(Quotation, id=pk, company=company)
        is_ajax = (
            request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in request.META.get('HTTP_ACCEPT', '') or
            request.content_type == 'application/json'
        )

        try:
            if request.content_type == 'application/json' and request.body:
                data = json.loads(request.body)
            else:
                data = request.POST.dict()
                if 'items_json' in request.POST:
                    data['items'] = request.POST['items_json']
        except Exception:
            data = request.POST.dict()

        result = QuotationService.update_quotation(company, request.user, quotation, data)
        if result.get('success'):
            messages.success(request, result.get('message', 'Sales quotation updated successfully!'))
            if is_ajax:
                return JsonResponse(result, status=200, encoder=DjangoJSONEncoder)
            return redirect('quotation_list')
        else:
            err_msg = result.get('message', 'Unable to update sales quotation. Please check the entered details.')
            messages.error(request, err_msg)
            if is_ajax:
                return JsonResponse(result, status=400, encoder=DjangoJSONEncoder)
            customers = Customer.objects.filter(company=company, is_active=True)
            products = Product.objects.filter(company=company, is_active=True)
            return render(request, 'company/quotation_edit.html', {
                'customers': customers,
                'products': products,
                'products_json': build_products_json(products),
                'quotation': quotation,
                'form_data': data,
                'errors': result.get('errors', {})
            })


class QuotationDetailView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = Quotation
    template_name = 'company/quotation_detail.html'
    context_object_name = 'quotation'


def build_quotation_context(quotation):
    from .utils import parse_product_specifications, build_hsn_sac_tax_summary
    company = quotation.company
    customer = quotation.customer
    
    company_state_code = str(company.state_code or '').strip().zfill(2)
    customer_state_code = str(customer.billing_state_code or company_state_code).strip().zfill(2)
    is_interstate = (company_state_code != customer_state_code)
    
    enriched_items = []
    for item in quotation.items.all().select_related('product', 'product__unit', 'product__hsn_sac'):
        prod = item.product
        desc = prod.description or ''
        parsed_spec = parse_product_specifications(desc)
        
        enriched_items.append({
            'item': item,
            'product': prod,
            'product_name': prod.name if prod else 'Item',
            'product_image_url': prod.get_image_url() if prod else None,
            'specs': parsed_spec['specs'],
            'notes': parsed_spec['notes'],
            'unit_code': prod.unit.code if prod and prod.unit else 'PCS',
            'hsn_code': item.hsn_code,
            'quantity': item.quantity,
            'rate': item.rate,
            'discount': item.discount,
            'taxable_value': item.taxable_value,
            'gst_rate': item.gst_rate,
            'cgst_amount': item.cgst_amount,
            'sgst_amount': item.sgst_amount,
            'igst_amount': item.igst_amount,
            'total_amount': item.total_amount,
        })
        
    hsn_summary, total_quantity = build_hsn_sac_tax_summary(quotation.items.all(), company_state_code, customer_state_code)
    total_tax = quotation.cgst_total + quotation.sgst_total + quotation.igst_total
    selected_terms = list(quotation.selected_terms.all().order_by('display_order', 'id'))
    if selected_terms:
        terms_list = [st.term_text for st in selected_terms]
    else:
        raw_terms = quotation.terms or company.terms_and_conditions or ""
        terms_list = [t.strip() for t in raw_terms.splitlines() if t.strip()]

    return {
        'quotation': quotation,
        'company': company,
        'customer': customer,
        'enriched_items': enriched_items,
        'total_quantity': total_quantity,
        'hsn_summary': hsn_summary,
        'is_interstate': is_interstate,
        'total_tax': total_tax,
        'terms_list': terms_list,
    }


class QuotationPDFView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = Quotation
    template_name = 'company/quotation_pdf.html'
    context_object_name = 'quotation'

    def get(self, request, *args, **kwargs):
        try:
            self.object = self.get_object()
            context = self.get_context_data(object=self.object)
            q_context = build_quotation_context(self.object)
            context.update(q_context)
            return render(request, self.template_name, context)
        except Exception as e:
            messages.error(request, f"Unable to generate Quotation PDF. Please try again. ({str(e)})")
            return HttpResponse(f"<div style='font-family: sans-serif; padding: 2rem; color: #721c24; background-color: #f8d7da; border: 1px solid #f5c6cb; border-radius: 4px;'><h3>Unable to generate Quotation PDF</h3><p>{str(e)}</p></div>", status=500)



def quotation_convert_to_invoice(request, pk):
    company = request.user.company
    q = get_object_or_404(Quotation, id=pk, company=company)
    
    if q.status == 'CONVERTED' and q.converted_to_invoice:
        messages.info(request, f"Quotation {q.quotation_number} has already been converted to Tax Invoice {q.converted_to_invoice.invoice_number}.")
        return redirect('invoice_view', pk=q.converted_to_invoice.id)
    
    with transaction.atomic():
        # Generate Invoice Number
        count = Invoice.objects.filter(company=company).count() + 1
        inv_no = f"{company.invoice_prefix}{company.financial_year}-{str(count).zfill(company.invoice_padding)}"
        
        # Create Invoice
        invoice = Invoice.objects.create(
            company=company, customer=q.customer, invoice_number=inv_no,
            invoice_date=date.today(), due_date=date.today() + timedelta(days=30),
            place_of_supply=q.customer.billing_state, place_of_supply_code=q.customer.billing_state_code,
            status='DRAFT', notes=q.notes
        )
        
        for item in q.items.all():
            InvoiceItem.objects.create(
                invoice=invoice, product=item.product, quantity=item.quantity,
                rate=item.rate, discount=item.discount, taxable_value=item.taxable_value,
                gst_rate=item.gst_rate, cgst_amount=item.cgst_amount, sgst_amount=item.sgst_amount,
                igst_amount=item.igst_amount, total_amount=item.total_amount,
                hsn_sac_code=item.hsn_sac_code
            )
            
        recalculate_invoice_totals(invoice)
        q.status = 'CONVERTED'
        q.converted_to_invoice = invoice
        q.save()
        
    messages.success(request, f"Quotation converted to Tax Invoice: {inv_no}")
    return redirect('invoice_view', pk=invoice.id)


@transaction.atomic
def quotation_convert_to_sales_order(request, pk):
    company = request.user.company
    q = get_object_or_404(Quotation, id=pk, company=company)
    
    if q.status == 'CONVERTED' and q.converted_to_sales_order:
        messages.info(request, f"Quotation {q.quotation_number} has already been converted to Sales Order {q.converted_to_sales_order.order_number}.")
        return redirect('sales_order_view', pk=q.converted_to_sales_order.id)
    
    count = SalesOrder.objects.filter(company=company).count() + 1
    so_no = f"SO-{company.financial_year}-{str(count).zfill(5)}"
    
    so = SalesOrder.objects.create(
        company=company, customer=q.customer, order_number=so_no,
        order_date=date.today(), expected_delivery=date.today() + timedelta(days=7),
        status='PENDING', notes=q.notes
    )
    
    for item in q.items.all():
        SalesOrderItem.objects.create(
            sales_order=so, product=item.product, quantity=item.quantity,
            rate=item.rate, discount=item.discount, taxable_value=item.taxable_value,
            gst_rate=item.gst_rate, cgst_amount=item.cgst_amount, sgst_amount=item.sgst_amount,
            igst_amount=item.igst_amount, total_amount=item.total_amount,
            hsn_sac_code=item.hsn_sac_code
        )
        
    so.grand_total = q.grand_total
    so.save()
    
    q.status = 'CONVERTED'
    q.converted_to_sales_order = so
    q.save()
    
    messages.success(request, f"Quotation converted to Sales Order: {so_no}")
    return redirect('sales_order_view', pk=so.id)


# --- SALES ORDERS ---

class SalesOrderListView(CompanyRequiredMixin, CompanyQuerySetMixin, PaginationMixin, ListView):
    model = SalesOrder
    template_name = 'company/sales_order_list.html'
    context_object_name = 'orders'

    def get_queryset(self):
        qs = super().get_queryset().order_by('-order_date', '-id')
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(order_number__icontains=search) |
                Q(customer__name__icontains=search) |
                Q(customer__gstin__icontains=search) |
                Q(status__icontains=search)
            )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        else:
            qs = qs.exclude(status='CANCELLED')
        return qs


class SalesOrderCreateView(CompanyRequiredMixin, View):
    def get(self, request):
        company = request.user.company
        customers = Customer.objects.filter(company=company, is_active=True)
        products = Product.objects.filter(company=company, is_active=True)
        count = SalesOrder.objects.filter(company=company).count() + 1
        so_no = f"SO-{company.financial_year}-{str(count).zfill(5)}"
        return render(request, 'company/sales_order_add.html', {
            'customers': customers,
            'products': products,
            'products_json': build_products_json(products),
            'order_number': so_no
        })

    @transaction.atomic
    def post(self, request):
        company = request.user.company
        is_ajax = (
            request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in request.META.get('HTTP_ACCEPT', '') or
            request.content_type == 'application/json'
        )

        try:
            if request.content_type == 'application/json' and request.body:
                data = json.loads(request.body)
            else:
                data = request.POST.dict()
                if 'items_json' in request.POST:
                    data['items'] = request.POST['items_json']
        except Exception:
            data = request.POST.dict()

        try:
            cust_id = data.get('customer_id') or data.get('customer')
            if not cust_id:
                raise ValueError('Please select a customer.')

            customer = get_object_or_404(Customer, id=int(str(cust_id).strip()), company=company)

            so_no = (data.get('order_number') or '').strip()
            if not so_no:
                count = SalesOrder.objects.filter(company=company).count() + 1
                so_no = f"SO-{company.financial_year}-{str(count).zfill(5)}"

            # Duplicate check
            if SalesOrder.objects.filter(company=company, order_number=so_no).exists():
                raise ValueError(f"Sales Order number {so_no} already exists!")

            so_date_raw = data.get('date') or data.get('order_date')
            try:
                so_date = datetime.strptime(str(so_date_raw).strip(), '%Y-%m-%d').date() if so_date_raw else date.today()
            except (ValueError, TypeError):
                so_date = date.today()

            delivery_raw = data.get('expected_delivery')
            try:
                delivery = datetime.strptime(str(delivery_raw).strip(), '%Y-%m-%d').date() if delivery_raw else (so_date + timedelta(days=14))
            except (ValueError, TypeError):
                delivery = so_date + timedelta(days=14)

            items_data = data.get('items', [])
            if isinstance(items_data, str):
                try:
                    items_data = json.loads(items_data)
                except Exception:
                    items_data = []

            if not items_data:
                raise ValueError('Please select at least one valid product in the items grid.')

            so = SalesOrder.objects.create(
                company=company, customer=customer, order_number=so_no,
                order_date=so_date, expected_delivery=delivery, status='PENDING',
                notes=(data.get('notes') or '').strip()
            )
            
            grand = Decimal('0.00')
            created_items_count = 0
            for item in items_data:
                prod_id = item.get('product_id') or item.get('product')
                if not prod_id:
                    continue
                prod = get_object_or_404(Product, id=int(str(prod_id).strip()), company=company)
                qty = parse_money(item.get('quantity', 1))
                rate = parse_money(item.get('rate', 0))
                disc = parse_money(item.get('discount', 0))
                
                if qty <= Decimal('0.00'):
                    continue

                taxable = (qty * rate) - disc
                rate_gst = prod.hsn_sac.gst_rate if prod.hsn_sac else Decimal('0.00')
                cgst, sgst, igst, tot_gst = calculate_item_gst(company.state_code, customer.billing_state_code or company.state_code, taxable, rate_gst)
                tot = taxable + tot_gst
                
                hsn_code = prod.hsn_sac.code if prod.hsn_sac else ''
                SalesOrderItem.objects.create(
                    sales_order=so, product=prod, quantity=qty, rate=rate,
                    discount=disc, taxable_value=Decimal('0.00'), gst_rate=rate_gst,
                    cgst_amount=Decimal('0.00'), sgst_amount=Decimal('0.00'), igst_amount=Decimal('0.00'),
                    hsn_sac_code=hsn_code, total_amount=Decimal('0.00')
                )
                created_items_count += 1
                
            if created_items_count == 0:
                raise ValueError("At least one valid item is required in the sales order.")

            recalculate_sales_order_totals(so)
            log_action(request.user, 'CREATE_SALES_ORDER', 'SALES_ORDER', so.id, request=request)
            messages.success(request, f"Sales Order {so.order_number} saved successfully!")
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'status': 'success',
                    'message': f"Sales Order {so.order_number} saved successfully!",
                    'order_id': so.id,
                    'redirect_url': f"/company/sales-orders/{so.id}/"
                }, encoder=DjangoJSONEncoder)
            return redirect('sales_order_view', pk=so.id)
        except Exception as e:
            messages.error(request, str(e))
            if is_ajax:
                return JsonResponse({'success': False, 'status': 'error', 'message': str(e)}, status=400, encoder=DjangoJSONEncoder)
            customers = Customer.objects.filter(company=company, is_active=True)
            products = Product.objects.filter(company=company, is_active=True)
            return render(request, 'company/sales_order_add.html', {
                'customers': customers,
                'products': products,
                'products_json': build_products_json(products),
                'order_number': data.get('order_number', ''),
                'form_data': data
            })


class SalesOrderDetailView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = SalesOrder
    template_name = 'company/sales_order_detail.html'
    context_object_name = 'order'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .utils import build_hsn_sac_tax_summary, recalculate_sales_order_totals
        order = self.object
        recalculate_sales_order_totals(order)
        comp_code = str(order.company.state_code or '').strip().zfill(2)
        pos_code = str(order.customer.billing_state_code or comp_code).strip().zfill(2)
        summary_list, total_qty = build_hsn_sac_tax_summary(order.items.all(), comp_code, pos_code)
        context['hsn_summary'] = summary_list
        context['total_quantity'] = total_qty
        context['is_interstate'] = (comp_code != pos_code)
        return context


class SalesOrderPDFView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = SalesOrder
    template_name = 'company/sales_order_pdf.html'
    context_object_name = 'order'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .utils import build_hsn_sac_tax_summary, recalculate_sales_order_totals
        order = self.object
        recalculate_sales_order_totals(order)
        comp_code = str(order.company.state_code or '').strip().zfill(2)
        pos_code = str(order.customer.billing_state_code or comp_code).strip().zfill(2)
        summary_list, total_qty = build_hsn_sac_tax_summary(order.items.all(), comp_code, pos_code)
        context['hsn_summary'] = summary_list
        context['total_quantity'] = total_qty
        context['is_interstate'] = (comp_code != pos_code)
        return context


def sales_order_convert_to_invoice(request, pk):
    company = request.user.company
    so = get_object_or_404(SalesOrder, id=pk, company=company)
    with transaction.atomic():
        count = Invoice.objects.filter(company=company).count() + 1
        inv_no = f"{company.invoice_prefix}{company.financial_year}-{str(count).zfill(company.invoice_padding)}"
        
        invoice = Invoice.objects.create(
            company=company, customer=so.customer, invoice_number=inv_no,
            invoice_date=date.today(), due_date=date.today() + timedelta(days=30),
            place_of_supply=so.customer.billing_state, place_of_supply_code=so.customer.billing_state_code,
            status='DRAFT', notes=so.notes
        )
        for item in so.items.all():
            InvoiceItem.objects.create(
                invoice=invoice, product=item.product, quantity=item.quantity,
                rate=item.rate, discount=item.discount, taxable_value=item.taxable_value,
                gst_rate=item.gst_rate, cgst_amount=item.cgst_amount, sgst_amount=item.sgst_amount,
                igst_amount=item.igst_amount, total_amount=item.total_amount,
                hsn_sac_code=item.hsn_sac_code
            )
        recalculate_invoice_totals(invoice)
        so.status = 'INVOICED'
        so.save()
        
    messages.success(request, f"Sales Order converted to invoice: {inv_no}")
    return redirect('invoice_view', pk=invoice.id)


# --- TAX INVOICES ---

class InvoiceListView(CompanyRequiredMixin, CompanyQuerySetMixin, PaginationMixin, ListView):
    model = Invoice
    template_name = 'company/invoice_list.html'
    context_object_name = 'invoices'

    def get_queryset(self):
        qs = super().get_queryset().order_by('-invoice_date', '-id')
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(invoice_number__icontains=search) |
                Q(customer__name__icontains=search) |
                Q(customer__gstin__icontains=search) |
                Q(place_of_supply__icontains=search) |
                Q(status__icontains=search)
            )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        else:
            qs = qs.exclude(status='CANCELLED')
        return qs


class InvoiceCreateView(CompanyRequiredMixin, View):
    def get(self, request):
        company = request.user.company
        customers = Customer.objects.filter(company=company, is_active=True)
        products = Product.objects.filter(company=company, is_active=True)
        warehouses = Warehouse.objects.filter(company=company, is_active=True)
        
        # Formulate next sequence number
        count = Invoice.objects.filter(company=company).count() + 1
        inv_no = f"{company.invoice_prefix}{company.financial_year}-{str(count).zfill(company.invoice_padding)}"
        
        return render(request, 'company/invoice_add.html', {
            'customers': customers,
            'products': products,
            'products_json': build_products_json(products),
            'warehouses': warehouses,
            'invoice_number': inv_no
        })

    @transaction.atomic
    def post(self, request):
        company = request.user.company
        is_ajax = (
            request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in request.META.get('HTTP_ACCEPT', '') or
            request.content_type == 'application/json'
        )

        try:
            if request.content_type == 'application/json' and request.body:
                data = json.loads(request.body)
            else:
                data = request.POST.dict()
                if 'items_json' in request.POST:
                    data['items'] = request.POST['items_json']
        except Exception:
            data = request.POST.dict()

        try:
            cust_id = data.get('customer_id') or data.get('customer')
            if not cust_id:
                raise ValueError('Please select a customer.')

            customer = get_object_or_404(Customer, id=int(str(cust_id).strip()), company=company)

            inv_no = (data.get('invoice_number') or '').strip()
            if not inv_no:
                count = Invoice.objects.filter(company=company).count() + 1
                inv_no = f"{company.invoice_prefix}{company.financial_year}-{str(count).zfill(company.invoice_padding)}"

            # Duplicate check
            if Invoice.objects.filter(company=company, invoice_number=inv_no).exists():
                raise ValueError(f"Invoice number {inv_no} already exists!")

            inv_date_raw = data.get('invoice_date') or data.get('date')
            try:
                inv_date = datetime.strptime(str(inv_date_raw).strip(), '%Y-%m-%d').date() if inv_date_raw else date.today()
            except (ValueError, TypeError):
                inv_date = date.today()

            due_date_raw = data.get('due_date')
            try:
                due_date = datetime.strptime(str(due_date_raw).strip(), '%Y-%m-%d').date() if due_date_raw else (inv_date + timedelta(days=30))
            except (ValueError, TypeError):
                due_date = inv_date + timedelta(days=30)

            place_of_supply = (data.get('place_of_supply') or customer.billing_state or company.state or 'Maharashtra').strip()
            place_of_supply_code = (data.get('place_of_supply_code') or customer.billing_state_code or company.state_code or '27').strip().zfill(2)
            reverse_charge = bool(data.get('reverse_charge', False))
            notes = (data.get('notes') or '').strip()
            
            items_data = data.get('items', [])
            if isinstance(items_data, str):
                try:
                    items_data = json.loads(items_data)
                except Exception:
                    items_data = []

            if not items_data:
                raise ValueError('Please select at least one valid product in the items grid.')

            wh_id = data.get('warehouse_id')
            warehouse = None
            if wh_id:
                try:
                    warehouse = Warehouse.objects.filter(id=int(wh_id), company=company).first()
                except (ValueError, TypeError):
                    pass
            if not warehouse:
                warehouse = Warehouse.objects.filter(company=company, is_active=True).first() or Warehouse.objects.filter(company=company).first()
            if not warehouse:
                warehouse = Warehouse.objects.create(company=company, name="Main Warehouse", code="WH-MAIN")

            # Create Invoice base
            invoice = Invoice.objects.create(
                company=company, customer=customer, invoice_number=inv_no,
                invoice_date=inv_date, due_date=due_date,
                place_of_supply=place_of_supply, place_of_supply_code=place_of_supply_code,
                reverse_charge=reverse_charge, status='POSTED', notes=notes, terms=company.terms_and_conditions
            )
            
            created_items_count = 0
            for item in items_data:
                prod_id = item.get('product_id') or item.get('product')
                if not prod_id:
                    continue
                prod = get_object_or_404(Product, id=int(str(prod_id).strip()), company=company)
                qty = parse_money(item.get('quantity', 1))
                rate = parse_money(item.get('rate', 0))
                disc = parse_money(item.get('discount', 0))
                
                if qty <= Decimal('0.00'):
                    continue

                # Check negative stock permissions
                if prod.track_inventory and not prod.allow_negative_stock:
                    if prod.current_stock < qty:
                        raise ValueError(f"Insufficient stock for product '{prod.name}' (Available: {prod.current_stock}, Requested: {qty})")
                
                hsn_code = prod.hsn_sac.code if prod.hsn_sac else ''
                gst_rate = prod.hsn_sac.gst_rate if prod.hsn_sac else Decimal('0.00')
                InvoiceItem.objects.create(
                    invoice=invoice, product=prod, quantity=qty, rate=rate,
                    discount=disc, taxable_value=Decimal('0.00'), hsn_sac_code=hsn_code,
                    gst_rate=gst_rate, total_amount=Decimal('0.00')
                )
                
                # Stock Movement Log
                if prod.track_inventory:
                    StockMovement.objects.create(
                        company=company, product=prod, warehouse=warehouse,
                        quantity=-qty, movement_type='SALE', reference_id=invoice.id,
                        reference_no=invoice.invoice_number, created_by=request.user
                    )
                    update_product_stock(prod.id)
                created_items_count += 1
                    
            if created_items_count == 0:
                raise ValueError("At least one valid item is required in the invoice.")

            apply_round_off = None
            if 'round_off_applied' in data:
                apply_round_off = bool(data.get('round_off_applied'))
            elif 'round_off' in data:
                try:
                    apply_round_off = (Decimal(str(data.get('round_off') or 0)) != Decimal('0.00'))
                except Exception:
                    apply_round_off = False

            recalculate_invoice_totals(invoice, advance_amount=adv_amt, amount_paid_now=amt_now, payment_percentage=pmt_pct, advance_paid=adv_paid, payment_status=pmt_status, apply_round_off=apply_round_off)
            
            # Update customer receivable balance
            customer.outstanding_balance += invoice.balance_due
            customer.save()
            
            record_invoice_accounting(invoice)
            
            log_action(request.user, 'CREATE_INVOICE', 'INVOICE', invoice.id, new_values={'amount': str(invoice.grand_total)}, request=request)
            messages.success(request, f"Invoice {invoice.invoice_number} saved and posted successfully!")
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'status': 'success',
                    'message': f"Invoice {invoice.invoice_number} saved and posted successfully!",
                    'invoice_id': invoice.id,
                    'redirect_url': f"/company/invoices/{invoice.id}/"
                }, encoder=DjangoJSONEncoder)
            return redirect('invoice_view', pk=invoice.id)
        except Exception as e:
            messages.error(request, str(e))
            if is_ajax:
                return JsonResponse({'success': False, 'status': 'error', 'message': str(e)}, status=400, encoder=DjangoJSONEncoder)
            customers = Customer.objects.filter(company=company, is_active=True)
            products = Product.objects.filter(company=company, is_active=True)
            warehouses = Warehouse.objects.filter(company=company, is_active=True)
            return render(request, 'company/invoice_add.html', {
                'customers': customers,
                'products': products,
                'products_json': build_products_json(products),
                'warehouses': warehouses,
                'invoice_number': data.get('invoice_number', ''),
                'form_data': data
            })


class InvoiceUpdateView(CompanyRequiredMixin, View):
    def get(self, request, pk):
        company = request.user.company
        invoice = get_object_or_404(Invoice, id=pk, company=company)
        if invoice.status == 'CANCELLED':
            messages.error(request, "Cancelled invoices cannot be edited.")
            return redirect('invoice_view', pk=invoice.id)
            
        customers = Customer.objects.filter(company=company, is_active=True)
        products = Product.objects.filter(company=company, is_active=True)
        warehouses = Warehouse.objects.filter(company=company, is_active=True)
        
        return render(request, 'company/invoice_edit.html', {
            'invoice': invoice,
            'customers': customers,
            'products': products,
            'products_json': build_products_json(products),
            'warehouses': warehouses,
        })

    @transaction.atomic
    def post(self, request, pk):
        company = request.user.company
        invoice = get_object_or_404(Invoice, id=pk, company=company)
        if invoice.status == 'CANCELLED':
            raise ValueError("Cancelled invoices cannot be edited.")

        if request.user.role not in ('ADMIN', 'ACCOUNTANT', 'SUPERADMIN') and not request.user.is_superuser:
            raise PermissionDenied("Only Company Admin and Accountant are authorized to modify invoice payment details.")

        is_ajax = (
            request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in request.META.get('HTTP_ACCEPT', '') or
            request.content_type == 'application/json'
        )

        try:
            if request.content_type == 'application/json' and request.body:
                data = json.loads(request.body)
            else:
                data = request.POST.dict()
                if 'items_json' in request.POST:
                    data['items'] = request.POST['items_json']
        except Exception:
            data = request.POST.dict()

        try:
            old_balance_due = invoice.balance_due or max(Decimal('0.00'), invoice.grand_total - invoice.paid_amount)
            old_customer = invoice.customer

            cust_id = data.get('customer_id') or data.get('customer')
            if cust_id:
                customer = get_object_or_404(Customer, id=int(str(cust_id).strip()), company=company)
                invoice.customer = customer

            inv_date_raw = data.get('invoice_date') or data.get('date')
            if inv_date_raw:
                try:
                    invoice.invoice_date = datetime.strptime(str(inv_date_raw).strip(), '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    pass

            due_date_raw = data.get('due_date')
            if due_date_raw:
                try:
                    invoice.due_date = datetime.strptime(str(due_date_raw).strip(), '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    pass

            if data.get('place_of_supply'):
                invoice.place_of_supply = data.get('place_of_supply').strip()
            if data.get('place_of_supply_code'):
                invoice.place_of_supply_code = data.get('place_of_supply_code').strip().zfill(2)
            if 'reverse_charge' in data:
                invoice.reverse_charge = bool(data.get('reverse_charge'))
            if 'notes' in data:
                invoice.notes = (data.get('notes') or '').strip()

            items_data = data.get('items', [])
            if isinstance(items_data, str):
                try:
                    items_data = json.loads(items_data)
                except Exception:
                    items_data = []

            if items_data:
                for item in invoice.items.all():
                    if item.product and item.product.track_inventory:
                        StockMovement.objects.create(
                            company=company, product=item.product, warehouse=Warehouse.objects.filter(company=company).first(),
                            quantity=item.quantity, movement_type='ADJUSTMENT', reference_id=invoice.id,
                            reference_no=f"EDIT-{invoice.invoice_number}", created_by=request.user
                        )
                        update_product_stock(item.product.id)
                invoice.items.all().delete()

                wh_id = data.get('warehouse_id')
                warehouse = Warehouse.objects.filter(id=int(wh_id), company=company).first() if wh_id else Warehouse.objects.filter(company=company, is_active=True).first()

                for item in items_data:
                    prod_id = item.get('product_id') or item.get('product')
                    if not prod_id:
                        continue
                    prod = get_object_or_404(Product, id=int(str(prod_id).strip()), company=company)
                    qty = parse_money(item.get('quantity', 1))
                    rate = parse_money(item.get('rate', 0))
                    disc = parse_money(item.get('discount', 0))
                    if qty <= Decimal('0.00'):
                        continue
                    if prod.track_inventory and not prod.allow_negative_stock:
                        if prod.current_stock < qty:
                            raise ValueError(f"Insufficient stock for product '{prod.name}' (Available: {prod.current_stock}, Requested: {qty})")

                    hsn_code = prod.hsn_sac.code if prod.hsn_sac else ''
                    gst_rate = prod.hsn_sac.gst_rate if prod.hsn_sac else Decimal('0.00')
                    InvoiceItem.objects.create(
                        invoice=invoice, product=prod, quantity=qty, rate=rate,
                        discount=disc, taxable_value=Decimal('0.00'), hsn_sac_code=hsn_code,
                        gst_rate=gst_rate, total_amount=Decimal('0.00')
                    )

                    if prod.track_inventory:
                        StockMovement.objects.create(
                            company=company, product=prod, warehouse=warehouse,
                            quantity=-qty, movement_type='SALE', reference_id=invoice.id,
                            reference_no=invoice.invoice_number, created_by=request.user
                        )
                        update_product_stock(prod.id)

            apply_round_off = None
            if 'round_off_applied' in data:
                apply_round_off = bool(data.get('round_off_applied'))
            elif 'round_off' in data:
                try:
                    apply_round_off = (Decimal(str(data.get('round_off') or 0)) != Decimal('0.00'))
                except Exception:
                    apply_round_off = False

            recalculate_invoice_totals(invoice, advance_amount=adv_amt, amount_paid_now=amt_now, payment_percentage=pmt_pct, advance_paid=adv_paid, payment_status=pmt_status, apply_round_off=apply_round_off)

            if old_customer != invoice.customer:
                old_customer.outstanding_balance -= old_balance_due
                old_customer.save()
                invoice.customer.outstanding_balance += invoice.balance_due
                invoice.customer.save()
            else:
                diff = invoice.balance_due - old_balance_due
                invoice.customer.outstanding_balance += diff
                invoice.customer.save()

            record_invoice_accounting(invoice)

            log_action(request.user, 'EDIT_INVOICE', 'INVOICE', invoice.id, new_values={'amount': str(invoice.grand_total), 'paid': str(invoice.total_payment_received)}, request=request)
            messages.success(request, f"Invoice {invoice.invoice_number} updated successfully!")
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'status': 'success',
                    'message': f"Invoice {invoice.invoice_number} updated successfully!",
                    'invoice_id': invoice.id,
                    'redirect_url': f"/company/invoices/{invoice.id}/"
                }, encoder=DjangoJSONEncoder)
            return redirect('invoice_view', pk=invoice.id)
        except Exception as e:
            messages.error(request, str(e))
            if is_ajax:
                return JsonResponse({'success': False, 'status': 'error', 'message': str(e)}, status=400, encoder=DjangoJSONEncoder)
            customers = Customer.objects.filter(company=company, is_active=True)
            products = Product.objects.filter(company=company, is_active=True)
            warehouses = Warehouse.objects.filter(company=company, is_active=True)
            return render(request, 'company/invoice_edit.html', {
                'invoice': invoice,
                'customers': customers,
                'products': products,
                'products_json': build_products_json(products),
                'warehouses': warehouses,
                'form_data': data
            })



def add_invoice_hsn_summary_to_context(invoice, context):
    from .utils import build_hsn_sac_tax_summary
    company_state_code = str(invoice.company.state_code or '').strip().zfill(2)
    pos_state_code = str(invoice.place_of_supply_code or getattr(invoice.customer, 'billing_state_code', company_state_code)).strip().zfill(2)
    is_interstate = (company_state_code != pos_state_code)

    summary_list, total_qty = build_hsn_sac_tax_summary(invoice.items.all(), company_state_code, pos_state_code)
    context['hsn_summary'] = summary_list
    context['total_quantity'] = total_qty
    context['is_interstate'] = is_interstate
    return context


class InvoiceDetailView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = Invoice
    template_name = 'company/invoice_detail.html'
    context_object_name = 'invoice'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        invoice = self.get_object()
        company = invoice.company
        
        # Payment QR link
        qr_string = generate_upi_qr_string(
            company.upi_id, company.trade_name or company.name,
            invoice.outstanding_amount(), invoice.invoice_number
        )
        context['upi_qr_string'] = qr_string
        context = add_invoice_hsn_summary_to_context(invoice, context)
        return context


class InvoicePDFView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = Invoice
    template_name = 'company/invoice_pdf.html'
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        context = self.get_context_data(object=self.object)
        
        # Dynamic QR generate inside template
        qr_string = generate_upi_qr_string(
            self.object.company.upi_id, self.object.company.trade_name or self.object.company.name,
            self.object.outstanding_amount(), self.object.invoice_number
        )
        context['upi_qr_string'] = qr_string
        context = add_invoice_hsn_summary_to_context(self.object, context)
        return render(request, self.template_name, context)


@transaction.atomic
def invoice_cancel(request, pk):
    company = request.user.company
    invoice = get_object_or_404(Invoice, id=pk, company=company)
    
    if invoice.status == 'CANCELLED':
        messages.error(request, "Invoice is already cancelled.")
        return redirect('invoice_view', pk=invoice.id)
        
    # Reverse stock adjustments
    movements = StockMovement.objects.filter(company=company, reference_id=invoice.id, movement_type='SALE')
    for mv in movements:
        # Create counter-balancing movement
        StockMovement.objects.create(
            company=company, product=mv.product, warehouse=mv.warehouse,
            quantity=abs(mv.quantity), movement_type='SALES_RETURN', reference_id=invoice.id,
            reference_no=f"CNL: {invoice.invoice_number}", created_by=request.user
        )
        update_product_stock(mv.product.id)
        
    # Rebalance customer ledger
    customer = invoice.customer
    customer.outstanding_balance -= invoice.grand_total
    customer.save()
    
    cancel_invoice_accounting(invoice)
    
    invoice.status = 'CANCELLED'
    invoice.save()
    
    log_action(request.user, 'CANCEL_INVOICE', 'INVOICE', invoice.id, request=request)
    messages.success(request, f"Invoice {invoice.invoice_number} cancelled and customer balance restored.")
    return redirect('invoice_view', pk=invoice.id)


@transaction.atomic
def invoice_post(request, pk):
    company = request.user.company
    invoice = get_object_or_404(Invoice, id=pk, company=company)
    
    if invoice.status != 'DRAFT':
        messages.error(request, "Only draft invoices can be posted.")
        return redirect('invoice_view', pk=invoice.id)
        
    # Check warehouse and items
    warehouse = Warehouse.objects.filter(company=company, is_active=True).first()
    if not warehouse:
        messages.error(request, "No active warehouse found. Please create one first.")
        return redirect('invoice_view', pk=invoice.id)
        
    # Recalculate totals
    recalculate_invoice_totals(invoice)
    
    # Process items: stock out
    for item in invoice.items.all():
        prod = item.product
        qty = item.quantity
        
        if prod.track_inventory and not prod.allow_negative_stock:
            if prod.current_stock < qty:
                messages.error(request, f"Insufficient stock for product '{prod.name}' (Available: {prod.current_stock}, Requested: {qty})")
                return redirect('invoice_view', pk=invoice.id)
                
        if prod.track_inventory:
            StockMovement.objects.create(
                company=company, product=prod, warehouse=warehouse,
                quantity=-qty, movement_type='SALE', reference_id=invoice.id,
                reference_no=invoice.invoice_number, created_by=request.user
            )
            update_product_stock(prod.id)
            
    # Rebalance customer receivable
    customer = invoice.customer
    customer.outstanding_balance += invoice.grand_total
    customer.save()
    
    invoice.status = 'POSTED'
    invoice.save()
    
    # Record accounting
    record_invoice_accounting(invoice)
    
    log_action(request.user, 'POST_INVOICE', 'INVOICE', invoice.id, request=request)
    messages.success(request, f"Invoice {invoice.invoice_number} posted successfully.")
    return redirect('invoice_view', pk=invoice.id)


# --- PURCHASE ORDERS ---

class PurchaseOrderListView(CompanyRequiredMixin, CompanyQuerySetMixin, PaginationMixin, ListView):
    model = PurchaseOrder
    template_name = 'company/purchase_order_list.html'
    context_object_name = 'purchase_orders'

    def get_queryset(self):
        qs = super().get_queryset().select_related('supplier', 'warehouse', 'created_by').prefetch_related('items').order_by('-po_date', '-id')
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(po_number__icontains=search) |
                Q(supplier__name__icontains=search) |
                Q(supplier__business_name__icontains=search) |
                Q(supplier__gstin__icontains=search) |
                Q(status__icontains=search)
            )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        
        supplier_id = self.request.GET.get('supplier')
        if supplier_id:
            try:
                qs = qs.filter(supplier_id=int(supplier_id))
            except (ValueError, TypeError):
                pass
                
        date_from = self.request.GET.get('date_from')
        if date_from:
            qs = qs.filter(po_date__gte=date_from)
            
        date_to = self.request.GET.get('date_to')
        if date_to:
            qs = qs.filter(po_date__lte=date_to)
            
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.user.company
        context['suppliers'] = Supplier.objects.filter(company=company, is_active=True).order_by('name')
        context['statuses'] = PurchaseOrder.STATUS_CHOICES
        context['selected_status'] = self.request.GET.get('status', '')
        context['selected_supplier'] = self.request.GET.get('supplier', '')
        context['selected_search'] = self.request.GET.get('search', '')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        return context


class PurchaseOrderCreateView(CompanyRequiredMixin, View):
    def get(self, request):
        company = request.user.company
        suppliers = Supplier.objects.filter(company=company, is_active=True).order_by('name')
        products = Product.objects.filter(company=company, is_active=True).order_by('name')
        warehouses = Warehouse.objects.filter(company=company, is_active=True).order_by('name')
        po_no = PurchaseOrderService.generate_po_number(company)

        return render(request, 'company/purchase_order_add.html', {
            'suppliers': suppliers,
            'products': products,
            'products_json': build_products_json(products),
            'warehouses': warehouses,
            'po_number': po_no,
            'today_date': date.today().strftime('%Y-%m-%d')
        })

    def post(self, request):
        company = request.user.company
        is_ajax = (
            request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in request.META.get('HTTP_ACCEPT', '') or
            request.content_type == 'application/json'
        )

        try:
            if request.content_type == 'application/json' and request.body:
                data = json.loads(request.body)
            else:
                data = request.POST.dict()
                if 'items_json' in request.POST:
                    data['items'] = request.POST['items_json']
        except Exception:
            data = request.POST.dict()

        result = PurchaseOrderService.create_purchase_order(company, request.user, data, files=request.FILES)
        if result.get('success'):
            messages.success(request, result.get('message', 'Purchase Order saved successfully!'))
            if is_ajax:
                return JsonResponse(result, status=200, encoder=DjangoJSONEncoder)
            return redirect('purchase_order_view', pk=result.get('po_id'))
        else:
            err_msg = result.get('message', 'Unable to save Purchase Order. Please check the entered details.')
            messages.error(request, err_msg)
            if is_ajax:
                return JsonResponse(result, status=400, encoder=DjangoJSONEncoder)
            suppliers = Supplier.objects.filter(company=company, is_active=True).order_by('name')
            products = Product.objects.filter(company=company, is_active=True).order_by('name')
            warehouses = Warehouse.objects.filter(company=company, is_active=True).order_by('name')
            return render(request, 'company/purchase_order_add.html', {
                'suppliers': suppliers,
                'products': products,
                'products_json': build_products_json(products),
                'warehouses': warehouses,
                'po_number': data.get('po_number', ''),
                'form_data': data,
                'errors': result.get('errors', {})
            })


class PurchaseOrderUpdateView(CompanyRequiredMixin, View):
    def get(self, request, pk):
        company = request.user.company
        po = get_object_or_404(PurchaseOrder, id=pk, company=company)

        suppliers = Supplier.objects.filter(company=company, is_active=True).order_by('name')
        products = Product.objects.filter(company=company, is_active=True).order_by('name')
        warehouses = Warehouse.objects.filter(company=company, is_active=True).order_by('name')

        from .utils import parse_state_and_code
        supplier_state_raw = po.supplier_state_code_snapshot or po.supplier_state_snapshot or (po.supplier.state_code if po.supplier else '') or (po.supplier.state if po.supplier else '') or ''
        supplier_state_name, supplier_state_code = parse_state_and_code(supplier_state_raw)

        items_json = []
        for item in po.items.all():
            items_json.append({
                'product_id': item.product.id if item.product else 'OTHER',
                'product_name': item.product_name_snapshot,
                'description': item.description_snapshot or '',
                'hsn_sac': item.hsn_sac_snapshot or '',
                'uqc': item.uqc_snapshot or '',
                'quantity': float(item.quantity),
                'rate': float(item.rate),
                'discount': float(item.discount),
                'taxable_amount': float(item.taxable_amount),
                'gst_rate': float(item.gst_rate),
                'cgst_amount': float(item.cgst_amount),
                'sgst_amount': float(item.sgst_amount),
                'igst_amount': float(item.igst_amount),
                'cess_amount': float(item.cess_amount),
                'total_amount': float(item.total_amount),
                'image_url': ''
            })

        return render(request, 'company/purchase_order_edit.html', {
            'po': po,
            'suppliers': suppliers,
            'products': products,
            'products_json': build_products_json(products),
            'warehouses': warehouses,
            'po_items_json': json.dumps(items_json),
            'statuses': PurchaseOrder.STATUS_CHOICES,
            'supplier_state_name': supplier_state_name,
            'supplier_state_code': supplier_state_code,
        })

    def post(self, request, pk):
        company = request.user.company
        is_ajax = (
            request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in request.META.get('HTTP_ACCEPT', '') or
            request.content_type == 'application/json'
        )

        try:
            if request.content_type == 'application/json' and request.body:
                data = json.loads(request.body)
            else:
                data = request.POST.dict()
                if 'items_json' in request.POST:
                    data['items'] = request.POST['items_json']
        except Exception:
            data = request.POST.dict()

        result = PurchaseOrderService.update_purchase_order(company, request.user, pk, data, files=request.FILES)
        if result.get('success'):
            messages.success(request, result.get('message', 'Purchase Order updated successfully!'))
            if is_ajax:
                return JsonResponse(result, status=200, encoder=DjangoJSONEncoder)
            return redirect('purchase_order_view', pk=pk)
        else:
            err_msg = result.get('message', 'Unable to update Purchase Order. Please check the entered details.')
            messages.error(request, err_msg)
            if is_ajax:
                return JsonResponse(result, status=400, encoder=DjangoJSONEncoder)
            po = get_object_or_404(PurchaseOrder, id=pk, company=company)
            suppliers = Supplier.objects.filter(company=company, is_active=True).order_by('name')
            products = Product.objects.filter(company=company, is_active=True).order_by('name')
            warehouses = Warehouse.objects.filter(company=company, is_active=True).order_by('name')
            return render(request, 'company/purchase_order_edit.html', {
                'po': po,
                'suppliers': suppliers,
                'products': products,
                'products_json': build_products_json(products),
                'warehouses': warehouses,
                'form_data': data,
                'errors': result.get('errors', {}),
                'statuses': PurchaseOrder.STATUS_CHOICES
            })


class PurchaseOrderDetailView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = PurchaseOrder
    template_name = 'company/purchase_order_detail.html'
    context_object_name = 'po'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .utils import build_hsn_sac_tax_summary, recalculate_purchase_order_totals
        po = self.object
        recalculate_purchase_order_totals(po)
        company = po.company
        supplier = po.supplier
        
        company_state_code = str(company.state_code or '').strip().zfill(2)
        supplier_state_code = str(po.supplier_state_code_snapshot or (supplier.state_code if supplier else company_state_code)).strip().zfill(2)
        is_interstate = (company_state_code != supplier_state_code)
        summary_list, total_qty = build_hsn_sac_tax_summary(po.items.all(), company_state_code, supplier_state_code)
        
        context['is_interstate'] = is_interstate
        context['hsn_summary'] = summary_list
        context['total_quantity'] = total_qty
        context['statuses'] = PurchaseOrder.STATUS_CHOICES
        return context


class PurchaseOrderPDFView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = PurchaseOrder
    template_name = 'company/purchase_order_pdf.html'
    context_object_name = 'po'

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        context = self.get_context_data(object=self.object)
        from .utils import build_hsn_sac_tax_summary, recalculate_purchase_order_totals
        po = self.object
        recalculate_purchase_order_totals(po)
        company = po.company
        supplier = po.supplier

        company_state_code = str(company.state_code or '').strip().zfill(2)
        supplier_state_code = str(po.supplier_state_code_snapshot or (supplier.state_code if supplier else company_state_code)).strip().zfill(2)
        is_interstate = (company_state_code != supplier_state_code)
        summary_list, total_qty = build_hsn_sac_tax_summary(po.items.all(), company_state_code, supplier_state_code)

        raw_terms = company.terms_and_conditions or ""
        terms_list = [t.strip() for t in raw_terms.splitlines() if t.strip()]

        context.update({
            'po': po,
            'company': company,
            'supplier': supplier,
            'is_interstate': is_interstate,
            'hsn_summary': summary_list,
            'total_quantity': total_qty,
            'terms_list': terms_list,
        })
        return self.render_to_response(context)


def purchase_order_status_change(request, pk, status):
    if not request.user.is_authenticated or not getattr(request.user, 'company', None):
        return JsonResponse({'success': False, 'message': 'Authentication required.'}, status=403)
    
    company = request.user.company
    po = get_object_or_404(PurchaseOrder, id=pk, company=company)
    
    valid_statuses = [choice[0] for choice in PurchaseOrder.STATUS_CHOICES]
    status_upper = status.upper().strip()
    if status_upper not in valid_statuses:
        return JsonResponse({'success': False, 'message': f'Invalid status "{status}".'}, status=400)

    po.status = status_upper
    po.save()
    log_action(request.user, 'CHANGE_PO_STATUS', 'PURCHASE_ORDER', po.id, new_values={'status': status_upper}, request=request)
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.META.get('HTTP_ACCEPT', ''):
        return JsonResponse({'success': True, 'message': f"Purchase Order status changed to {po.get_status_display()}.", 'status': po.status, 'status_display': po.get_status_display()})
    
    messages.success(request, f"Purchase Order status updated to {po.get_status_display()}.")
    return redirect('purchase_order_view', pk=po.id)


def purchase_order_duplicate(request, pk):
    if not request.user.is_authenticated or not getattr(request.user, 'company', None):
        messages.error(request, "Authentication required.")
        return redirect('login')

    company = request.user.company
    try:
        new_po = PurchaseOrderService.duplicate_purchase_order(company, request.user, pk)
        messages.success(request, f"Purchase Order duplicated successfully as {new_po.po_number}.")
        return redirect('purchase_order_edit', pk=new_po.id)
    except Exception as e:
        messages.error(request, f"Error duplicating Purchase Order: {str(e)}")
        return redirect('purchase_order_list')


def purchase_order_send_email(request, pk):
    if not request.user.is_authenticated or getattr(request.user, 'company', None) is None:
        return JsonResponse({'success': False, 'message': 'Authentication required.'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST request required.'}, status=405)

    company = request.user.company
    po = get_object_or_404(PurchaseOrder, id=pk, company=company)

    try:
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()

        recipient_email = str(data.get('recipient_email') or data.get('email') or (po.supplier.email if po.supplier else '')).strip()
        subject_input = str(data.get('subject') or '').strip()
        message_input = str(data.get('message') or '').strip()
    except Exception:
        return JsonResponse({'success': False, 'message': 'Invalid input data.'}, status=400)

    import re
    email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not recipient_email or not re.match(email_regex, recipient_email):
        return JsonResponse({'success': False, 'message': 'Please enter a valid email address.'}, status=400)

    company_state_code = str(company.state_code or '').strip().zfill(2)
    supplier_state_code = str(po.supplier.state_code or company_state_code if po.supplier else company_state_code).strip().zfill(2)
    is_interstate = (company_state_code != supplier_state_code)
    raw_terms = company.terms_and_conditions or ""
    terms_list = [t.strip() for t in raw_terms.splitlines() if t.strip()]

    from .utils import build_hsn_sac_tax_summary
    summary_list, total_qty = build_hsn_sac_tax_summary(po.items.all(), company_state_code, supplier_state_code)

    context = {
        'po': po,
        'company': company,
        'supplier': po.supplier,
        'is_interstate': is_interstate,
        'hsn_summary': summary_list,
        'total_quantity': total_qty,
        'terms_list': terms_list
    }

    from django.template.loader import render_to_string
    html_content = render_to_string('company/purchase_order_pdf.html', context, request=request)

    pdf_data = html_to_pdf_bytes(html_content)
    if not pdf_data or len(pdf_data) == 0:
        return JsonResponse({'success': False, 'message': 'Failed to generate Purchase Order PDF document.'}, status=500)

    from django.core.mail import EmailMessage
    from django.conf import settings

    supplier_name = po.supplier.name if po.supplier else "Valued Supplier"
    subject = subject_input or f"Purchase Order {po.po_number} from {company.name}"
    date_str = po.po_date.strftime('%d %b %Y') if hasattr(po.po_date, 'strftime') else str(po.po_date)
    
    default_body = (
        f"Dear {supplier_name},\n\n"
        f"Please find attached Purchase Order {po.po_number} from {company.name}.\n\n"
        f"PO Date: {date_str}\n"
        f"Total Amount: ₹{po.grand_total}\n\n"
        f"Please confirm receipt and expected delivery date.\n\n"
        f"Regards,\n"
        f"{company.name}"
    )
    body = message_input or default_body

    try:
        from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', None) or getattr(settings, 'EMAIL_HOST_USER', 'noreply@gblbilling.com')
        email = EmailMessage(
            subject,
            body,
            from_email,
            [recipient_email]
        )
        safe_filename = f"PurchaseOrder-{po.po_number}.pdf".replace(' ', '_').replace('/', '_')
        email.attach(safe_filename, pdf_data, 'application/pdf')
        email.send()

        po.status = 'SENT'
        po.save()
        log_action(request.user, 'SEND_EMAIL_PO', 'PURCHASE_ORDER', po.id, new_values={'recipient': recipient_email}, request=request)
        return JsonResponse({'success': True, 'message': f"Purchase Order PDF sent successfully to {recipient_email}."})
    except Exception as e:
        import logging
        logging.getLogger('django').error(f"SMTP Error sending Purchase Order {po.po_number}: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Unable to send email: {str(e)}'}, status=500)


def purchase_order_convert_to_bill(request, pk):
    if not request.user.is_authenticated or not getattr(request.user, 'company', None):
        messages.error(request, "Authentication required.")
        return redirect('login')

    company = request.user.company
    result = PurchaseOrderService.convert_to_purchase_bill(company, request.user, pk)
    if result.get('success'):
        messages.success(request, result.get('message'))
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.META.get('HTTP_ACCEPT', ''):
            return JsonResponse(result)
        return redirect('purchase_bill_view', pk=result.get('bill_id'))
    else:
        messages.error(request, result.get('message'))
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.META.get('HTTP_ACCEPT', ''):
            return JsonResponse(result, status=400)
        return redirect('purchase_order_view', pk=pk)


def purchase_order_delete(request, pk):
    if not request.user.is_authenticated or not getattr(request.user, 'company', None):
        messages.error(request, "Authentication required.")
        return redirect('login')

    company = request.user.company
    po = get_object_or_404(PurchaseOrder, id=pk, company=company)

    if po.converted_to_purchase_bill:
        messages.error(request, f"Purchase Order {po.po_number} cannot be deleted because it has been converted to Purchase Bill {po.converted_to_purchase_bill.supplier_bill_no}.")
        return redirect('purchase_order_view', pk=po.id)

    if request.method == 'POST' or request.GET.get('confirm') == 'true':
        po_no = po.po_number
        with transaction.atomic():
            po.items.all().delete()
            po.delete()
            log_action(request.user, 'DELETE_PURCHASE_ORDER', 'PURCHASE_ORDER', pk, request=request)
        messages.success(request, f"Purchase Order {po_no} deleted successfully.")
        return redirect('purchase_order_list')

    messages.error(request, "Deletion requires POST confirmation.")
    return redirect('purchase_order_view', pk=po.id)


# --- PURCHASE BILLS ---

class PurchaseBillListView(CompanyRequiredMixin, CompanyQuerySetMixin, PaginationMixin, ListView):
    model = PurchaseBill
    template_name = 'company/purchase_bill_list.html'
    context_object_name = 'bills'

    def get_queryset(self):
        qs = super().get_queryset().order_by('-bill_date', '-id')
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(supplier_bill_no__icontains=search) |
                Q(supplier__name__icontains=search) |
                Q(supplier__gstin__icontains=search) |
                Q(status__icontains=search)
            )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        else:
            qs = qs.exclude(status='CANCELLED')
        return qs


class PurchaseBillCreateView(CompanyRequiredMixin, View):
    def get(self, request):
        company = request.user.company
        suppliers = Supplier.objects.filter(company=company, is_active=True)
        products = Product.objects.filter(company=company, is_active=True)
        warehouses = Warehouse.objects.filter(company=company, is_active=True)
        return render(request, 'company/purchase_bill_add.html', {
            'suppliers': suppliers,
            'products': products,
            'products_json': build_products_json(products),
            'warehouses': warehouses
        })

    @transaction.atomic
    def post(self, request):
        company = request.user.company
        is_ajax = (
            request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in request.META.get('HTTP_ACCEPT', '') or
            request.content_type == 'application/json'
        )

        try:
            if request.content_type == 'application/json' and request.body:
                data = json.loads(request.body)
            else:
                data = request.POST.dict()
                if 'items_json' in request.POST:
                    data['items'] = request.POST['items_json']
        except Exception:
            data = request.POST.dict()

        try:
            supp_id = data.get('supplier_id') or data.get('supplier')
            if not supp_id:
                raise ValueError('Please select a supplier.')

            bill_no = (data.get('supplier_bill_no') or '').strip()
            if not bill_no:
                raise ValueError('Supplier Bill / Invoice number is required.')

            bill_date_raw = data.get('bill_date')
            due_date_raw = data.get('due_date')
            
            try:
                bill_date = datetime.strptime(str(bill_date_raw).strip(), '%Y-%m-%d').date() if bill_date_raw else date.today()
            except (ValueError, TypeError):
                bill_date = date.today()

            try:
                due_date = datetime.strptime(str(due_date_raw).strip(), '%Y-%m-%d').date() if due_date_raw else (bill_date + timedelta(days=30))
            except (ValueError, TypeError):
                due_date = bill_date + timedelta(days=30)

            items_data = data.get('items', [])
            if isinstance(items_data, str):
                try:
                    items_data = json.loads(items_data)
                except Exception:
                    items_data = []

            if not items_data:
                raise ValueError('Please select at least one valid product in the items grid.')

            wh_id = data.get('warehouse_id')
            warehouse = None
            if wh_id:
                try:
                    warehouse = Warehouse.objects.filter(id=int(wh_id), company=company).first()
                except (ValueError, TypeError):
                    pass
            if not warehouse:
                warehouse = Warehouse.objects.filter(company=company, is_active=True).first() or Warehouse.objects.filter(company=company).first()
            if not warehouse:
                warehouse = Warehouse.objects.create(company=company, name="Main Warehouse", code="WH-MAIN")

            supplier = get_object_or_404(Supplier, id=int(str(supp_id).strip()), company=company)
            
            # Check duplicate purchase bill
            if PurchaseBill.objects.filter(company=company, supplier_bill_no=bill_no, supplier=supplier).exists():
                raise ValueError(f"Purchase bill number '{bill_no}' already exists for supplier '{supplier.name}'!")
                
            bill = PurchaseBill.objects.create(
                company=company, supplier=supplier, supplier_bill_no=bill_no,
                bill_date=bill_date, due_date=due_date, status='POSTED'
            )
            
            created_items_count = 0
            for item in items_data:
                prod_id = item.get('product_id') or item.get('product')
                if not prod_id:
                    continue
                prod = get_object_or_404(Product, id=int(str(prod_id).strip()), company=company)
                qty = parse_money(item.get('quantity', 1))
                rate = parse_money(item.get('rate', 0))
                disc = parse_money(item.get('discount', 0))
                
                if qty <= Decimal('0.00'):
                    continue

                hsn_code = prod.hsn_sac.code if prod.hsn_sac else ''
                gst_rate = prod.hsn_sac.gst_rate if prod.hsn_sac else Decimal('0.00')
                PurchaseBillItem.objects.create(
                    purchase_bill=bill, product=prod, quantity=qty, rate=rate,
                    discount=disc, taxable_value=Decimal('0.00'), hsn_sac_code=hsn_code,
                    gst_rate=gst_rate, total_amount=Decimal('0.00')
                )
                
                # Inwards stock log
                if prod.track_inventory:
                    StockMovement.objects.create(
                        company=company, product=prod, warehouse=warehouse,
                        quantity=qty, movement_type='PURCHASE', reference_id=bill.id,
                        reference_no=bill.supplier_bill_no, created_by=request.user
                    )
                    update_product_stock(prod.id)
                created_items_count += 1

            if created_items_count == 0:
                raise ValueError("At least one valid product item is required in the purchase bill.")

            recalculate_purchase_totals(bill)
            
            # Supplier outstanding payable
            supplier.outstanding_balance += bill.grand_total
            supplier.save()
            
            record_purchase_accounting(bill)

            if request.FILES:
                for file_key, upload_file in request.FILES.items():
                    if upload_file:
                        filename = upload_file.name
                        ext = os.path.splitext(filename)[1].lower().strip('.')
                        ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png', 'webp', 'doc', 'docx', 'xls', 'xlsx'}
                        DANGEROUS_EXTENSIONS = {'exe', 'bat', 'sh', 'js', 'py', 'php', 'cmd', 'vbs', 'ps1', 'cgi', 'pl', 'jar'}
                        if ext not in DANGEROUS_EXTENSIONS and ext in ALLOWED_EXTENSIONS and upload_file.size <= 10 * 1024 * 1024:
                            PurchaseBillDocument.objects.create(
                                purchase_bill=bill,
                                file=upload_file,
                                file_name=filename,
                                file_type=ext.upper(),
                                file_size=upload_file.size,
                                uploaded_by=request.user if request.user and request.user.is_authenticated else None
                            )
            
            log_action(request.user, 'CREATE_PURCHASE_BILL', 'PURCHASE_BILL', bill.id, request=request)
            messages.success(request, f"Purchase bill {bill.supplier_bill_no} posted successfully!")
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'status': 'success',
                    'message': f"Purchase bill {bill.supplier_bill_no} posted successfully!",
                    'bill_id': bill.id,
                    'redirect_url': f"/company/purchase-bills/{bill.id}/"
                }, encoder=DjangoJSONEncoder)
            return redirect('purchase_bill_view', pk=bill.id)
        except Exception as e:
            messages.error(request, str(e))
            if is_ajax:
                return JsonResponse({'success': False, 'status': 'error', 'message': str(e)}, status=400, encoder=DjangoJSONEncoder)
            suppliers = Supplier.objects.filter(company=company, is_active=True)
            products = Product.objects.filter(company=company, is_active=True)
            warehouses = Warehouse.objects.filter(company=company, is_active=True)
            return render(request, 'company/purchase_bill_add.html', {
                'suppliers': suppliers,
                'products': products,
                'products_json': build_products_json(products),
                'warehouses': warehouses,
                'form_data': data
            })


class PurchaseBillDetailView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = PurchaseBill
    template_name = 'company/purchase_bill_detail.html'
    context_object_name = 'bill'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .utils import build_hsn_sac_tax_summary, recalculate_purchase_totals
        bill = self.object
        recalculate_purchase_totals(bill)
        company = bill.company
        supplier = bill.supplier
        company_state_code = str(company.state_code or '').strip().zfill(2)
        supplier_state_code = str(supplier.state_code if supplier else company_state_code).strip().zfill(2)
        summary_list, total_qty = build_hsn_sac_tax_summary(bill.items.all(), company_state_code, supplier_state_code)
        context['hsn_summary'] = summary_list
        context['total_quantity'] = total_qty
        context['is_interstate'] = (company_state_code != supplier_state_code)
        return context


class PurchaseBillPDFView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = PurchaseBill
    template_name = 'company/purchase_bill_pdf.html'
    context_object_name = 'bill'

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        context = self.get_context_data(object=self.object)
        from .utils import build_hsn_sac_tax_summary, recalculate_purchase_totals
        bill = self.object
        recalculate_purchase_totals(bill)
        company = bill.company
        supplier = bill.supplier
        
        company_state_code = str(company.state_code or '').strip().zfill(2)
        supplier_state_code = str(supplier.state_code if supplier else company_state_code).strip().zfill(2)
        is_interstate = (company_state_code != supplier_state_code)
        
        summary_list, total_qty = build_hsn_sac_tax_summary(bill.items.all(), company_state_code, supplier_state_code)
        raw_terms = company.terms_and_conditions or ""
        terms_list = [t.strip() for t in raw_terms.splitlines() if t.strip()]
        
        context.update({
            'bill': bill,
            'company': company,
            'supplier': supplier,
            'is_interstate': is_interstate,
            'hsn_summary': summary_list,
            'total_quantity': total_qty,
            'terms_list': terms_list,
        })
        return self.render_to_response(context)


def warehouse_detail_api(request, pk):
    if not request.user.is_authenticated or not getattr(request.user, 'company', None):
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
    company = request.user.company
    try:
        w = Warehouse.objects.get(id=pk, company=company, is_active=True)
        return JsonResponse({
            'status': 'success',
            'warehouse': {
                'id': w.id,
                'name': w.name,
                'code': w.code or '',
                'manager': w.manager or 'Not specified',
                'contact': w.contact or 'Not specified',
                'address': w.address or 'No address recorded',
                'city': getattr(w, 'city', company.city or ''),
                'state': getattr(w, 'state', company.state or ''),
                'state_code': getattr(w, 'state_code', company.state_code or ''),
                'pincode': getattr(w, 'pincode', company.pincode or ''),
                'gstin': getattr(w, 'gstin', company.gstin or ''),
            }
        })
    except Warehouse.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Unable to load warehouse details.'}, status=404)


def purchase_bill_upload_document(request, pk):
    if not request.user.is_authenticated or not getattr(request.user, 'company', None):
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST request required.'}, status=405)

    company = request.user.company
    bill = get_object_or_404(PurchaseBill, id=pk, company=company)

    upload_file = request.FILES.get('file') or request.FILES.get('document_file')
    if not upload_file:
        return JsonResponse({'status': 'error', 'message': 'No file uploaded.'}, status=400)

    filename = upload_file.name
    ext = os.path.splitext(filename)[1].lower().strip('.')
    
    ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png', 'webp', 'doc', 'docx', 'xls', 'xlsx'}
    DANGEROUS_EXTENSIONS = {'exe', 'bat', 'sh', 'js', 'py', 'php', 'cmd', 'vbs', 'ps1', 'cgi', 'pl', 'jar'}

    if ext in DANGEROUS_EXTENSIONS or ext not in ALLOWED_EXTENSIONS:
        return JsonResponse({'status': 'error', 'message': f'Invalid file format (.{ext}). Allowed formats: PDF, JPG, PNG, WEBP, DOC, DOCX, XLS, XLSX.'}, status=400)

    MAX_SIZE = 10 * 1024 * 1024  # 10MB
    if upload_file.size > MAX_SIZE:
        return JsonResponse({'status': 'error', 'message': 'File size exceeds maximum limit of 10MB.'}, status=400)

    try:
        doc = PurchaseBillDocument.objects.create(
            purchase_bill=bill,
            file=upload_file,
            file_name=filename,
            file_type=ext.upper(),
            file_size=upload_file.size,
            uploaded_by=request.user
        )
        log_action(request.user, 'UPLOAD_PURCHASE_BILL_DOC', 'PURCHASE_BILL', bill.id, request=request)
        return JsonResponse({
            'status': 'success',
            'message': 'Document uploaded successfully.',
            'document': {
                'id': doc.id,
                'file_name': doc.file_name,
                'file_type': doc.file_type,
                'file_size': doc.file_size,
                'uploaded_by': doc.uploaded_by.get_full_name() or doc.uploaded_by.username if doc.uploaded_by else 'User',
                'uploaded_at': doc.uploaded_at.strftime('%d %b %Y %I:%M %p'),
                'view_url': f'/company/purchase-bills/documents/{doc.id}/view/',
                'delete_url': f'/company/purchase-bills/documents/{doc.id}/delete/'
            }
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Unable to upload document: {str(e)}'}, status=500)


def purchase_bill_delete_document(request, doc_id):
    if not request.user.is_authenticated or not getattr(request.user, 'company', None):
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
    if request.method not in ['POST', 'DELETE']:
        return JsonResponse({'status': 'error', 'message': 'POST/DELETE request required.'}, status=405)

    company = request.user.company
    doc = get_object_or_404(PurchaseBillDocument, id=doc_id, purchase_bill__company=company)

    try:
        doc_name = doc.file_name
        bill_id = doc.purchase_bill.id
        try:
            if doc.file and hasattr(doc.file, 'path') and os.path.isfile(doc.file.path):
                os.remove(doc.file.path)
        except Exception:
            pass
        doc.delete()
        log_action(request.user, 'DELETE_PURCHASE_BILL_DOC', 'PURCHASE_BILL', bill_id, request=request)
        return JsonResponse({'status': 'success', 'message': f'Document "{doc_name}" deleted successfully.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Unable to delete document: {str(e)}'}, status=500)


def purchase_bill_view_document(request, doc_id):
    if not request.user.is_authenticated or not getattr(request.user, 'company', None):
        raise PermissionDenied("Authentication required.")

    company = request.user.company
    doc = get_object_or_404(PurchaseBillDocument, id=doc_id, purchase_bill__company=company)

    if not doc.file or not os.path.isfile(doc.file.path):
        raise Http404("Requested document file does not exist on disk.")

    mime_type, _ = mimetypes.guess_type(doc.file.path)
    if not mime_type:
        mime_type = 'application/octet-stream'

    ext = doc.file_type.lower()
    is_previewable = ext in ['pdf', 'jpg', 'jpeg', 'png', 'webp']

    response = FileResponse(open(doc.file.path, 'rb'), content_type=mime_type)
    disposition = 'inline' if is_previewable else 'attachment'
    response['Content-Disposition'] = f'{disposition}; filename="{doc.file_name}"'
    return response


@transaction.atomic
def purchase_bill_cancel(request, pk):
    company = request.user.company
    bill = get_object_or_404(PurchaseBill, id=pk, company=company)
    
    if bill.status == 'CANCELLED':
        messages.error(request, "Purchase bill is already cancelled.")
        return redirect('purchase_bill_view', pk=bill.id)
        
    # Reverse inventory
    movements = StockMovement.objects.filter(company=company, reference_id=bill.id, movement_type='PURCHASE')
    for mv in movements:
        StockMovement.objects.create(
            company=company, product=mv.product, warehouse=mv.warehouse,
            quantity=-mv.quantity, movement_type='PURCHASE_RETURN', reference_id=bill.id,
            reference_no=f"CNL: {bill.supplier_bill_no}", created_by=request.user
        )
        update_product_stock(mv.product.id)
        
    # Reverse outstanding
    supplier = bill.supplier
    supplier.outstanding_balance -= bill.grand_total
    supplier.save()
    
    cancel_purchase_accounting(bill)
    
    bill.status = 'CANCELLED'
    bill.save()
    
    log_action(request.user, 'CANCEL_PURCHASE_BILL', 'PURCHASE_BILL', bill.id, request=request)
    messages.success(request, f"Purchase bill {bill.supplier_bill_no} cancelled.")
    return redirect('purchase_bill_view', pk=bill.id)


# --- PAYMENTS ---

class PaymentListView(CompanyRequiredMixin, CompanyQuerySetMixin, PaginationMixin, ListView):
    model = Payment
    template_name = 'company/payment_list.html'
    context_object_name = 'payments'

    def get_queryset(self):
        qs = super().get_queryset().order_by('-payment_date', '-id')
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(id__icontains=search) |
                Q(reference_no__icontains=search) |
                Q(customer__name__icontains=search) |
                Q(supplier__name__icontains=search) |
                Q(payment_method__icontains=search) |
                Q(payment_type__icontains=search)
            )
        return qs


class PaymentReceiptCreateView(CompanyRequiredMixin, CreateView):
    model = Payment
    fields = ['customer', 'invoice', 'amount', 'payment_date', 'payment_method', 'reference_no', 'notes']
    template_name = 'company/payment_receipt_add.html'
    success_url = reverse_lazy('payment_list')

    def is_ajax(self):
        return (
            self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in self.request.META.get('HTTP_ACCEPT', '') or
            self.request.content_type == 'application/json'
        )

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        company = self.request.user.company
        form.fields['customer'].queryset = Customer.objects.filter(company=company)
        form.fields['invoice'].queryset = Invoice.objects.filter(company=company).exclude(status__in=['PAID', 'CANCELLED'])
        form.fields['payment_date'].widget = forms.DateInput(attrs={'type': 'date', 'class': 'form-control'})
        form.fields['payment_date'].initial = timezone.now().date()
        for name, field in form.fields.items():
            if name != 'payment_date':
                field.widget.attrs['class'] = 'form-control'
        return form

    def form_invalid(self, form):
        if self.is_ajax():
            errors = []
            for field, err_list in form.errors.items():
                for err in err_list:
                    errors.append(f"{field.replace('_', ' ').capitalize()}: {err}")
            msg = "; ".join(errors) if errors else "Please fill all required fields correctly."
            return JsonResponse({'status': 'error', 'message': msg}, status=400)
        return super().form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        company = self.request.user.company
        form.instance.company = company
        form.instance.payment_type = 'RECEIPT'
        
        # Check customer/invoice correlation
        if not form.instance.customer and form.instance.invoice:
            form.instance.customer = form.instance.invoice.customer
            
        if not form.instance.customer and not form.instance.invoice:
            form.add_error('customer', 'Please select a customer or an invoice.')
            return self.form_invalid(form)

        if form.instance.amount <= Decimal('0.00'):
            form.add_error('amount', 'Payment amount must be greater than zero.')
            return self.form_invalid(form)

        response = super().form_valid(form)
        payment = self.object
        
        # Deduct from customer receivable balance
        customer = payment.customer
        if customer:
            customer.outstanding_balance -= payment.amount
            customer.save()
            
            # Customer Ledger entry
            CustomerLedger.objects.create(
                company=company,
                customer=customer,
                date=payment.payment_date,
                entry_type='PAYMENT',
                reference_id=payment.id,
                reference_no=f"RCPT #{payment.id}",
                description=f"Payment received via {payment.get_payment_method_display()} (Ref: {payment.reference_no or 'N/A'})",
                credit=payment.amount,
                debit=Decimal('0.00'),
                running_balance=customer.outstanding_balance
            )
            
        # Deduct from invoice balance if selected
        invoice = payment.invoice
        if invoice:
            invoice.paid_amount += payment.amount
            if invoice.paid_amount >= invoice.grand_total:
                invoice.status = 'PAID'
            else:
                invoice.status = 'PARTIALLY_PAID'
            invoice.save()
            
        record_payment_accounting(payment)
        log_action(self.request.user, 'RECEIVE_PAYMENT', 'PAYMENT', payment.id, request=self.request)
        messages.success(self.request, "Payment receipt logged successfully!")
        
        if self.is_ajax():
            return JsonResponse({
                'status': 'success',
                'message': 'Payment receipt logged successfully!',
                'payment_id': payment.id
            })
        return response


class PaymentSupplierCreateView(CompanyRequiredMixin, CreateView):
    model = Payment
    fields = ['supplier', 'purchase_bill', 'amount', 'payment_date', 'payment_method', 'reference_no', 'notes']
    template_name = 'company/payment_supplier_add.html'
    success_url = reverse_lazy('payment_list')

    def is_ajax(self):
        return (
            self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in self.request.META.get('HTTP_ACCEPT', '') or
            self.request.content_type == 'application/json'
        )

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        company = self.request.user.company
        form.fields['supplier'].queryset = Supplier.objects.filter(company=company)
        form.fields['purchase_bill'].queryset = PurchaseBill.objects.filter(company=company).exclude(status__in=['PAID', 'CANCELLED'])
        form.fields['payment_date'].widget = forms.DateInput(attrs={'type': 'date', 'class': 'form-control'})
        form.fields['payment_date'].initial = timezone.now().date()
        for name, field in form.fields.items():
            if name != 'payment_date':
                field.widget.attrs['class'] = 'form-control'
        return form

    def form_invalid(self, form):
        if self.is_ajax():
            errors = []
            for field, err_list in form.errors.items():
                for err in err_list:
                    errors.append(f"{field.replace('_', ' ').capitalize()}: {err}")
            msg = "; ".join(errors) if errors else "Please fill all required fields correctly."
            return JsonResponse({'status': 'error', 'message': msg}, status=400)
        return super().form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        company = self.request.user.company
        form.instance.company = company
        form.instance.payment_type = 'PAYMENT'
        
        # Check supplier/bill correlation
        if not form.instance.supplier and form.instance.purchase_bill:
            form.instance.supplier = form.instance.purchase_bill.supplier

        if not form.instance.supplier and not form.instance.purchase_bill:
            form.add_error('supplier', 'Please select a supplier or a purchase bill.')
            return self.form_invalid(form)

        if form.instance.amount <= Decimal('0.00'):
            form.add_error('amount', 'Payment amount must be greater than zero.')
            return self.form_invalid(form)

        response = super().form_valid(form)
        payment = self.object
        
        # Deduct from supplier payable balance
        supplier = payment.supplier
        if supplier:
            supplier.outstanding_balance -= payment.amount
            supplier.save()
            
            # Supplier Ledger entry
            SupplierLedger.objects.create(
                company=company,
                supplier=supplier,
                date=payment.payment_date,
                entry_type='PAYMENT',
                reference_id=payment.id,
                reference_no=f"PAY #{payment.id}",
                description=f"Payment made via {payment.get_payment_method_display()} (Ref: {payment.reference_no or 'N/A'})",
                debit=payment.amount,
                credit=Decimal('0.00'),
                running_balance=supplier.outstanding_balance
            )
            
        # Deduct from bill balance if selected
        bill = payment.purchase_bill
        if bill:
            bill.paid_amount += payment.amount
            if bill.paid_amount >= bill.grand_total:
                bill.status = 'PAID'
            else:
                bill.status = 'PARTIALLY_PAID'
            bill.save()
            
        record_payment_accounting(payment)
        log_action(self.request.user, 'SUPPLIER_PAYMENT', 'PAYMENT', payment.id, request=self.request)
        messages.success(self.request, "Supplier payment logged successfully!")

        if self.is_ajax():
            return JsonResponse({
                'success': True,
                'status': 'success',
                'message': 'Supplier payment saved successfully!',
                'payment_id': payment.id
            }, encoder=DjangoJSONEncoder)
        return response


def api_customer_unpaid_invoices(request, pk):
    if not request.user.is_authenticated or not request.user.company:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
    invoices = Invoice.objects.filter(
        company=request.user.company,
        customer_id=pk
    ).exclude(status__in=['PAID', 'CANCELLED']).order_by('-invoice_date')
    data = []
    for inv in invoices:
        data.append({
            'id': inv.id,
            'invoice_number': inv.invoice_number,
            'grand_total': float(inv.grand_total),
            'paid_amount': float(inv.paid_amount),
            'outstanding_amount': float(inv.outstanding_amount())
        })
    return JsonResponse({'status': 'success', 'invoices': data})


def api_supplier_unpaid_bills(request, pk):
    if not request.user.is_authenticated or not request.user.company:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
    bills = PurchaseBill.objects.filter(
        company=request.user.company,
        supplier_id=pk
    ).exclude(status__in=['PAID', 'CANCELLED']).order_by('-bill_date')
    data = []
    for pb in bills:
        data.append({
            'id': pb.id,
            'bill_number': pb.supplier_bill_no or f"BILL-{pb.id}",
            'grand_total': float(pb.grand_total),
            'paid_amount': float(pb.paid_amount),
            'outstanding_amount': float(pb.outstanding_amount())
        })
    return JsonResponse({'status': 'success', 'bills': data})


# --- EXPENSES ---

class ExpenseListView(CompanyRequiredMixin, CompanyQuerySetMixin, PaginationMixin, ListView):
    model = Expense
    template_name = 'company/expense_list.html'
    context_object_name = 'expenses'

    def get_queryset(self):
        qs = super().get_queryset().order_by('-created_at', '-id')
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(id__icontains=search) |
                Q(title__icontains=search) |
                Q(vendor__icontains=search) |
                Q(category__name__icontains=search) |
                Q(reference_no__icontains=search)
            )
        return qs


class ExpenseCreateView(CompanyRequiredMixin, CreateView):
    model = Expense
    form_class = ExpenseForm
    template_name = 'company/expense_add.html'
    success_url = reverse_lazy('expense_list')

    def is_ajax(self):
        return (
            self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in self.request.META.get('HTTP_ACCEPT', '') or
            self.request.content_type == 'application/json'
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.user.company
        return kwargs

    def form_valid(self, form):
        form.instance.company = self.request.user.company
        amt = Decimal(form.instance.amount)
        rate = Decimal(form.instance.gst_rate)
        form.instance.gst_amount = quantize_amount(amt * (rate / Decimal('100.00')))
        response = super().form_valid(form)
        log_action(self.request.user, 'CREATE_EXPENSE', 'EXPENSE', self.object.id, new_values=form.cleaned_data, request=self.request)
        messages.success(self.request, "Overhead expense logged successfully!")
        if self.is_ajax():
            return JsonResponse({
                'success': True,
                'status': 'success',
                'message': "Overhead expense logged successfully!",
                'expense_id': self.object.id,
                'redirect_url': str(self.success_url)
            }, encoder=DjangoJSONEncoder)
        return response

    def form_invalid(self, form):
        if self.is_ajax():
            errors = [f"{field.capitalize()}: {errs[0]}" for field, errs in form.errors.items()]
            return JsonResponse({'success': False, 'status': 'error', 'message': ', '.join(errors)}, status=400, encoder=DjangoJSONEncoder)
        return super().form_invalid(form)


# --- GST REPORTS ---

class GSTCheckDetailsView(CompanyRequiredMixin, TemplateView):
    template_name = 'company/gst_check_details.html'


class GSTDashboardView(CompanyRequiredMixin, TemplateView):
    template_name = 'company/gst_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.user.company

        outward = GSTTransaction.objects.filter(
            company=company,
            transaction_type__in=['SALES', 'CREDIT_NOTE'],
            is_cancelled=False
        ).aggregate(
            cgst=Sum('cgst_amount'), sgst=Sum('sgst_amount'), igst=Sum('igst_amount'), total=Sum('taxable_value')
        )
        
        inward = GSTTransaction.objects.filter(
            company=company,
            transaction_type__in=['PURCHASE', 'DEBIT_NOTE'],
            is_cancelled=False
        ).aggregate(
            cgst=Sum('cgst_amount'), sgst=Sum('sgst_amount'), igst=Sum('igst_amount'), total=Sum('taxable_value')
        )

        out_cgst = quantize_amount(outward['cgst'] or Decimal('0.00'))
        out_sgst = quantize_amount(outward['sgst'] or Decimal('0.00'))
        out_igst = quantize_amount(outward['igst'] or Decimal('0.00'))
        out_taxable = quantize_amount(outward['total'] or Decimal('0.00'))

        # Fallback to posted Invoice records if GSTTransaction total is 0
        if (out_cgst + out_sgst + out_igst + out_taxable) == Decimal('0.00'):
            inv_agg = Invoice.objects.filter(company=company, status='POSTED').aggregate(
                cgst=Sum('cgst_total'), sgst=Sum('sgst_total'), igst=Sum('igst_total'), total=Sum('taxable_value')
            )
            out_cgst = quantize_amount(inv_agg['cgst'] or Decimal('0.00'))
            out_sgst = quantize_amount(inv_agg['sgst'] or Decimal('0.00'))
            out_igst = quantize_amount(inv_agg['igst'] or Decimal('0.00'))
            out_taxable = quantize_amount(inv_agg['total'] or Decimal('0.00'))

        in_cgst = quantize_amount(inward['cgst'] or Decimal('0.00'))
        in_sgst = quantize_amount(inward['sgst'] or Decimal('0.00'))
        in_igst = quantize_amount(inward['igst'] or Decimal('0.00'))
        in_taxable = quantize_amount(inward['total'] or Decimal('0.00'))

        # Fallback to posted PurchaseBill records if GSTTransaction total is 0
        if (in_cgst + in_sgst + in_igst + in_taxable) == Decimal('0.00'):
            bill_agg = PurchaseBill.objects.filter(company=company, status='POSTED').aggregate(
                cgst=Sum('cgst_total'), sgst=Sum('sgst_total'), igst=Sum('igst_total'), total=Sum('taxable_value')
            )
            in_cgst = quantize_amount(bill_agg['cgst'] or Decimal('0.00'))
            in_sgst = quantize_amount(bill_agg['sgst'] or Decimal('0.00'))
            in_igst = quantize_amount(bill_agg['igst'] or Decimal('0.00'))
            in_taxable = quantize_amount(bill_agg['total'] or Decimal('0.00'))

        output_tax = quantize_amount(out_cgst + out_sgst + out_igst)
        input_tax = quantize_amount(in_cgst + in_sgst + in_igst)
        net_payable = quantize_amount(max(Decimal('0.00'), output_tax - input_tax))

        context['out_cgst'] = f"{out_cgst:.2f}"
        context['out_sgst'] = f"{out_sgst:.2f}"
        context['out_igst'] = f"{out_igst:.2f}"
        context['out_taxable'] = f"{out_taxable:.2f}"
        context['output_tax'] = f"{output_tax:.2f}"

        context['in_cgst'] = f"{in_cgst:.2f}"
        context['in_sgst'] = f"{in_sgst:.2f}"
        context['in_igst'] = f"{in_igst:.2f}"
        context['in_taxable'] = f"{in_taxable:.2f}"
        context['input_tax'] = f"{input_tax:.2f}"

        context['net_payable'] = f"{net_payable:.2f}"
        return context


class GSTR1View(CompanyRequiredMixin, TemplateView):
    template_name = 'company/gst_report_gstr1.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.user.company

        b2b = GSTTransaction.objects.filter(
            company=company,
            transaction_type__in=['SALES', 'CREDIT_NOTE'],
            is_cancelled=False
        ).exclude(gstin__in=[None, ''])
        context['b2b_invoices'] = b2b
        context['b2b_totals'] = b2b.aggregate(
            taxable=Sum('taxable_value'), cgst=Sum('cgst_amount'),
            sgst=Sum('sgst_amount'), igst=Sum('igst_amount'), count=Count('id')
        )

        b2c = GSTTransaction.objects.filter(
            company=company,
            transaction_type__in=['SALES', 'CREDIT_NOTE'],
            is_cancelled=False,
            gstin__in=[None, '']
        )
        context['b2c_invoices'] = b2c
        context['b2c_totals'] = b2c.aggregate(
            taxable=Sum('taxable_value'), cgst=Sum('cgst_amount'),
            sgst=Sum('sgst_amount'), igst=Sum('igst_amount'), count=Count('id')
        )

        hsn_summary = GSTTransaction.objects.filter(
            company=company,
            transaction_type__in=['SALES', 'CREDIT_NOTE'],
            is_cancelled=False
        ).values('hsn_sac_code', 'uqc_unit', 'gst_rate').annotate(
            total_qty=Sum('quantity'),
            total_taxable=Sum('taxable_value'),
            total_cgst=Sum('cgst_amount'),
            total_sgst=Sum('sgst_amount'),
            total_igst=Sum('igst_amount'),
            total_cess=Sum('cess_amount')
        ).order_by('hsn_sac_code')
        
        context['hsn_summary'] = hsn_summary
        return context


class GSTR3BView(CompanyRequiredMixin, TemplateView):
    template_name = 'company/gst_report_gstr3b.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.user.company

        sales = GSTTransaction.objects.filter(
            company=company,
            transaction_type__in=['SALES', 'CREDIT_NOTE'],
            is_cancelled=False
        ).aggregate(
            val=Sum('taxable_value'), cgst=Sum('cgst_amount'), sgst=Sum('sgst_amount'), igst=Sum('igst_amount')
        )
        context['sales_val'] = sales['val'] or Decimal('0.00')
        context['sales_cgst'] = sales['cgst'] or Decimal('0.00')
        context['sales_sgst'] = sales['sgst'] or Decimal('0.00')
        context['sales_igst'] = sales['igst'] or Decimal('0.00')

        purchases = GSTTransaction.objects.filter(
            company=company,
            transaction_type__in=['PURCHASE', 'DEBIT_NOTE'],
            is_cancelled=False
        ).aggregate(
            val=Sum('taxable_value'), cgst=Sum('cgst_amount'), sgst=Sum('sgst_amount'), igst=Sum('igst_amount')
        )
        context['itc_val'] = purchases['val'] or Decimal('0.00')
        context['itc_cgst'] = purchases['cgst'] or Decimal('0.00')
        context['itc_sgst'] = purchases['sgst'] or Decimal('0.00')
        context['itc_igst'] = purchases['igst'] or Decimal('0.00')

        return context


# --- STUB DUMMY/FALLBACK FOR OMITTED VIEWS ---

class CreditNoteListView(CompanyRequiredMixin, CompanyQuerySetMixin, PaginationMixin, ListView):
    model = CreditNote
    template_name = 'company/credit_note_list.html'
    context_object_name = 'credit_notes'

    def get_queryset(self):
        qs = super().get_queryset().order_by('-note_date', '-id')
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(note_number__icontains=search) |
                Q(invoice__invoice_number__icontains=search) |
                Q(customer__name__icontains=search)
            )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        else:
            qs = qs.exclude(status='CANCELLED')
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.forms import modelform_factory
        CreditNoteForm = modelform_factory(CreditNote, fields=['invoice', 'note_number', 'note_date', 'reason', 'subtotal', 'notes'])
        form = CreditNoteForm()
        form.fields['invoice'].queryset = Invoice.objects.filter(company=self.request.user.company, status='POSTED')
        for name, field in form.fields.items():
            field.widget.attrs['class'] = 'form-control'
            field.widget.attrs['id'] = f'cn-{name}'
        
        company = self.request.user.company
        last_note = CreditNote.objects.filter(company=company).order_by('-id').first()
        if last_note:
            try:
                import re
                num_match = re.search(r'\d+', last_note.note_number)
                if num_match:
                    num = int(num_match.group())
                    next_no = last_note.note_number.replace(num_match.group(), str(num+1).zfill(len(num_match.group())))
                else:
                    next_no = f"CN-{CreditNote.objects.filter(company=company).count() + 1:05d}"
            except Exception:
                next_no = f"CN-{CreditNote.objects.filter(company=company).count() + 1:05d}"
        else:
            next_no = "CN-00001"
        form.initial['note_number'] = next_no
        form.initial['note_date'] = timezone.now().date()
        context['form'] = form
        return context


class CreditNoteCreateView(CompanyRequiredMixin, CreateView):
    model = CreditNote
    fields = ['invoice', 'note_number', 'note_date', 'reason', 'subtotal', 'notes']
    template_name = 'company/credit_note_add.html'
    success_url = reverse_lazy('credit_note_list')

    def get_initial(self):
        initial = super().get_initial()
        company = self.request.user.company
        count = CreditNote.objects.filter(company=company).count() + 1
        initial['note_number'] = f"CN-{company.financial_year}-{str(count).zfill(5)}"
        initial['note_date'] = date.today().strftime('%Y-%m-%d')
        return initial

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        company = self.request.user.company
        count = CreditNote.objects.filter(company=company).count() + 1
        default_no = f"CN-{company.financial_year}-{str(count).zfill(5)}"
        default_date = date.today().strftime('%Y-%m-%d')
        
        if self.request.content_type == 'application/json':
            try:
                import json
                data = json.loads(self.request.body)
                if not data.get('note_number'):
                    data['note_number'] = default_no
                if not data.get('note_date'):
                    data['note_date'] = default_date
                if 'subtotal' in data:
                    try:
                        data['subtotal'] = str(parse_money(data['subtotal']))
                    except Exception:
                        pass
                kwargs['data'] = data
            except Exception:
                pass
        elif self.request.method in ('POST', 'PUT'):
            post_data = self.request.POST.copy()
            if not post_data.get('note_number'):
                post_data['note_number'] = default_no
            if not post_data.get('note_date'):
                post_data['note_date'] = default_date
            if 'subtotal' in post_data:
                try:
                    post_data['subtotal'] = str(parse_money(post_data['subtotal']))
                except Exception:
                    pass
            kwargs['data'] = post_data
        return kwargs

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields['invoice'].queryset = Invoice.objects.filter(company=self.request.user.company, status='POSTED')
        for name, field in form.fields.items():
            field.widget.attrs['class'] = 'form-control'
        return form

    @transaction.atomic
    def form_valid(self, form):
        company = self.request.user.company
        form.instance.company = company
        original_invoice = form.instance.invoice
        subtotal = form.instance.subtotal
        
        # Validation: Amount must not exceed original invoice taxable value
        if subtotal > original_invoice.taxable_value:
            msg = f"Credit note taxable value (₹{subtotal}) cannot exceed original invoice taxable value (₹{original_invoice.taxable_value})."
            if self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or self.request.content_type == 'application/json':
                return JsonResponse({'status': 'error', 'message': msg}, status=400)
            form.add_error('subtotal', msg)
            return self.form_invalid(form)

        # Calculate GST rate from invoice first item or default
        gst_rate = Decimal('18.00')
        if original_invoice.items.exists():
            gst_rate = original_invoice.items.first().gst_rate
            
        cgst, sgst, igst, total_gst = calculate_item_gst(
            company.state_code,
            original_invoice.place_of_supply_code,
            subtotal,
            gst_rate
        )
        
        form.instance.taxable_value = subtotal
        form.instance.discount_total = Decimal('0.00')
        form.instance.cess_total = Decimal('0.00')
        form.instance.cgst_total = cgst
        form.instance.sgst_total = sgst
        form.instance.igst_total = igst
        
        calculated_grand = subtotal + cgst + sgst + igst
        rounded_grand = quantize_amount(calculated_grand.quantize(Decimal('1.'), rounding=ROUND_HALF_UP))
        form.instance.round_off = quantize_amount(rounded_grand - calculated_grand)
        form.instance.grand_total = rounded_grand
        form.instance.status = 'POSTED'
        
        response = super().form_valid(form)
        note = self.object
        recalculate_credit_note_totals(note)
        
        # Adjust customer receivable outstanding balance
        customer = original_invoice.customer
        customer.outstanding_balance -= note.grand_total
        customer.save()
        
        # Adjust stock if Sales Return
        if note.reason == 'SALES_RETURN':
            ratio = subtotal / original_invoice.taxable_value if original_invoice.taxable_value > 0 else Decimal('1')
            wh = Warehouse.objects.filter(company=company, is_active=True).first() or Warehouse.objects.filter(company=company).first()
            for item in original_invoice.items.all():
                returned_qty = quantize_amount(item.quantity * ratio)
                if returned_qty > 0 and item.product.track_inventory:
                    StockMovement.objects.create(
                        company=company,
                        product=item.product,
                        warehouse=wh,
                        quantity=returned_qty,
                        movement_type='SALES_RETURN',
                        reference_id=note.id,
                        reference_no=note.note_number,
                        created_by=self.request.user
                    )
                    update_product_stock(item.product.id)
                    
        # Record accounting
        record_credit_note_accounting(note)
        
        log_action(self.request.user, 'CREATE_CREDIT_NOTE', 'CREDIT_NOTE', note.id, request=self.request)
        messages.success(self.request, "Credit note issued successfully!")
        
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or self.request.content_type == 'application/json':
            return JsonResponse({
                'success': True,
                'status': 'success',
                'message': "Credit note issued successfully!",
                'note_id': note.id,
                'redirect_url': str(self.success_url)
            }, encoder=DjangoJSONEncoder)
        return response

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or self.request.content_type == 'application/json':
            errors = []
            for field, errs in form.errors.items():
                errors.append(f"{field.capitalize()}: {errs[0]}")
            return JsonResponse({'success': False, 'status': 'error', 'message': ', '.join(errors)}, status=400, encoder=DjangoJSONEncoder)
        return super().form_invalid(form)


class CreditNoteDetailView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = CreditNote
    template_name = 'company/credit_note_detail.html'
    context_object_name = 'note'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .utils import build_hsn_sac_tax_summary, recalculate_credit_note_totals
        note = self.object
        recalculate_credit_note_totals(note)
        company = note.company
        customer = note.invoice.customer
        company_state_code = str(company.state_code or '').strip().zfill(2)
        pos_code = str(note.invoice.place_of_supply_code or customer.billing_state_code or company_state_code).strip().zfill(2)
        summary_list, total_qty = build_hsn_sac_tax_summary(note.items.all(), company_state_code, pos_code)
        context['hsn_summary'] = summary_list
        context['total_quantity'] = total_qty
        context['is_interstate'] = (company_state_code != pos_code)
        return context


class CreditNotePDFView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = CreditNote
    template_name = 'company/credit_note_pdf.html'
    context_object_name = 'note'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .utils import build_hsn_sac_tax_summary, recalculate_credit_note_totals
        note = self.object
        recalculate_credit_note_totals(note)
        company = note.company
        customer = note.invoice.customer
        company_state_code = str(company.state_code or '').strip().zfill(2)
        pos_code = str(note.invoice.place_of_supply_code or customer.billing_state_code or company_state_code).strip().zfill(2)
        summary_list, total_qty = build_hsn_sac_tax_summary(note.items.all(), company_state_code, pos_code)
        context['hsn_summary'] = summary_list
        context['total_quantity'] = total_qty
        context['is_interstate'] = (company_state_code != pos_code)
        return context


class DebitNoteListView(CompanyRequiredMixin, CompanyQuerySetMixin, PaginationMixin, ListView):
    model = DebitNote
    template_name = 'company/debit_note_list.html'
    context_object_name = 'debit_notes'

    def get_queryset(self):
        qs = super().get_queryset().order_by('-note_date', '-id')
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(note_number__icontains=search) |
                Q(purchase_bill__supplier_bill_no__icontains=search) |
                Q(supplier__name__icontains=search)
            )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        else:
            qs = qs.exclude(status='CANCELLED')
        return qs


class DebitNoteCreateView(CompanyRequiredMixin, CreateView):
    model = DebitNote
    fields = ['purchase_bill', 'note_number', 'note_date', 'reason', 'subtotal', 'notes']
    template_name = 'company/debit_note_add.html'
    success_url = reverse_lazy('debit_note_list')

    def is_ajax(self):
        return (
            self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in self.request.META.get('HTTP_ACCEPT', '') or
            self.request.content_type == 'application/json'
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        company = self.request.user.company
        count = DebitNote.objects.filter(company=company).count() + 1
        default_no = f"DN-{company.financial_year}-{str(count).zfill(5)}"
        default_date = date.today().strftime('%Y-%m-%d')
        
        if self.request.content_type == 'application/json':
            try:
                import json
                data = json.loads(self.request.body)
                if not data.get('note_number'):
                    data['note_number'] = default_no
                if not data.get('note_date'):
                    data['note_date'] = default_date
                kwargs['data'] = data
            except Exception:
                pass
        elif self.request.method in ('POST', 'PUT'):
            post_data = self.request.POST.copy()
            if not post_data.get('note_number'):
                post_data['note_number'] = default_no
            if not post_data.get('note_date'):
                post_data['note_date'] = default_date
            kwargs['data'] = post_data
        return kwargs

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields['purchase_bill'].queryset = PurchaseBill.objects.filter(company=self.request.user.company, status='POSTED')
        for name, field in form.fields.items():
            field.widget.attrs['class'] = 'form-control'
        return form

    @transaction.atomic
    def form_valid(self, form):
        company = self.request.user.company
        form.instance.company = company
        bill = form.instance.purchase_bill
        subtotal = form.instance.subtotal
        
        # Calculate GST rate from bill first item or default
        gst_rate = Decimal('18.00')
        if bill.items.exists():
            gst_rate = bill.items.first().gst_rate
            
        cgst, sgst, igst, total_gst = calculate_item_gst(
            company.state_code,
            bill.supplier.state_code,
            subtotal,
            gst_rate
        )
        
        form.instance.taxable_value = subtotal
        form.instance.discount_total = Decimal('0.00')
        form.instance.cess_total = Decimal('0.00')
        form.instance.cgst_total = cgst
        form.instance.sgst_total = sgst
        form.instance.igst_total = igst
        
        calculated_grand = subtotal + cgst + sgst + igst
        rounded_grand = quantize_amount(calculated_grand.quantize(Decimal('1.'), rounding=ROUND_HALF_UP))
        form.instance.round_off = quantize_amount(rounded_grand - calculated_grand)
        form.instance.grand_total = rounded_grand
        form.instance.status = 'POSTED'
        
        response = super().form_valid(form)
        note = self.object
        recalculate_debit_note_totals(note)
        
        # Adjust supplier outstanding payable balance
        supplier = bill.supplier
        supplier.outstanding_balance -= note.grand_total
        supplier.save()
        
        # Adjust stock if Purchase Return
        if note.reason == 'PURCHASE_RETURN':
            ratio = subtotal / bill.taxable_value if bill.taxable_value > 0 else Decimal('1')
            wh = Warehouse.objects.filter(company=company, is_active=True).first() or Warehouse.objects.filter(company=company).first()
            for item in bill.items.all():
                returned_qty = quantize_amount(item.quantity * ratio)
                if returned_qty > 0 and item.product.track_inventory:
                    StockMovement.objects.create(
                        company=company,
                        product=item.product,
                        warehouse=wh,
                        quantity=-returned_qty,  # Negative: stock leaves warehouse
                        movement_type='PURCHASE_RETURN',
                        reference_id=note.id,
                        reference_no=note.note_number,
                        created_by=self.request.user
                    )
                    update_product_stock(item.product.id)
                    
        # Record accounting
        record_debit_note_accounting(note)
        
        log_action(self.request.user, 'CREATE_DEBIT_NOTE', 'DEBIT_NOTE', note.id, request=self.request)
        messages.success(self.request, "Debit note saved successfully!")
        if self.is_ajax():
            return JsonResponse({
                'success': True,
                'status': 'success',
                'message': "Debit note saved successfully!",
                'note_id': note.id,
                'redirect_url': str(self.success_url)
            }, encoder=DjangoJSONEncoder)
        return response

    def form_invalid(self, form):
        if self.is_ajax():
            errors = [f"{field.capitalize()}: {errs[0]}" for field, errs in form.errors.items()]
            return JsonResponse({'success': False, 'status': 'error', 'message': ', '.join(errors)}, status=400, encoder=DjangoJSONEncoder)
        return super().form_invalid(form)

class DebitNoteDetailView(CompanyRequiredMixin, CompanyQuerySetMixin, DetailView):
    model = DebitNote
    template_name = 'company/debit_note_detail.html'
    context_object_name = 'note'

class ReportsHubView(CompanyRequiredMixin, CompanyQuerySetMixin, PaginationMixin, ListView):
    model = Invoice
    template_name = 'company/reports_hub.html'
    context_object_name = 'invoices'

    def get_filtered_invoices(self):
        qs = Invoice.objects.filter(company=self.request.user.company)
        
        # Apply search filter
        search = self.request.GET.get('search')
        if search:
            qs = qs.filter(Q(invoice_number__icontains=search) | Q(customer__name__icontains=search))

        # Apply status filter
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)

        # Apply date filters
        date_range = self.request.GET.get('date_range', 'this_month')
        today = timezone.localtime(timezone.now()).date()
        
        start_date = None
        end_date = None

        if date_range == 'today':
            start_date = today
            end_date = today
        elif date_range == 'yesterday':
            start_date = today - timedelta(days=1)
            end_date = today - timedelta(days=1)
        elif date_range == 'this_week':
            start_date = today - timedelta(days=today.weekday())
            end_date = today
        elif date_range == 'this_month':
            start_date = today.replace(day=1)
            end_date = today
        elif date_range == 'last_month':
            last_month_end = today.replace(day=1) - timedelta(days=1)
            start_date = last_month_end.replace(day=1)
            end_date = last_month_end
        elif date_range == 'this_year':
            start_date = today.replace(month=1, day=1)
            end_date = today
        elif date_range == 'custom':
            sd = self.request.GET.get('start_date')
            ed = self.request.GET.get('end_date')
            if sd:
                try:
                    start_date = datetime.strptime(sd, '%Y-%m-%d').date()
                except ValueError:
                    pass
            if ed:
                try:
                    end_date = datetime.strptime(ed, '%Y-%m-%d').date()
                except ValueError:
                    pass

        if start_date:
            qs = qs.filter(invoice_date__gte=start_date)
        if end_date:
            qs = qs.filter(invoice_date__lte=end_date)
            
        return qs

    def get_queryset(self):
        qs = self.get_filtered_invoices()
        view_type = self.request.GET.get('view_type', 'detailed')
        
        if view_type == 'day_wise':
            qs = qs.values('invoice_date').annotate(
                total_records=Count('id'),
                completed=Count('id', filter=Q(status='PAID')),
                pending=Count('id', filter=Q(status__in=['POSTED', 'PARTIALLY_PAID', 'DRAFT'])),
                cancelled=Count('id', filter=Q(status='CANCELLED')),
                total_amount=Sum('grand_total')
            )
            sort_by = self.request.GET.get('sort_by')
            if sort_by == 'amount_asc':
                qs = qs.order_by('total_amount')
            elif sort_by == 'amount_desc':
                qs = qs.order_by('-total_amount')
            elif sort_by == 'date_asc':
                qs = qs.order_by('invoice_date')
            else:
                qs = qs.order_by('-invoice_date')
                
        elif view_type == 'month_wise':
            from django.db.models.functions import TruncMonth
            qs = qs.annotate(month=TruncMonth('invoice_date')).values('month').annotate(
                total_records=Count('id'),
                completed=Count('id', filter=Q(status='PAID')),
                pending=Count('id', filter=Q(status__in=['POSTED', 'PARTIALLY_PAID', 'DRAFT'])),
                cancelled=Count('id', filter=Q(status='CANCELLED')),
                total_amount=Sum('grand_total')
            )
            sort_by = self.request.GET.get('sort_by')
            if sort_by == 'amount_asc':
                qs = qs.order_by('total_amount')
            elif sort_by == 'amount_desc':
                qs = qs.order_by('-total_amount')
            elif sort_by == 'date_asc':
                qs = qs.order_by('month')
            else:
                qs = qs.order_by('-month')
                
        else: # detailed
            sort_by = self.request.GET.get('sort_by')
            if sort_by == 'date_asc':
                qs = qs.order_by('invoice_date', 'id')
            elif sort_by == 'date_desc':
                qs = qs.order_by('-invoice_date', '-id')
            elif sort_by == 'amount_asc':
                qs = qs.order_by('grand_total', 'id')
            elif sort_by == 'amount_desc':
                qs = qs.order_by('-grand_total', '-id')
            elif sort_by == 'name_asc':
                qs = qs.order_by('customer__name', 'id')
            elif sort_by == 'name_desc':
                qs = qs.order_by('-customer__name', '-id')
            elif sort_by == 'status_asc':
                qs = qs.order_by('status', 'id')
            elif sort_by == 'status_desc':
                qs = qs.order_by('-status', '-id')
            else:
                qs = qs.order_by('-invoice_date', '-id')
                
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        base_qs = self.get_filtered_invoices()
        
        total_records = base_qs.count()
        today = timezone.localtime(timezone.now()).date()
        
        today_records = base_qs.filter(invoice_date=today).count()
        this_month_records = base_qs.filter(invoice_date__year=today.year, invoice_date__month=today.month).count()
        this_year_records = base_qs.filter(invoice_date__year=today.year).count()
        
        total_amount = base_qs.aggregate(sum=Sum('grand_total'))['sum'] or Decimal('0.00')

        context['summary_metrics'] = {
            'total_records': total_records,
            'today_records': today_records,
            'this_month_records': this_month_records,
            'this_year_records': this_year_records,
            'total_amount': total_amount,
        }

        footer_totals = base_qs.aggregate(
            records=Count('id'),
            completed=Count('id', filter=Q(status='PAID')),
            pending=Count('id', filter=Q(status__in=['POSTED', 'PARTIALLY_PAID', 'DRAFT'])),
            cancelled=Count('id', filter=Q(status='CANCELLED')),
            amount=Sum('grand_total')
        )
        context['footer_totals'] = {
            'records': footer_totals['records'] or 0,
            'completed': footer_totals['completed'] or 0,
            'pending': footer_totals['pending'] or 0,
            'cancelled': footer_totals['cancelled'] or 0,
            'amount': footer_totals['amount'] or Decimal('0.00')
        }

        date_range = self.request.GET.get('date_range', 'this_month')
        context['date_range'] = date_range
        context['view_type'] = self.request.GET.get('view_type', 'detailed')
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        context['current_time'] = timezone.localtime(timezone.now())
        
        return context

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format in ['csv', 'excel', 'pdf']:
            return self.export_data(export_format)
            
        print_format = request.GET.get('print')
        if print_format == 'true':
            self.object_list = self.get_queryset()
            context = self.get_context_data()
            return render(request, 'company/reports_hub_print.html', context)
            
        return super().get(request, *args, **kwargs)

    def export_data(self, export_format):
        import csv
        import io
        from django.http import HttpResponse
        
        qs = self.get_queryset()
        view_type = self.request.GET.get('view_type', 'detailed')
        company = self.request.user.company
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Build table rows based on view_type
        if view_type == 'day_wise':
            headers = ['Date', 'Total Records', 'Completed', 'Pending', 'Cancelled', 'Total Amount (INR)']
            rows = []
            for row in qs:
                d_str = row['invoice_date'].strftime('%Y-%m-%d') if row.get('invoice_date') else '-'
                rows.append([
                    d_str,
                    row.get('total_records', 0),
                    row.get('completed', 0),
                    row.get('pending', 0),
                    row.get('cancelled', 0),
                    float(row.get('total_amount') or 0.00)
                ])
        elif view_type == 'month_wise':
            headers = ['Month', 'Total Records', 'Completed', 'Pending', 'Cancelled', 'Total Amount (INR)']
            rows = []
            for row in qs:
                m_str = row['month'].strftime('%B %Y') if row.get('month') else '-'
                rows.append([
                    m_str,
                    row.get('total_records', 0),
                    row.get('completed', 0),
                    row.get('pending', 0),
                    row.get('cancelled', 0),
                    float(row.get('total_amount') or 0.00)
                ])
        else: # detailed
            headers = ['Invoice Number', 'Date', 'Customer Name', 'Status', 'Taxable Value (INR)', 'GST Total (INR)', 'Grand Total (INR)']
            rows = []
            for inv in qs:
                gst_total = inv.cgst_total + inv.sgst_total + inv.igst_total
                rows.append([
                    inv.invoice_number,
                    inv.invoice_date.strftime('%Y-%m-%d') if inv.invoice_date else '-',
                    inv.customer.name,
                    inv.get_status_display(),
                    float(inv.taxable_value),
                    float(gst_total),
                    float(inv.grand_total)
                ])

        if export_format == 'excel':
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Report_{view_type}"

            # Add Company Header
            ws.append([f"{company.name} - Billing Report ({view_type.replace('_', ' ').title()})"])
            ws.append([f"Generated on: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"])
            ws.append([]) # Blank row

            # Style report title
            ws['A1'].font = Font(name="Arial", size=14, bold=True, color="1E293B")
            ws['A2'].font = Font(name="Arial", size=9, italic=True, color="64748B")

            # Table Header
            header_row_idx = 4
            ws.append(headers)
            header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )

            for cell in ws[header_row_idx]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Table Data
            for row in rows:
                ws.append(row)

            for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border
                    cell.font = Font(name="Arial", size=9)
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="Billing_Report_{view_type}_{timestamp}.xlsx"'
            return response

        elif export_format == 'pdf':
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),
                rightMargin=20,
                leftMargin=20,
                topMargin=25,
                bottomMargin=20
            )
            elements = []
            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'ReportTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1E293B'),
                spaceAfter=4
            )
            meta_style = ParagraphStyle(
                'ReportMeta',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#64748B'),
                spaceAfter=12
            )

            elements.append(Paragraph(f"<b>{company.name}</b> — Billing Report ({view_type.replace('_', ' ').title()})", title_style))
            elements.append(Paragraph(f"Generated on {datetime.now().strftime('%d %b %Y, %I:%M %p')} | Total Records: {len(rows)}", meta_style))

            table_data = [headers]
            for r in rows:
                formatted_row = []
                for val in r:
                    if isinstance(val, float):
                        formatted_row.append(f"{val:,.2f}")
                    else:
                        formatted_row.append(str(val))
                table_data.append(formatted_row)

            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#FFFFFF'), colors.HexColor('#F8FAFC')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ]))
            elements.append(table)
            doc.build(elements)

            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="Billing_Report_{view_type}_{timestamp}.pdf"'
            return response

        else: # csv
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="Billing_Report_{view_type}_{timestamp}.csv"'
            writer = csv.writer(response)
            writer.writerow(headers)
            for r in rows:
                writer.writerow(r)
            return response

class ProfitLossReportView(CompanyRequiredMixin, TemplateView):
    template_name = 'company/report_profit_loss.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.user.company
        
        # Revenue
        sales = Invoice.objects.filter(company=company, status__in=['POSTED', 'PARTIALLY_PAID', 'PAID']).aggregate(sum=Sum('taxable_value'))['sum'] or Decimal('0.00')
        
        # Cost of Goods Sold (simplified estimation as purchases)
        cogs = PurchaseBill.objects.filter(company=company, status__in=['POSTED', 'PARTIALLY_PAID', 'PAID']).aggregate(sum=Sum('taxable_value'))['sum'] or Decimal('0.00')
        
        # Expenses
        expenses = Expense.objects.filter(company=company).aggregate(sum=Sum('amount'))['sum'] or Decimal('0.00')
        
        context['sales_total'] = sales
        context['cogs_total'] = cogs
        context['expenses_total'] = expenses
        context['net_profit'] = sales - cogs - expenses
        
        return context


class HSNSACReportView(CompanyRequiredMixin, TemplateView):
    template_name = 'company/report_hsn_sac.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.user.company
        
        tx_type = self.request.GET.get('type', 'SALES')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        hsn_code = self.request.GET.get('hsn_code')
        
        qs = GSTTransaction.objects.filter(company=company, is_cancelled=False)
        if tx_type == 'PURCHASE':
            qs = qs.filter(transaction_type__in=['PURCHASE', 'DEBIT_NOTE'])
        else:
            qs = qs.filter(transaction_type__in=['SALES', 'CREDIT_NOTE'])
            
        if start_date:
            qs = qs.filter(date__gte=start_date)
        if end_date:
            qs = qs.filter(date__lte=end_date)
        if hsn_code:
            qs = qs.filter(hsn_sac_code__icontains=hsn_code)
            
        summary = qs.values('hsn_sac_code', 'uqc_unit', 'gst_rate').annotate(
            total_qty=Sum('quantity'),
            total_taxable=Sum('taxable_value'),
            total_cgst=Sum('cgst_amount'),
            total_sgst=Sum('sgst_amount'),
            total_igst=Sum('igst_amount'),
            total_cess=Sum('cess_amount')
        ).order_by('hsn_sac_code')
        
        context['hsn_summary'] = summary
        context['selected_type'] = tx_type
        context['start_date'] = start_date or ''
        context['end_date'] = end_date or ''
        context['hsn_code'] = hsn_code or ''
        return context


# --- PASSWORD RESET FLOWS ---

def mask_email(email):
    if not email or '@' not in email:
        return "user@example.com"
    parts = email.split('@')
    name = parts[0]
    domain = parts[1]
    if len(name) > 2:
        masked_name = name[0] + '*' * (len(name) - 2) + name[-1]
    else:
        masked_name = name[0] + '*'
    return f"{masked_name}@{domain}"


def forgot_password_view(request):
    from django.conf import settings
    from django.core.mail import send_mail
    from django.utils.html import strip_tags
    import random
    import hashlib

    # Rate limit check on forgot password submissions
    cooldown = request.session.get('otp_resend_cooldown', 0)
    now = int(timezone.now().timestamp())
    
    import sys
    is_testing = 'test' in sys.argv or (getattr(settings, 'EMAIL_BACKEND', '') == 'django.core.mail.backends.locmem.EmailBackend')
    
    if request.method == 'POST':
        if not is_testing and cooldown > now:
            messages.error(request, f"Please wait {cooldown - now} seconds before requesting a new OTP.")
            return render(request, 'auth/forgot_password.html')

        username_or_email = request.POST.get('username', '').strip()
        
        # Avoid user enumeration
        user = None
        if '@' in username_or_email:
            user = CustomUser.objects.filter(email__iexact=username_or_email).first()
        else:
            user = CustomUser.objects.filter(username__iexact=username_or_email).first()
            
        if user and user.email:
            otp = str(random.randint(100000, 999999))
            otp_hash = hashlib.sha256(otp.encode()).hexdigest()
            expiry = timezone.now() + timedelta(minutes=10)
            
            # Deactivate any previous OTPs
            PasswordResetOTP.objects.filter(user=user).update(is_verified=True)
            
            # Save OTP record
            PasswordResetOTP.objects.create(
                user=user,
                otp_hash=otp_hash,
                expires_at=expiry
            )
            
            # Send HTML email
            subject = "Your Greenbacks Billing Password Reset OTP"
            html_message = f"""
<div style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; max-width: 500px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 8px; overflow: hidden; background-color: #ffffff; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);">
    <div style="background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 100%); padding: 24px; text-align: center; color: #ffffff;">
        <h2 style="margin: 0; font-size: 20px; font-weight: 700; letter-spacing: 0.5px; text-transform: uppercase;">GREENBACKS BILLING</h2>
        <p style="margin: 4px 0 0 0; font-size: 14px; opacity: 0.9;">Password Reset</p>
    </div>
    <div style="padding: 32px 24px; color: #334155; line-height: 1.6;">
        <p style="margin-top: 0; font-size: 16px;">Hello,</p>
        <p style="font-size: 15px;">We received a request to reset your password.</p>
        <p style="font-size: 15px;">Your OTP is:</p>
        <div style="text-align: center; margin: 30px 0; background-color: #f1f5f9; padding: 16px; border-radius: 8px; border: 1px dashed #cbd5e1;">
            <span style="font-size: 32px; font-weight: 800; color: #1e3a8a; letter-spacing: 6px; display: inline-block;">{otp}</span>
        </div>
        <p style="font-size: 14px; color: #64748b;">This OTP will expire in 10 minutes.</p>
        <p style="font-size: 14px; color: #64748b;">If you did not request this password reset, you can safely ignore this email.</p>
    </div>
    <div style="background-color: #f8fafc; border-top: 1px solid #e2e8f0; padding: 16px; text-align: center; color: #94a3b8; font-size: 12px;">
        <p style="margin: 0; font-weight: 600;">Designed & Managed by</p>
        <p style="margin: 4px 0 0 0; color: #64748b; font-weight: 700;">Greenbacks Lexverse Pvt. Ltd.</p>
    </div>
</div>
            """
            plain_message = strip_tags(html_message)
            try:
                send_mail(
                    subject,
                    plain_message,
                    getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@gblbilling.com'),
                    [user.email],
                    html_message=html_message,
                    fail_silently=False,
                )
                request.session['reset_user_id'] = user.id
                request.session['reset_masked_email'] = mask_email(user.email)
            except Exception as e:
                print("Failed to send OTP email:", e)
                messages.error(request, "Unable to send email. Please check the email configuration and try again.")
                return render(request, 'auth/forgot_password.html')
        else:
            # Dummy verify state for enumeration protection
            request.session['reset_user_id'] = None
            dummy_email = username_or_email if '@' in username_or_email else "user@example.com"
            request.session['reset_masked_email'] = mask_email(dummy_email)
            
        request.session['otp_resend_cooldown'] = int(timezone.now().timestamp()) + 30
        messages.success(request, "OTP sent successfully. Please check your email.")
        return redirect('verify_otp')
        
    return render(request, 'auth/forgot_password.html')


def verify_otp_view(request):
    from django.conf import settings
    from django.core.mail import send_mail
    from django.utils.html import strip_tags
    import random
    import hashlib

    masked_email = request.session.get('reset_masked_email', 'your email')
    user_id = request.session.get('reset_user_id')
    cooldown = request.session.get('otp_resend_cooldown', 0)
    now = int(timezone.now().timestamp())
    cooldown_remaining = max(0, cooldown - now)
    
    if request.method == 'POST':
        # Handling resend OTP request
        if 'resend' in request.POST:
            if cooldown_remaining > 0:
                messages.error(request, f"Please wait {cooldown_remaining} seconds before resending.")
                return redirect('verify_otp')
                
            if user_id:
                user = CustomUser.objects.filter(id=user_id).first()
                if user and user.email:
                    otp = str(random.randint(100000, 999999))
                    otp_hash = hashlib.sha256(otp.encode()).hexdigest()
                    expiry = timezone.now() + timedelta(minutes=10)
                    
                    PasswordResetOTP.objects.filter(user=user).update(is_verified=True)
                    PasswordResetOTP.objects.create(
                        user=user,
                        otp_hash=otp_hash,
                        expires_at=expiry
                    )
                    
                    html_message = f"""
<div style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; max-width: 500px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 8px; overflow: hidden; background-color: #ffffff; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);">
    <div style="background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 100%); padding: 24px; text-align: center; color: #ffffff;">
        <h2 style="margin: 0; font-size: 20px; font-weight: 700; letter-spacing: 0.5px; text-transform: uppercase;">GREENBACKS BILLING</h2>
        <p style="margin: 4px 0 0 0; font-size: 14px; opacity: 0.9;">Password Reset</p>
    </div>
    <div style="padding: 32px 24px; color: #334155; line-height: 1.6;">
        <p style="margin-top: 0; font-size: 16px;">Hello,</p>
        <p style="font-size: 15px;">We received a request to reset your password.</p>
        <p style="font-size: 15px;">Your OTP is:</p>
        <div style="text-align: center; margin: 30px 0; background-color: #f1f5f9; padding: 16px; border-radius: 8px; border: 1px dashed #cbd5e1;">
            <span style="font-size: 32px; font-weight: 800; color: #1e3a8a; letter-spacing: 6px; display: inline-block;">{otp}</span>
        </div>
        <p style="font-size: 14px; color: #64748b;">This OTP will expire in 10 minutes.</p>
        <p style="font-size: 14px; color: #64748b;">If you did not request this password reset, you can safely ignore this email.</p>
    </div>
    <div style="background-color: #f8fafc; border-top: 1px solid #e2e8f0; padding: 16px; text-align: center; color: #94a3b8; font-size: 12px;">
        <p style="margin: 0; font-weight: 600;">Designed & Managed by</p>
        <p style="margin: 4px 0 0 0; color: #64748b; font-weight: 700;">Greenbacks Lexverse Pvt. Ltd.</p>
    </div>
</div>
                    """
                    plain_message = strip_tags(html_message)
                    try:
                        send_mail(
                            "Your Greenbacks Billing Password Reset OTP",
                            plain_message,
                            getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@gblbilling.com'),
                            [user.email],
                            html_message=html_message,
                            fail_silently=False,
                        )
                        request.session['otp_resend_cooldown'] = int(timezone.now().timestamp()) + 30
                        messages.success(request, "OTP resent successfully. Please check your email.")
                    except Exception as e:
                        print("Failed to send OTP email:", e)
                        messages.error(request, "Unable to send email. Please check the email configuration and try again.")
                        return redirect('verify_otp')
            else:
                # Fake success to prevent enumeration
                request.session['otp_resend_cooldown'] = int(timezone.now().timestamp()) + 30
                messages.success(request, "OTP resent successfully. Please check your email.")
            return redirect('verify_otp')

        otp_code = request.POST.get('otp', '').strip()
        if not otp_code:
            messages.error(request, "OTP is required.")
            return render(request, 'auth/verify_otp.html', {'masked_email': masked_email, 'cooldown': cooldown_remaining})
            
        if not user_id:
            messages.error(request, "Invalid OTP. Please enter the correct OTP.")
            return render(request, 'auth/verify_otp.html', {'masked_email': masked_email, 'cooldown': cooldown_remaining})

        user = CustomUser.objects.filter(id=user_id).first()
        if not user:
            messages.error(request, "User not found.")
            return redirect('forgot_password')
            
        otp_hash = hashlib.sha256(otp_code.encode()).hexdigest()
        otp_record = PasswordResetOTP.objects.filter(user=user, is_verified=False).order_by('-created_at').first()
        
        if not otp_record:
            messages.error(request, "No active OTP request found. Please request a new OTP.")
            return redirect('forgot_password')
            
        if otp_record.is_expired():
            messages.error(request, "OTP expired. Please request a new OTP.")
            return render(request, 'auth/verify_otp.html', {'masked_email': masked_email, 'cooldown': cooldown_remaining})
            
        if otp_record.attempts >= 3:
            messages.error(request, "Too many failed attempts. Please request a new OTP.")
            return redirect('forgot_password')
            
        if otp_record.otp_hash != otp_hash:
            otp_record.attempts += 1
            otp_record.save()
            if otp_record.attempts >= 3:
                messages.error(request, "Too many failed attempts. Please request a new OTP.")
                return redirect('forgot_password')
            messages.error(request, f"Invalid OTP. Please enter the correct OTP. ({3 - otp_record.attempts} attempts remaining)")
            return render(request, 'auth/verify_otp.html', {'masked_email': masked_email, 'cooldown': cooldown_remaining})
            
        # OTP verified successfully
        otp_record.is_verified = True
        otp_record.save()
        
        request.session['otp_verified_user_id'] = user.id
        request.session['otp_verified_time'] = int(timezone.now().timestamp())
        
        messages.success(request, "OTP verified successfully. Please enter your new password.")
        return redirect('reset_password')
        
    return render(request, 'auth/verify_otp.html', {'masked_email': masked_email, 'cooldown': cooldown_remaining})


def reset_password_view(request):
    from django.conf import settings
    from django.core.mail import send_mail
    from django.utils.html import strip_tags

    verified_user_id = request.session.get('otp_verified_user_id')
    verified_time = request.session.get('otp_verified_time', 0)
    
    if not verified_user_id or int(timezone.now().timestamp()) - verified_time > 15 * 60:
        messages.error(request, "Session expired or invalid. Please start over.")
        return redirect('forgot_password')
        
    user = CustomUser.objects.filter(id=verified_user_id).first()
    if not user:
        messages.error(request, "User not found.")
        return redirect('forgot_password')
        
    if request.method == 'POST':
        new_password = request.POST.get('new_password', '')
        confirm_password = request.POST.get('confirm_password', '')
        
        if not new_password:
            messages.error(request, "Password is required.")
            return render(request, 'auth/reset_password.html')
            
        if len(new_password) < 8:
            messages.error(request, "Password must be at least 8 characters long.")
            return render(request, 'auth/reset_password.html')
            
        if new_password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return render(request, 'auth/reset_password.html')
            
        user.set_password(new_password)
        user.save()
        
        # Send confirmation email (No plaintext password)
        subject = "GREENBACKS BILLING - Password Successfully Changed"
        html_message = f"""
<div style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; max-width: 500px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 8px; overflow: hidden; background-color: #ffffff; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);">
    <div style="background: linear-gradient(135deg, #16a34a 0%, #4ade80 100%); padding: 24px; text-align: center; color: #ffffff;">
        <h2 style="margin: 0; font-size: 20px; font-weight: 700; letter-spacing: 0.5px; text-transform: uppercase;">GREENBACKS BILLING</h2>
        <p style="margin: 4px 0 0 0; font-size: 14px; opacity: 0.9;">Password Successfully Changed</p>
    </div>
    <div style="padding: 32px 24px; color: #334155; line-height: 1.6;">
        <p style="margin-top: 0; font-size: 16px;">Hello,</p>
        <p style="font-size: 15px;">Your password was successfully changed.</p>
        <p style="font-size: 15px;">If you made this change, no further action is required.</p>
        <p style="font-size: 15px; color: #ef4444; font-weight: 600; margin-top: 20px;">If you did not make this change, please contact your administrator immediately.</p>
    </div>
    <div style="background-color: #f8fafc; border-top: 1px solid #e2e8f0; padding: 16px; text-align: center; color: #94a3b8; font-size: 12px;">
        <p style="margin: 0; font-weight: 600;">Designed & Managed by</p>
        <p style="margin: 4px 0 0 0; color: #64748b; font-weight: 700;">Greenbacks Lexverse Pvt. Ltd.</p>
    </div>
</div>
        """
        plain_message = strip_tags(html_message)
        try:
            send_mail(
                subject,
                plain_message,
                getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@gblbilling.com'),
                [user.email],
                html_message=html_message,
                fail_silently=False
            )
        except Exception as e:
            print("Failed to send password confirmation email:", e)
            
        # Clean up session
        request.session.pop('otp_verified_user_id', None)
        request.session.pop('otp_verified_time', None)
        request.session.pop('reset_user_id', None)
        request.session.pop('reset_masked_email', None)
        
        log_action(user, 'RESET_PASSWORD', 'USER', user.id, request=request)
        return render(request, 'auth/reset_success.html')
        
    return render(request, 'auth/reset_password.html')


# --- BRANDING ASSETS API VIEWS ---

def api_company_branding_upload(request):
    if not request.user.is_authenticated or not request.user.company:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST method required.'}, status=405)

    asset_type = request.POST.get('type')
    if asset_type not in ['logo', 'signature', 'stamp']:
        return JsonResponse({'status': 'error', 'message': 'Invalid asset type.'}, status=400)

    if asset_type not in request.FILES:
        return JsonResponse({'status': 'error', 'message': 'Please select a file to upload.'}, status=400)

    upload_file = request.FILES[asset_type]

    if upload_file.size > 5 * 1024 * 1024:
        return JsonResponse({'status': 'error', 'message': 'File size must be less than 5MB.'}, status=400)

    ext = os.path.splitext(upload_file.name)[1].lower()
    if ext not in ['.jpg', '.jpeg', '.png', '.webp']:
        return JsonResponse({'status': 'error', 'message': 'Please select a valid image file (JPG, JPEG, PNG, or WEBP).'}, status=400)

    try:
        company = request.user.company
        
        # Delete existing file
        existing_file = getattr(company, asset_type)
        if existing_file:
            try:
                if existing_file.path and os.path.isfile(existing_file.path):
                    os.remove(existing_file.path)
            except Exception:
                pass
            try:
                existing_file.delete(save=False)
            except Exception:
                pass

        setattr(company, asset_type, upload_file)
        company.save()
        
        log_action(request.user, f'UPDATE_{asset_type.upper()}', 'COMPANY', company.id, request=request)

        get_url_fn = getattr(company, f'get_{asset_type}_url')
        get_name_fn = getattr(company, f'get_{asset_type}_filename')

        return JsonResponse({
            'status': 'success',
            'message': f'{asset_type.capitalize()} uploaded successfully.',
            'url': get_url_fn(),
            'filename': get_name_fn()
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'{asset_type.capitalize()} upload failed. Please try again.'}, status=400)


def api_company_branding_remove(request):
    if not request.user.is_authenticated or not request.user.company:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST method required.'}, status=405)

    try:
        data = json.loads(request.body)
    except Exception:
        data = request.POST

    asset_type = data.get('type')
    if asset_type not in ['logo', 'signature', 'stamp']:
        return JsonResponse({'status': 'error', 'message': 'Invalid asset type.'}, status=400)

    try:
        company = request.user.company
        existing_file = getattr(company, asset_type)
        if existing_file:
            try:
                if existing_file.path and os.path.isfile(existing_file.path):
                    os.remove(existing_file.path)
            except Exception:
                pass
            try:
                existing_file.delete(save=False)
            except Exception:
                pass

        setattr(company, asset_type, None)
        company.save()

        log_action(request.user, f'REMOVE_{asset_type.upper()}', 'COMPANY', company.id, request=request)

        return JsonResponse({
            'status': 'success',
            'message': f'{asset_type.capitalize()} removed successfully.'
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'{asset_type.capitalize()} removal failed. Please try again.'}, status=400)


# --- MY ACCOUNT & USER MANAGEMENT PANELS ---

class CompanyAccountsListView(CompanyRequiredMixin, PaginationMixin, ListView):
    model = CustomUser
    template_name = 'company/my_account.html'
    context_object_name = 'accounts'

    def dispatch(self, request, *args, **kwargs):
        # Allow only ADMIN and ACCOUNTANT roles
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if request.user.role not in ['ADMIN', 'ACCOUNTANT']:
            raise PermissionDenied("Unauthorized panel access.")
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        # Strict tenant data isolation
        qs = CustomUser.objects.filter(company=self.request.user.company, is_active=True).order_by('-id')
        
        # Debounced search
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(first_name__icontains=search) |
                Q(username__icontains=search) |
                Q(email__icontains=search) |
                Q(mobile__icontains=search)
            )
            
        # Role filters
        role = self.request.GET.get('role')
        if role:
            qs = qs.filter(role=role)
            
        # Status filters
        status = self.request.GET.get('status')
        if status == 'active':
            qs = qs.filter(is_active=True)
        elif status == 'inactive':
            qs = qs.filter(is_active=False)
            
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['roles'] = [('ADMIN', 'Company Admin'), ('ACCOUNTANT', 'Accountant')]
        return context


@transaction.atomic
def company_user_create_api(request):
    if not request.user.is_authenticated or request.user.company is None:
        return JsonResponse({'success': False, 'message': 'Authentication required.'}, status=403)
    
    if request.user.role != 'ADMIN':
        return JsonResponse({'success': False, 'message': 'Only Company Admins can add accounts.'}, status=403)
        
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST method required.'}, status=405)
        
    try:
        data = json.loads(request.body) if request.body and request.content_type == 'application/json' else request.POST
        name = data.get('name', '').strip()
        username = data.get('username', '').strip()
        email = data.get('email', '').strip()
        mobile = data.get('mobile', '').strip()
        role = data.get('role', '').strip()
        password = data.get('password', '').strip()
        confirm_password = data.get('confirm_password', '').strip()
        status_val = data.get('status', 'active')
        
        if not name or not username or not email or not role or not password or not confirm_password:
            return JsonResponse({'success': False, 'message': 'All fields marked with * are required.'}, status=400)
            
        if password != confirm_password:
            return JsonResponse({'success': False, 'message': 'Passwords do not match.'}, status=400)
            
        if len(password) < 6:
            return JsonResponse({'success': False, 'message': 'Password must be at least 6 characters long.'}, status=400)
            
        if role not in ['ADMIN', 'ACCOUNTANT']:
            return JsonResponse({'success': False, 'message': 'Invalid role assignment.'}, status=400)
            
        if CustomUser.objects.filter(username__iexact=username).exists():
            return JsonResponse({'success': False, 'message': f"Username '{username}' is already taken."}, status=400)
            
        if CustomUser.objects.filter(email__iexact=email).exists():
            return JsonResponse({'success': False, 'message': f"Email '{email}' is already in use."}, status=400)

        # Create user belonging to current company context
        user = CustomUser.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=name,
            role=role,
            mobile=mobile,
            company=request.user.company,
            is_active=(status_val == 'active')
        )
        
        log_action(request.user, 'CREATE_USER', 'USER', user.id, new_values={'username': username, 'role': role}, request=request)
        
        return JsonResponse({
            'success': True,
            'status': 'success',
            'message': 'Account created successfully.',
            'user': {
                'id': user.id,
                'name': user.first_name,
                'username': user.username,
                'email': user.email,
                'mobile': user.mobile or '',
                'role': user.get_role_display(),
                'status': 'Active' if user.is_active else 'Inactive',
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


def company_user_detail_api(request, pk):
    if not request.user.is_authenticated or request.user.company is None:
        return JsonResponse({'success': False, 'message': 'Authentication required.'}, status=403)
        
    company = request.user.company
    target_user = get_object_or_404(CustomUser, id=pk, company=company)
    
    # Calculate stats using AuditLog history
    quotations_created = AuditLog.objects.filter(user=target_user, action='CREATE_QUOTATION', module='QUOTATION').count()
    invoices_created = AuditLog.objects.filter(user=target_user, action='POST_INVOICE', module='INVOICE').count()
    payments_created = AuditLog.objects.filter(user=target_user, action__contains='PAYMENT').count()
    purchases_created = AuditLog.objects.filter(user=target_user, action__contains='PURCHASE').count()
    
    return JsonResponse({
        'success': True,
        'user': {
            'id': target_user.id,
            'name': target_user.first_name,
            'username': target_user.username,
            'email': target_user.email,
            'mobile': target_user.mobile or '',
            'role': target_user.role,
            'role_display': target_user.get_role_display(),
            'status': 'Active' if target_user.is_active else 'Inactive',
            'company': company.name,
            'created_at': target_user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if target_user.date_joined else '',
            'last_login': target_user.last_login.strftime('%Y-%m-%d %H:%M:%S') if target_user.last_login else 'Never',
            'activity': {
                'quotations_created': quotations_created,
                'invoices_created': invoices_created,
                'payments_created': payments_created,
                'purchases_created': purchases_created
            }
        }
    })


@transaction.atomic
def company_user_edit_api(request, pk):
    if not request.user.is_authenticated or request.user.company is None:
        return JsonResponse({'success': False, 'message': 'Authentication required.'}, status=403)
        
    if request.user.role != 'ADMIN':
        return JsonResponse({'success': False, 'message': 'Only Company Admins can edit accounts.'}, status=403)
        
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST method required.'}, status=405)
        
    company = request.user.company
    target_user = get_object_or_404(CustomUser, id=pk, company=company)
    
    try:
        data = json.loads(request.body) if request.body and request.content_type == 'application/json' else request.POST
        name = data.get('name', '').strip()
        username = data.get('username', '').strip()
        email = data.get('email', '').strip()
        mobile = data.get('mobile', '').strip()
        role = data.get('role', '').strip()
        status_val = data.get('status', 'active')
        
        if not name or not username or not email or not role:
            return JsonResponse({'success': False, 'message': 'Name, Username, Email and Role are required.'}, status=400)
            
        if role not in ['ADMIN', 'ACCOUNTANT']:
            return JsonResponse({'success': False, 'message': 'Invalid role assignment.'}, status=400)
            
        if CustomUser.objects.filter(username__iexact=username).exclude(id=pk).exists():
            return JsonResponse({'success': False, 'message': f"Username '{username}' is already taken."}, status=400)
            
        if CustomUser.objects.filter(email__iexact=email).exclude(id=pk).exists():
            return JsonResponse({'success': False, 'message': f"Email '{email}' is already in use."}, status=400)

        # Deactivation / Status validation guard
        if status_val == 'inactive' and target_user.is_active:
            if target_user.id == request.user.id:
                return JsonResponse({'success': False, 'message': 'You cannot deactivate your own active account.'}, status=400)
            if target_user.role == 'ADMIN':
                admin_count = CustomUser.objects.filter(company=company, role='ADMIN', is_active=True).count()
                if admin_count <= 1:
                    return JsonResponse({'success': False, 'message': 'You cannot deactivate the last active Company Admin.'}, status=400)

        old_values = {
            'first_name': target_user.first_name,
            'username': target_user.username,
            'email': target_user.email,
            'mobile': target_user.mobile,
            'role': target_user.role,
            'is_active': target_user.is_active
        }
        
        target_user.first_name = name
        target_user.username = username
        target_user.email = email
        target_user.mobile = mobile
        target_user.role = role
        target_user.is_active = (status_val == 'active')
        target_user.save()
        
        log_action(request.user, 'EDIT_USER', 'USER', target_user.id, old_values=old_values, new_values=data, request=request)
        
        return JsonResponse({
            'success': True,
            'message': 'Account updated successfully.',
            'user': {
                'id': target_user.id,
                'name': target_user.first_name,
                'username': target_user.username,
                'email': target_user.email,
                'mobile': target_user.mobile or '',
                'role': target_user.get_role_display(),
                'status': 'Active' if target_user.is_active else 'Inactive',
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@transaction.atomic
def company_user_change_password_api(request, pk):
    if not request.user.is_authenticated or request.user.company is None:
        return JsonResponse({'success': False, 'message': 'Authentication required.'}, status=403)
        
    if request.user.role != 'ADMIN':
        return JsonResponse({'success': False, 'message': 'Only Company Admins can manage passwords.'}, status=403)
        
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST method required.'}, status=405)
        
    company = request.user.company
    target_user = get_object_or_404(CustomUser, id=pk, company=company)
    
    try:
        data = json.loads(request.body) if request.body and request.content_type == 'application/json' else request.POST
        current_password = data.get('current_password', '').strip()
        new_password = data.get('new_password', '').strip()
        confirm_password = data.get('confirm_password', '').strip()
        
        if not current_password or not new_password or not confirm_password:
            return JsonResponse({'success': False, 'message': 'All password fields are required.'}, status=400)
            
        # Security verification: Verify current password of performing user
        if not request.user.check_password(current_password):
            return JsonResponse({'success': False, 'message': 'Incorrect current verification password.'}, status=400)
            
        if new_password != confirm_password:
            return JsonResponse({'success': False, 'message': 'New passwords do not match.'}, status=400)
            
        if len(new_password) < 6:
            return JsonResponse({'success': False, 'message': 'New password must be at least 6 characters long.'}, status=400)
            
        target_user.set_password(new_password)
        target_user.save()
        
        log_action(request.user, 'CHANGE_PASSWORD', 'USER', target_user.id, request=request)
        
        return JsonResponse({'success': True, 'message': 'Password changed successfully.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


def fetch_pdf_resources(uri, rel):
    import os
    from django.conf import settings

    if uri.startswith(('http://', 'https://')):
        return uri

    clean_uri = uri.split('?')[0]

    if clean_uri.startswith(settings.MEDIA_URL):
        path = os.path.join(settings.MEDIA_ROOT, clean_uri.replace(settings.MEDIA_URL, "", 1))
    elif clean_uri.startswith(settings.STATIC_URL):
        path = os.path.join(settings.STATIC_ROOT, clean_uri.replace(settings.STATIC_URL, "", 1))
    else:
        path = clean_uri

    if not os.path.isfile(path):
        alt_path = os.path.join(settings.BASE_DIR, clean_uri.lstrip('/'))
        if os.path.isfile(alt_path):
            return alt_path

    return path


def html_to_pdf_bytes(html_content):
    import io
    from xhtml2pdf import pisa
    result = io.BytesIO()
    pdf = pisa.pisaDocument(
        io.BytesIO(html_content.encode("UTF-8")),
        result,
        link_callback=fetch_pdf_resources
    )
    if not pdf.err:
        return result.getvalue()
    return None


def company_quotation_send_email(request, pk):
    if not request.user.is_authenticated or getattr(request.user, 'company', None) is None:
        return JsonResponse({'success': False, 'message': 'Authentication required.'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST method required.'}, status=405)

    company = request.user.company
    quotation = get_object_or_404(Quotation, id=pk, company=company)

    try:
        data = json.loads(request.body) if request.body and request.content_type == 'application/json' else request.POST
        recipient_email = data.get('email', '').strip()
    except Exception:
        return JsonResponse({'success': False, 'message': 'Invalid input data.'}, status=400)

    import re
    email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not recipient_email or not re.match(email_regex, recipient_email):
        return JsonResponse({'success': False, 'message': 'Please enter a valid email address.'}, status=400)

    from django.template.loader import render_to_string
    context = build_quotation_context(quotation)
    html_content = render_to_string('company/quotation_pdf.html', context, request=request)
    
    pdf_data = html_to_pdf_bytes(html_content)
    if not pdf_data or len(pdf_data) == 0:
        return JsonResponse({'success': False, 'message': 'Failed to generate PDF document.'}, status=500)

    from django.core.mail import EmailMessage
    from django.conf import settings
    
    customer_name = quotation.customer.name if quotation.customer else "Valued Customer"
    subject = f"Quotation {quotation.quotation_number} from {company.name}"
    date_str = quotation.date.strftime('%d %b %Y') if hasattr(quotation.date, 'strftime') else str(quotation.date)
    body = (
        f"Dear {customer_name},\n\n"
        f"Please find attached quotation {quotation.quotation_number} from {company.name}.\n\n"
        f"Quotation Date: {date_str}\n"
        f"Total Amount: ₹{quotation.grand_total}\n\n"
        f"Please review the attached quotation.\n\n"
        f"Regards,\n"
        f"{company.name}"
    )
    
    try:
        from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', None) or getattr(settings, 'EMAIL_HOST_USER', 'noreply@gblbilling.com')
        email = EmailMessage(
            subject,
            body,
            from_email,
            [recipient_email]
        )
        safe_filename = f"Quotation-{quotation.quotation_number}.pdf".replace(' ', '_').replace('/', '_')
        email.attach(safe_filename, pdf_data, 'application/pdf')
        email.send()
        
        log_action(request.user, 'SEND_EMAIL_QUOTATION', 'QUOTATION', quotation.id, new_values={'recipient': recipient_email}, request=request)
        return JsonResponse({'success': True, 'message': f"Quotation PDF sent successfully to {recipient_email}."})
    except Exception as e:
        import logging
        logging.getLogger('django').error(f"SMTP Error sending quotation {quotation.quotation_number}: {str(e)}")
        return JsonResponse({'success': False, 'message': 'Unable to send email. Please check your email configuration and try again.'}, status=500)


def company_sales_order_send_email(request, pk):
    if not request.user.is_authenticated or getattr(request.user, 'company', None) is None:
        return JsonResponse({'success': False, 'message': 'Authentication required.'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST method required.'}, status=405)

    company = request.user.company
    order = get_object_or_404(SalesOrder, id=pk, company=company)

    try:
        data = json.loads(request.body) if request.body and request.content_type == 'application/json' else request.POST
        recipient_email = data.get('email', '').strip()
    except Exception:
        return JsonResponse({'success': False, 'message': 'Invalid input data.'}, status=400)

    import re
    email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not recipient_email or not re.match(email_regex, recipient_email):
        return JsonResponse({'success': False, 'message': 'Please enter a valid email address.'}, status=400)

    from django.template.loader import render_to_string
    from .utils import build_hsn_sac_tax_summary
    comp_code = str(order.company.state_code or '').strip().zfill(2)
    pos_code = str(getattr(order.customer, 'billing_state_code', comp_code) or comp_code).strip().zfill(2)
    summary_list, total_qty = build_hsn_sac_tax_summary(order.items.all(), comp_code, pos_code)
    context = {
        'order': order,
        'company': company,
        'customer': order.customer,
        'is_interstate': (comp_code != pos_code),
        'hsn_summary': summary_list,
        'total_quantity': total_qty,
    }
    html_content = render_to_string('company/sales_order_pdf.html', context, request=request)
    
    pdf_data = html_to_pdf_bytes(html_content)
    if not pdf_data or len(pdf_data) == 0:
        return JsonResponse({'success': False, 'message': 'Failed to generate PDF document.'}, status=500)

    from django.core.mail import EmailMessage
    from django.conf import settings
    
    customer_name = order.customer.name if order.customer else "Valued Customer"
    subject = f"Sales Order {order.order_number} from {company.name}"
    date_str = order.order_date.strftime('%d %b %Y') if hasattr(order.order_date, 'strftime') else str(order.order_date)
    body = (
        f"Dear {customer_name},\n\n"
        f"Please find attached sales order {order.order_number} from {company.name}.\n\n"
        f"Order Date: {date_str}\n"
        f"Total Amount: ₹{order.grand_total}\n\n"
        f"Please review the attached sales order.\n\n"
        f"Regards,\n"
        f"{company.name}"
    )
    
    try:
        from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', None) or getattr(settings, 'EMAIL_HOST_USER', 'noreply@gblbilling.com')
        email = EmailMessage(
            subject,
            body,
            from_email,
            [recipient_email]
        )
        safe_filename = f"SalesOrder-{order.order_number}.pdf".replace(' ', '_').replace('/', '_')
        email.attach(safe_filename, pdf_data, 'application/pdf')
        email.send()
        
        log_action(request.user, 'SEND_EMAIL_SALES_ORDER', 'SALES_ORDER', order.id, new_values={'recipient': recipient_email}, request=request)
        return JsonResponse({'success': True, 'message': f"Sales Order PDF sent successfully to {recipient_email}."})
    except Exception as e:
        import logging
        logging.getLogger('django').error(f"SMTP Error sending sales order {order.order_number}: {str(e)}")
        return JsonResponse({'success': False, 'message': 'Unable to send email. Please check your email configuration and try again.'}, status=500)


def company_invoice_send_email(request, pk):
    if not request.user.is_authenticated or getattr(request.user, 'company', None) is None:
        return JsonResponse({'success': False, 'message': 'Authentication required.'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST method required.'}, status=405)

    company = request.user.company
    invoice = get_object_or_404(Invoice, id=pk, company=company)

    try:
        data = json.loads(request.body) if request.body and request.content_type == 'application/json' else request.POST
        recipient_email = data.get('email', '').strip()
    except Exception:
        return JsonResponse({'success': False, 'message': 'Invalid input data.'}, status=400)

    import re
    email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not recipient_email or not re.match(email_regex, recipient_email):
        return JsonResponse({'success': False, 'message': 'Please enter a valid email address.'}, status=400)

    from django.template.loader import render_to_string
    context = {'invoice': invoice}
    qr_string = generate_upi_qr_string(
        invoice.company.upi_id, invoice.company.trade_name or invoice.company.name,
        invoice.outstanding_amount(), invoice.invoice_number
    )
    context['upi_qr_string'] = qr_string
    context = add_invoice_hsn_summary_to_context(invoice, context)

    html_content = render_to_string('company/invoice_pdf.html', context, request=request)
    
    pdf_data = html_to_pdf_bytes(html_content)
    if not pdf_data or len(pdf_data) == 0:
        return JsonResponse({'success': False, 'message': 'Failed to generate PDF document.'}, status=500)

    from django.core.mail import EmailMessage
    from django.conf import settings
    
    customer_name = invoice.customer.name if invoice.customer else "Valued Customer"
    subject = f"Invoice {invoice.invoice_number} from {company.name}"
    date_str = invoice.invoice_date.strftime('%d %b %Y') if hasattr(invoice.invoice_date, 'strftime') else str(invoice.invoice_date)
    body = (
        f"Dear {customer_name},\n\n"
        f"Please find attached tax invoice {invoice.invoice_number} from {company.name}.\n\n"
        f"Invoice Date: {date_str}\n"
        f"Total Amount: ₹{invoice.grand_total}\n\n"
        f"Please review the attached invoice.\n\n"
        f"Regards,\n"
        f"{company.name}"
    )
    
    try:
        from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', None) or getattr(settings, 'EMAIL_HOST_USER', 'noreply@gblbilling.com')
        email = EmailMessage(
            subject,
            body,
            from_email,
            [recipient_email]
        )
        safe_filename = f"Invoice-{invoice.invoice_number}.pdf".replace(' ', '_').replace('/', '_')
        email.attach(safe_filename, pdf_data, 'application/pdf')
        email.send()
        
        log_action(request.user, 'SEND_EMAIL_INVOICE', 'INVOICE', invoice.id, new_values={'recipient': recipient_email}, request=request)
        return JsonResponse({'success': True, 'message': f"Invoice PDF sent successfully to {recipient_email}."})
    except Exception as e:
        import logging
        logging.getLogger('django').error(f"SMTP Error sending invoice {invoice.invoice_number}: {str(e)}")
        return JsonResponse({'success': False, 'message': 'Unable to send email. Please check your email configuration and try again.'}, status=500)


def company_credit_note_send_email(request, pk):
    if not request.user.is_authenticated or getattr(request.user, 'company', None) is None:
        return JsonResponse({'success': False, 'message': 'Authentication required.'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST method required.'}, status=405)

    company = request.user.company
    note = get_object_or_404(CreditNote, id=pk, company=company)

    try:
        data = json.loads(request.body) if request.body and request.content_type == 'application/json' else request.POST
        recipient_email = data.get('email', '').strip()
    except Exception:
        return JsonResponse({'success': False, 'message': 'Invalid input data.'}, status=400)

    import re
    email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not recipient_email or not re.match(email_regex, recipient_email):
        return JsonResponse({'success': False, 'message': 'Please enter a valid email address.'}, status=400)

    from django.template.loader import render_to_string
    from .utils import build_hsn_sac_tax_summary
    company_state_code = str(note.company.state_code or '').strip().zfill(2)
    pos_code = str(getattr(getattr(note, 'invoice', None), 'place_of_supply_code', getattr(getattr(getattr(note, 'invoice', None), 'customer', None), 'billing_state_code', company_state_code)) or company_state_code).strip().zfill(2)
    summary_list, total_qty = build_hsn_sac_tax_summary(note.items.all(), company_state_code, pos_code)
    context = {
        'note': note,
        'company': company,
        'is_interstate': (company_state_code != pos_code),
        'hsn_summary': summary_list,
        'total_quantity': total_qty,
    }
    html_content = render_to_string('company/credit_note_pdf.html', context, request=request)
    
    pdf_data = html_to_pdf_bytes(html_content)
    if not pdf_data or len(pdf_data) == 0:
        return JsonResponse({'success': False, 'message': 'Failed to generate PDF document.'}, status=500)

    from django.core.mail import EmailMessage
    from django.conf import settings
    
    customer_name = note.company.name
    subject = f"Credit Note {note.note_number} from {company.name}"
    date_str = note.note_date.strftime('%d %b %Y') if hasattr(note.note_date, 'strftime') else str(note.note_date)
    body = (
        f"Dear Customer,\n\n"
        f"Please find attached credit note {note.note_number} from {company.name}.\n\n"
        f"Note Date: {date_str}\n"
        f"Total Amount: ₹{note.grand_total}\n\n"
        f"Please review the attached credit note.\n\n"
        f"Regards,\n"
        f"{company.name}"
    )
    
    try:
        from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', None) or getattr(settings, 'EMAIL_HOST_USER', 'noreply@gblbilling.com')
        email = EmailMessage(
            subject,
            body,
            from_email,
            [recipient_email]
        )
        safe_filename = f"CreditNote-{note.note_number}.pdf".replace(' ', '_').replace('/', '_')
        email.attach(safe_filename, pdf_data, 'application/pdf')
        email.send()
        
        log_action(request.user, 'SEND_EMAIL_CREDIT_NOTE', 'CREDIT_NOTE', note.id, new_values={'recipient': recipient_email}, request=request)
        return JsonResponse({'success': True, 'message': f"Credit Note PDF sent successfully to {recipient_email}."})
    except Exception as e:
        import logging
        logging.getLogger('django').error(f"SMTP Error sending credit note {note.note_number}: {str(e)}")
        return JsonResponse({'success': False, 'message': 'Unable to send email. Please check your email configuration and try again.'}, status=500)


def api_product_detail(request, pk):
    if not request.user.is_authenticated or getattr(request.user, 'company', None) is None:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
    company = request.user.company
    prod = get_object_or_404(Product, id=pk, company=company)
    data = {
        'id': prod.id,
        'name': prod.name,
        'product_type': prod.get_product_type_display(),
        'sku': prod.sku or '-',
        'barcode': prod.barcode or '-',
        'category': prod.category.name if prod.category else '-',
        'brand': prod.brand.name if prod.brand else '-',
        'hsn_sac': prod.hsn_sac.code if prod.hsn_sac else '-',
        'gst_rate': str(prod.hsn_sac.gst_rate) if prod.hsn_sac else '0.00',
        'unit': prod.unit.name if prod.unit else 'PCS',
        'purchase_price': str(prod.purchase_price),
        'selling_price': str(prod.selling_price),
        'mrp': str(prod.mrp),
        'wholesale_price': str(prod.wholesale_price),
        'retail_price': str(prod.retail_price),
        'min_selling_price': str(prod.min_selling_price),
        'tax_inclusive': prod.tax_inclusive,
        'track_inventory': prod.track_inventory,
        'current_stock': str(prod.current_stock),
        'min_stock': str(prod.min_stock),
        'max_stock': str(prod.max_stock),
        'description': prod.description or '',
        'image_url': prod.get_image_url(),
        'is_active': prod.is_active,
        'created_at': prod.created_at.strftime('%d %b %Y %H:%M') if prod.created_at else '-'
    }
    return JsonResponse({'status': 'success', 'data': data})


def api_expense_detail(request, pk):
    if not request.user.is_authenticated or getattr(request.user, 'company', None) is None:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
    company = request.user.company
    exp = get_object_or_404(Expense, id=pk, company=company)
    total_amount = (exp.amount or Decimal('0.00')) + (exp.gst_amount or Decimal('0.00'))
    data = {
        'id': exp.id,
        'category_id': exp.category.id if exp.category else None,
        'category_name': exp.category.name if exp.category else '-',
        'vendor': exp.vendor or '-',
        'amount': str(exp.amount),
        'gst_amount': str(exp.gst_amount),
        'total_amount': str(total_amount),
        'payment_method': exp.payment_method or 'CASH',
        'reference_no': exp.reference_no or '-',
        'notes': exp.description or exp.notes if hasattr(exp, 'notes') else getattr(exp, 'description', '') or '',
        'date': exp.created_at.strftime('%Y-%m-%d') if hasattr(exp, 'created_at') and exp.created_at else '',
        'created_at': exp.created_at.strftime('%d %b %Y %H:%M') if hasattr(exp, 'created_at') and exp.created_at else '-'
    }
    return JsonResponse({'status': 'success', 'data': data})


def api_expense_edit(request, pk):
    if not request.user.is_authenticated or getattr(request.user, 'company', None) is None:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST method required.'}, status=405)
    
    company = request.user.company
    exp = get_object_or_404(Expense, id=pk, company=company)
    
    try:
        data = json.loads(request.body) if request.body and request.content_type == 'application/json' else request.POST
        if 'category_id' in data and data['category_id']:
            exp.category_id = int(data['category_id'])
        if 'vendor' in data:
            exp.vendor = str(data['vendor']).strip()
        if 'amount' in data:
            exp.amount = Decimal(str(data['amount']))
        if 'gst_amount' in data:
            exp.gst_amount = Decimal(str(data['gst_amount']))
        if 'payment_method' in data:
            exp.payment_method = str(data['payment_method'])
        if 'reference_no' in data:
            exp.reference_no = str(data['reference_no']).strip()
        if 'notes' in data:
            if hasattr(exp, 'description'):
                exp.description = str(data['notes'])
            if hasattr(exp, 'notes'):
                exp.notes = str(data['notes'])
        exp.save()
        return JsonResponse({'status': 'success', 'message': 'Expense updated successfully.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def api_payment_detail(request, pk):
    if not request.user.is_authenticated or getattr(request.user, 'company', None) is None:
        return JsonResponse({'status': 'error', 'message': 'Authentication required.'}, status=403)
    company = request.user.company
    pm = get_object_or_404(Payment, id=pk, company=company)
    
    party_name = pm.customer.name if pm.customer else (pm.supplier.name if pm.supplier else '-')
    ref_doc = '-'
    if pm.payment_type == 'RECEIPT' and pm.invoice:
        ref_doc = f"Invoice #{pm.invoice.invoice_number}"
    elif pm.payment_type == 'PAYMENT' and pm.purchase_bill:
        ref_doc = f"Purchase Bill #{pm.purchase_bill.supplier_bill_no}"
    else:
        ref_doc = "On Account"
        
    data = {
        'id': pm.id,
        'payment_date': pm.payment_date.strftime('%d %b %Y') if pm.payment_date else '-',
        'payment_type': pm.get_payment_type_display(),
        'party_name': party_name,
        'reference_doc': ref_doc,
        'amount': str(pm.amount),
        'payment_method': pm.get_payment_method_display(),
        'reference_no': pm.reference_no or '-',
        'notes': pm.notes or '-',
        'created_at': pm.created_at.strftime('%d %b %Y %H:%M') if pm.created_at else '-'
    }
    return JsonResponse({'status': 'success', 'data': data})


# --- APPLY GST & ADMIN APPLIED GST VIEWS ---

class ApplyGSTView(CompanyRequiredMixin, View):
    def get(self, request):
        company = request.user.company
        applications = GSTApplication.objects.filter(company=company).order_by('-created_at')
        initial_name = request.user.get_full_name() or request.user.first_name or ''
        initial_email = request.user.email or ''
        initial_phone = getattr(company, 'phone', '') or getattr(request.user, 'mobile', '') or ''
        
        return render(request, 'company/apply_gst.html', {
            'applications': applications,
            'initial_name': initial_name,
            'initial_email': initial_email,
            'initial_phone': initial_phone,
        })

    @transaction.atomic
    def post(self, request):
        company = request.user.company
        is_ajax = (
            request.headers.get('x-requested-with') == 'XMLHttpRequest' or
            'application/json' in request.META.get('HTTP_ACCEPT', '') or
            request.content_type == 'application/json'
        )

        try:
            if request.content_type == 'application/json' and request.body:
                data = json.loads(request.body)
            else:
                data = request.POST.dict()
        except Exception:
            data = request.POST.dict()

        full_name = (data.get('full_name') or '').strip()
        phone_number = (data.get('phone_number') or '').strip()
        email = (data.get('email') or '').strip()
        message = (data.get('message') or '').strip()

        if not full_name:
            if is_ajax:
                return JsonResponse({'success': False, 'status': 'error', 'message': 'Please enter your full name.'}, status=400)
            messages.error(request, 'Please enter your full name.')
            return redirect('apply_gst')

        if not phone_number:
            if is_ajax:
                return JsonResponse({'success': False, 'status': 'error', 'message': 'Please enter a valid phone number.'}, status=400)
            messages.error(request, 'Please enter a valid phone number.')
            return redirect('apply_gst')
        
        cleaned_phone = phone_number.replace('+', '').replace('-', '').replace(' ', '')
        if len(cleaned_phone) < 10 or not cleaned_phone.isdigit():
            if is_ajax:
                return JsonResponse({'success': False, 'status': 'error', 'message': 'Please enter a valid 10-digit phone number.'}, status=400)
            messages.error(request, 'Please enter a valid 10-digit phone number.')
            return redirect('apply_gst')

        if not email or '@' not in email or '.' not in email:
            if is_ajax:
                return JsonResponse({'success': False, 'status': 'error', 'message': 'Please enter a valid email address.'}, status=400)
            messages.error(request, 'Please enter a valid email address.')
            return redirect('apply_gst')

        app = GSTApplication.objects.create(
            company=company,
            user=request.user,
            full_name=full_name,
            phone_number=phone_number,
            email=email,
            message=message,
            status='Work Pending'
        )

        log_action(request.user, 'APPLY_GST', 'GSTApplication', app.id, request=request)

        # Send email notification to admin safely
        try:
            from django.core.mail import send_mail
            from django.utils.html import strip_tags

            admin_email = getattr(settings, 'BUSINESS_ADMIN_EMAIL', None) or getattr(settings, 'DEFAULT_FROM_EMAIL', None) or 'admin@example.com'
            company_name = company.name if company else 'N/A'
            created_str = app.created_at.strftime('%d %b %Y, %I:%M %p')

            subject = f"New GST Application Received - {app.full_name}"
            html_message = f"""
            <div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 8px; overflow: hidden; background: #ffffff;">
                <div style="background-color: #1e3a8a; color: #ffffff; padding: 20px; text-align: center;">
                    <h2 style="margin: 0; font-size: 20px;">New GST Application Received</h2>
                </div>
                <div style="padding: 24px; color: #334155; line-height: 1.6;">
                    <p style="font-size: 15px; margin-top: 0;">A new GST application has been submitted from the billing portal.</p>
                    <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
                        <tr><td style="padding: 8px 0; color: #64748b; font-weight: 600; width: 140px;">Applicant Name:</td><td style="padding: 8px 0; font-weight: 700; color: #0f172a;">{app.full_name}</td></tr>
                        <tr><td style="padding: 8px 0; color: #64748b; font-weight: 600;">Company:</td><td style="padding: 8px 0; font-weight: 700; color: #0f172a;">{company_name}</td></tr>
                        <tr><td style="padding: 8px 0; color: #64748b; font-weight: 600;">Phone Number:</td><td style="padding: 8px 0; font-weight: 700; color: #0f172a;">{app.phone_number}</td></tr>
                        <tr><td style="padding: 8px 0; color: #64748b; font-weight: 600;">Email Address:</td><td style="padding: 8px 0; font-weight: 700; color: #0f172a;">{app.email}</td></tr>
                        <tr><td style="padding: 8px 0; color: #64748b; font-weight: 600;">Status:</td><td style="padding: 8px 0;"><span style="background: #fef3c7; color: #b45309; padding: 3px 8px; border-radius: 4px; font-size: 12px; font-weight: 700;">Work Pending</span></td></tr>
                        <tr><td style="padding: 8px 0; color: #64748b; font-weight: 600;">Submitted At:</td><td style="padding: 8px 0; color: #475569;">{created_str}</td></tr>
                    </table>
                    <div style="background: #f8fafc; border: 1px solid #cbd5e1; border-radius: 6px; padding: 14px; margin-top: 15px;">
                        <strong style="color: #475569; display: block; margin-bottom: 4px;">Message:</strong>
                        <p style="margin: 0; color: #1e293b;">{app.message or 'No message provided.'}</p>
                    </div>
                </div>
            </div>
            """
            plain_message = strip_tags(html_message)
            send_mail(
                subject=subject,
                message=plain_message,
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@gblbilling.com'),
                recipient_list=[admin_email],
                html_message=html_message,
                fail_silently=True
            )
        except Exception as e:
            print("Failed to send admin email for GST application:", e)

        msg_text = "Your GST application has been submitted successfully. Our team will contact you shortly."
        if is_ajax:
            return JsonResponse({
                'success': True,
                'status': 'success',
                'message': msg_text,
                'application': {
                    'id': app.id,
                    'full_name': app.full_name,
                    'phone_number': app.phone_number,
                    'email': app.email,
                    'status': app.status,
                    'created_at': app.created_at.strftime('%d %b %Y, %I:%M %p')
                }
            })
        messages.success(request, msg_text)
        return redirect('apply_gst')


class AdminAppliedGSTListView(PaginationMixin, ListView):
    model = GSTApplication
    template_name = 'admin/applied_gst_list.html'
    context_object_name = 'applications'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or (request.user.role != 'SUPERADMIN' and not request.user.is_superuser):
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        qs = GSTApplication.objects.all().select_related('company', 'user')
        
        search = self.request.GET.get('search')
        if search:
            search = search.strip()
            qs = qs.filter(
                Q(full_name__icontains=search) |
                Q(phone_number__icontains=search) |
                Q(email__icontains=search) |
                Q(company__name__icontains=search) |
                Q(message__icontains=search)
            )

        status_filter = self.request.GET.get('status')
        if status_filter in ('Work Pending', 'Work Done'):
            qs = qs.filter(status=status_filter)

        date_filter = self.request.GET.get('date_filter')
        today = date.today()
        if date_filter == 'today':
            qs = qs.filter(created_at__date=today)
        elif date_filter == 'this_week':
            start_week = today - timedelta(days=today.weekday())
            qs = qs.filter(created_at__date__gte=start_week)
        elif date_filter == 'this_month':
            qs = qs.filter(created_at__year=today.year, created_at__month=today.month)
        elif date_filter == 'custom':
            date_from = self.request.GET.get('date_from')
            date_to = self.request.GET.get('date_to')
            if date_from:
                qs = qs.filter(created_at__date__gte=date_from)
            if date_to:
                qs = qs.filter(created_at__date__lte=date_to)

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_qs = GSTApplication.objects.all()
        context['total_count'] = all_qs.count()
        context['pending_count'] = all_qs.filter(status='Work Pending').count()
        context['done_count'] = all_qs.filter(status='Work Done').count()
        context['current_search'] = self.request.GET.get('search', '')
        context['current_status'] = self.request.GET.get('status', '')
        context['current_date_filter'] = self.request.GET.get('date_filter', '')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        return context


def admin_gst_application_detail_api(request, pk):
    if not request.user.is_authenticated or (request.user.role != 'SUPERADMIN' and not request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)

    app = get_object_or_404(GSTApplication.objects.select_related('company'), pk=pk)
    return JsonResponse({
        'success': True,
        'application': {
            'id': app.id,
            'full_name': app.full_name,
            'company_name': app.company.name if app.company else 'N/A',
            'phone_number': app.phone_number,
            'email': app.email,
            'message': app.message or 'N/A',
            'status': app.status,
            'admin_notes': app.admin_notes or '',
            'created_at': app.created_at.strftime('%d %b %Y, %I:%M %p'),
            'updated_at': app.updated_at.strftime('%d %b %Y, %I:%M %p')
        }
    })


@transaction.atomic
def admin_gst_application_edit_api(request, pk):
    if not request.user.is_authenticated or (request.user.role != 'SUPERADMIN' and not request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)

    app = get_object_or_404(GSTApplication, pk=pk)

    try:
        if request.content_type == 'application/json' and request.body:
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
    except Exception:
        data = request.POST.dict()

    full_name = (data.get('full_name') or app.full_name).strip()
    phone_number = (data.get('phone_number') or app.phone_number).strip()
    email = (data.get('email') or app.email).strip()
    message = data.get('message', app.message)
    new_status = data.get('status', app.status)
    admin_notes = data.get('admin_notes', app.admin_notes)

    if new_status not in ('Work Pending', 'Work Done'):
        return JsonResponse({'success': False, 'message': 'Invalid status choice.'}, status=400)

    old_status = app.status

    app.full_name = full_name
    app.phone_number = phone_number
    app.email = email
    app.message = message
    app.status = new_status
    app.admin_notes = admin_notes
    app.save()

    # Send status update email if status changed to Work Done
    if old_status != 'Work Done' and new_status == 'Work Done':
        try:
            from django.core.mail import send_mail
            from django.utils.html import strip_tags

            subject = "GST Application Status Updated"
            html_message = f"""
            <div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 8px; overflow: hidden; background: #ffffff;">
                <div style="background-color: #16a34a; color: #ffffff; padding: 20px; text-align: center;">
                    <h2 style="margin: 0; font-size: 20px;">GST Application Status Updated</h2>
                </div>
                <div style="padding: 24px; color: #334155; line-height: 1.6;">
                    <p style="font-size: 15px; margin-top: 0;">Hello <strong>{app.full_name}</strong>,</p>
                    <p style="font-size: 15px;">Your GST application status has been updated.</p>
                    <div style="background: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 6px; padding: 16px; margin: 20px 0; text-align: center;">
                        <span style="font-size: 14px; color: #166534; font-weight: 600; display: block; margin-bottom: 4px;">Current Status:</span>
                        <strong style="font-size: 18px; color: #15803d;">Work Done</strong>
                    </div>
                    <p style="font-size: 14px; color: #64748b;">Thank you.<br><strong>Billing Software Team</strong></p>
                </div>
            </div>
            """
            plain_message = strip_tags(html_message)
            send_mail(
                subject=subject,
                message=plain_message,
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@gblbilling.com'),
                recipient_list=[app.email],
                html_message=html_message,
                fail_silently=True
            )
        except Exception as e:
            print("Failed to send status update email:", e)

    return JsonResponse({
        'success': True,
        'message': 'GST application updated successfully.',
        'application': {
            'id': app.id,
            'full_name': app.full_name,
            'company_name': app.company.name if app.company else 'N/A',
            'phone_number': app.phone_number,
            'email': app.email,
            'status': app.status,
            'updated_at': app.updated_at.strftime('%d %b %Y, %I:%M %p')
        }
    })




